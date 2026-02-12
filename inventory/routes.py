from __future__ import annotations
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy import func, or_, and_
from models import IssuedPartRecord, WorkOrder, WorkOrderPart, TechReceiveLog, IssuedBatch, Part, ReceivingBatch, \
    ReceivingItem, OrderItem, IssuedBatch
from services.receiving import unpost_receiving_batch
import json
from services.receiving import post_receiving_batch
from flask import (
    Blueprint, render_template, request, redirect, url_for,
    flash, send_file, jsonify, after_this_request,
    current_app,abort, session,                   # NEW
)
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
from werkzeug.exceptions import abort
from werkzeug.security import generate_password_hash, check_password_hash
from urllib.parse import urlencode
import re,sqlite3,os

import pandas as pd
from io import BytesIO
from datetime import datetime, timedelta, time, date
from collections import defaultdict
from extensions import db
from utils.invoice_generator import generate_invoice_pdf
from models import User, ROLE_SUPERADMIN, ROLE_ADMIN, ROLE_USER, ROLE_VIEWER,ROLE_TECHNICIAN, Part, WorkOrder, WorkOrderPart
from reportlab.lib.pagesizes import letter, landscape
from compare_cart.run_compare import get_marcone_items, check_cart_items, export_to_docx
from compare_cart.run_compare_reliable import get_reliable_items
from .import_rules import load_table, normalize_table, build_receive_movements
from .import_ledger import has_key, add_key
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

UPLOAD_DIR = os.path.join(os.getcwd(), "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

inventory_bp = Blueprint('inventory', __name__)
EPORT_PATH = os.path.join(os.path.dirname(__file__), '..', 'marcone_inventory_report.docx')

DB_PATH = os.path.normpath(os.path.join(os.path.dirname(__file__), "..", "instance", "inventory.db"))

_MARKER_RX = re.compile(r'^\s*shipment\s*mark(?:ing)?\s*[:\-]?', re.I)


_PREFIX_RE = re.compile(r'^\s*(wo|job|pn|model|brand|serial)\s*:\s*(.+)\s*$', re.I)
_PN_TOKEN_RX = re.compile(r'^([A-Za-z0-9][A-Za-z0-9\-\/\._]*)\s+(.+)$')

_units_re = re.compile(r"^units\[(\d+)\]\[(brand|model|serial)\]$")
_rows_re  = re.compile(
    r"^units\[(\d+)\]\[rows\]\[(\d+)\]\[(part_number|part_name|quantity|alt_numbers|supplier|backorder_flag|line_status|unit_cost|unit_cost_base|location)\]$"
)
_rows_flat_re = re.compile(
    r"^rows\[(\d+)\]\[(part_number|part_name|quantity|alt_numbers|supplier|backorder_flag|line_status|unit_cost|unit_cost_base|location)\]$"
)

_PN_FIRST_TOKEN_RX = re.compile(r'^([A-Za-z0-9][A-Za-z0-9\-\/\._]*)\s+(.+)$')
_PN_SPLIT_RX  = re.compile(r'^([A-Za-z0-9][A-Za-z0-9\-\/\._]*)\s+(.+)$')
_TOTALS_RX    = re.compile(r'^\s*(order\s*total|orderinetot|ordertot|subtotal|total)\b', re.I)


# --- Invoice numbering baseline ---
INVOICE_START_AT = 140  # новые инвойсы начнутся с 000140

ALLOWED_PDF = {".pdf"}

# Ключ → дефолтная локация
SUPPLIER_LOC_DEFAULTS = {
    "reliable": "REL",
    "marcone":  "MAR",
    "marcon":   "MAR",
}

# inventory/routes.py  (добавь рядом с _create_batch_for_records)
# --- TECH ROLE HELPER (fallback, если нет security.py) ---
try:
    from security import is_technician  # основной путь
except Exception:
    from flask_login import current_user
    def is_technician() -> bool:
        return (getattr(current_user, "role", "") or "").strip().lower() == "technician"



def _query_technicians():
    """Return list of (id, username) for users with 'technician' role."""
    return (
        db.session.query(User.id, User.username)
        .filter(func.lower(func.trim(User.role)) == "technician")
        .order_by(func.lower(func.trim(User.username)).asc())
        .all()
    )

def _job_tokens_from_text(s: str) -> list[str]:
    raw = (s or "").upper()
    # берём числа (твоя canonical_job на этом же принципе)
    nums = re.findall(r"\d+", raw)
    # уникальные, сохраняем порядок
    out = []
    for n in nums:
        n = n.strip()
        if n and n not in out:
            out.append(n)
    return out

def _job_tokens(raw: str) -> list[str]:
    s = (raw or "").upper()
    tokens = re.findall(r"\d+[A-Za-z]?", s)  # 986238 or 986238A
    out = []
    for t in tokens:
        t = t.strip().upper()
        if t and t not in out:
            out.append(t)
    return out



# рядом с другими утилитами
def _is_superadmin_user(u) -> bool:
    role = (getattr(u, "role", "") or "").strip().lower()
    # поддерживаем все твои варианты флагов
    return (
        role in ("superadmin", "super admin")
        or bool(getattr(u, "is_superadmin", False))
        or bool(getattr(u, "is_super_admin", False))
    )


def _supplier_to_default_location(supplier_hint: str | None) -> str:
    """По имени/подсказке поставщика выбрать дефолтную локацию."""
    s = (supplier_hint or "").lower()
    for key, loc in SUPPLIER_LOC_DEFAULTS.items():
        if key in s:
            return loc
    return "MAIN"

def _reserve_invoice_number():
    """
    Берём максимальный invoice_number и в IssuedBatch, и в IssuedPartRecord
    и возвращаем следующий номер.
    """
    from models import IssuedBatch, IssuedPartRecord  # локальный импорт, чтобы не ловить циклы

    mb = (
        db.session.query(func.coalesce(func.max(IssuedBatch.invoice_number), 0))
        .scalar()
        or 0
    )
    ml = (
        db.session.query(func.coalesce(func.max(IssuedPartRecord.invoice_number), 0))
        .scalar()
        or 0
    )
    return max(int(mb), int(ml)) + 1


def _apply_default_location(obj, default_loc: str):
    """
    Проставить default_loc туда, где location пустая.
    Работает и с DataFrame, и со списком dict.
    """
    try:
        import pandas as pd  # локальный импорт — не мешаем остальному
        if hasattr(obj, "copy") and hasattr(obj, "columns"):  # вероятно, DataFrame
            df = obj.copy()
            if "location" not in df.columns:
                df["location"] = ""
            df["location"] = df["location"].fillna("").replace("", default_loc)
            return df
    except Exception:
        pass

    # список словарей
    if isinstance(obj, list):
        for r in obj:
            if not (r.get("location") or "").strip():
                r["location"] = default_loc
        return obj

    return obj

def _resolve_default_loc(df, default_loc: str | None, saved_path: str, supplier_hint: str | None = None) -> str:
    """
    Надёжно выбираем дефолтную локацию:
    1) Явный default_loc (если передали)
    2) supplier_hint (из контента / формы)
    3) df['supplier'] (если распознался текстом)
    4) имя файла (contains 'reliable'/'marcone' и т.п.)
    5) MAIN
    """
    # 1) уже переданный default_loc
    if default_loc:
        return str(default_loc).strip().upper()

    # 2) supplier_hint
    loc = _supplier_to_default_location(supplier_hint)
    if loc and loc != "MAIN":
        return loc

    # 3) колонка supplier в DF
    try:
        if df is not None and "supplier" in df.columns:
            vals = " ".join(str(x) for x in df["supplier"].dropna().astype(str).tolist()).lower()
            loc3 = _supplier_to_default_location(vals)
            if loc3 and loc3 != "MAIN":
                return loc3
    except Exception:
        pass

    # 4) имя файла
    base = (saved_path or "").lower()
    loc4 = _supplier_to_default_location(base)
    if loc4 and loc4 != "MAIN":
        return loc4

    # 5) запасной вариант
    return "MAIN"


def _clear_dedup_keys_for_batch(batch_id: int, supplier: str | None = None, invoice: str | None = None) -> int:
    """
    Удаляет записи анти-дедупа, связанные с batch_id / supplier / invoice.
    Возвращает количество удалённых записей.
    Стратегии:
      1) ImportDedupKey (если есть).
      2) inventory.dedup_store (если есть).
      3) Рефлексия БД по 'dedup' + 'import' + 'processed' + 'parse' + 'lock' + 'invoice' таблицам.
         Матч по:
           - batch_id/receiving_batch_id/goods_receipt_id
           - supplier/supplier_name/vendor/vendor_name
           - invoice/invoice_number/inv_no/inv_number/invoice_no/order_no
           - key/hash/fingerprint
           - source_file/filename/file_path
           - JSON-поля (meta_json/meta/data) по batch_id/supplier/invoice
    """
    deleted = 0
    sup = (supplier or "").strip()
    inv = (invoice  or "").strip()

    # --- Стратегия 1: ImportDedupKey (если модель существует) ---
    try:
        from models import ImportDedupKey
        from sqlalchemy import or_, cast, String, func

        q = ImportDedupKey.query
        or_conds = []

        # JSON-поля: batch_id, supplier, invoice
        for mc in ("meta_json", "meta", "data"):
            if hasattr(ImportDedupKey, mc):
                col = getattr(ImportDedupKey, mc)
                or_conds.append(cast(col, String).ilike(f'%\"batch_id\": {batch_id}%'))
                or_conds.append(cast(col, String).ilike(f'%\"batch_id\":\"{batch_id}\"%'))
                if sup:
                    or_conds.append(cast(col, String).ilike(f'%\"supplier\":\"{sup}\"%'))
                    or_conds.append(cast(col, String).ilike(f'%\"supplier_name\":\"{sup}\"%'))
                if inv:
                    or_conds.append(cast(col, String).ilike(f'%\"invoice\":\"{inv}\"%'))
                    or_conds.append(cast(col, String).ilike(f'%\"invoice_number\":\"{inv}\"%'))
                    or_conds.append(cast(col, String).ilike(f'%\"invoice_no\":\"{inv}\"%'))
                    or_conds.append(cast(col, String).ilike(f'%\"order_no\":\"{inv}\"%'))

        # key/hash
        for name in ("key", "hash", "fingerprint"):
            if hasattr(ImportDedupKey, name):
                col = getattr(ImportDedupKey, name)
                if batch_id:
                    or_conds.append(col.ilike(f"%|B{batch_id}"))
                if sup and inv:
                    or_conds.append(func.upper(col) == (sup + inv).upper())

        # supplier/invoice прямыми колонками
        if sup and hasattr(ImportDedupKey, "supplier"):
            or_conds.append(func.upper(getattr(ImportDedupKey, "supplier")) == sup.upper())
        if sup and hasattr(ImportDedupKey, "supplier_name"):
            or_conds.append(func.upper(getattr(ImportDedupKey, "supplier_name")) == sup.upper())
        for inv_col in ("invoice_number", "invoice", "invoice_no", "order_no"):
            if inv and hasattr(ImportDedupKey, inv_col):
                or_conds.append(func.upper(getattr(ImportDedupKey, inv_col)) == inv.upper())

        # файловые колонки
        for file_col in ("source_file", "filename", "file_path"):
            if hasattr(ImportDedupKey, file_col):
                col = getattr(ImportDedupKey, file_col)
                # если в meta нет - всё равно почистим по наличию файла, см. откуда вызывает парсер
                # оставляем без условия — зависит от твоего кейса

        if or_conds:
            rows = q.filter(or_(*or_conds)).all()
            for row in rows:
                db.session.delete(row)
                deleted += 1
            if rows:
                db.session.commit()
    except ImportError:
        pass
    except Exception:
        db.session.rollback()

    # --- Стратегия 2: dedup_store API (если есть) ---
    try:
        try:
            from inventory.dedup_store import iter_keys, del_key  # type: ignore
        except Exception:
            from inventory.import_ledger import iter_keys, del_key  # fallback

        for k, meta in iter_keys():
            try:
                meta = meta or {}
                bid = str(meta.get("batch_id", ""))
                if bid == str(batch_id) or f"|B{batch_id}" in str(k):
                    if sup and str(meta.get("supplier", "")).strip():
                        if str(meta.get("supplier")).strip().upper() != sup.upper():
                            continue
                    if inv and str(meta.get("invoice", "")).strip():
                        if str(meta.get("invoice")).strip().upper() != inv.upper():
                            continue
                    if del_key(k):
                        deleted += 1
            except Exception:
                continue
    except Exception:
        pass

    # 2b) del_keys_by_batch
    try:
        try:
            from inventory.dedup_store import del_keys_by_batch  # type: ignore
        except Exception:
            from inventory.import_ledger import del_keys_by_batch  # fallback
        deleted += int(del_keys_by_batch(batch_id) or 0)
    except Exception:
        pass

    # --- Стратегия 3: расширенная рефлексия БД ---
    try:
        import sqlalchemy as sa
        engine = db.session.get_bind()
        insp   = sa.inspect(engine)

        TABLE_NAME_HINTS = ("dedup", "import", "processed", "parse", "lock", "invoice")
        POSSIBLE_BATCH   = ("batch_id", "goods_receipt_id", "receiving_batch_id")
        POSSIBLE_SUP     = ("supplier", "supplier_name", "vendor", "vendor_name")
        POSSIBLE_INV     = ("invoice", "invoice_number", "inv_no", "inv_number", "invoice_no", "order_no")
        POSSIBLE_KEY     = ("key", "hash", "fingerprint")
        POSSIBLE_FILE    = ("source_file", "filename", "file_path")
        META_CANDS       = ("meta_json", "meta", "data")

        for tname in insp.get_table_names():
            lname = tname.lower()
            if not any(h in lname for h in TABLE_NAME_HINTS):
                continue

            meta = sa.MetaData()
            try:
                table = sa.Table(tname, meta, autoload_with=engine)
            except Exception:
                continue

            cols = {c.name.lower(): c for c in table.columns}

            col_batch = next((cols[n] for n in POSSIBLE_BATCH if n in cols), None)
            col_sup   = next((cols[n] for n in POSSIBLE_SUP   if n in cols), None)
            col_inv   = next((cols[n] for n in POSSIBLE_INV   if n in cols), None)
            col_key   = next((cols[n] for n in POSSIBLE_KEY   if n in cols), None)
            col_file  = next((cols[n] for n in POSSIBLE_FILE  if n in cols), None)

            or_conds = []
            if col_batch is not None:
                or_conds.append(col_batch == int(batch_id))
            if sup and col_sup is not None:
                or_conds.append(sa.func.upper(col_sup) == sup.upper())
            if inv and col_inv is not None:
                or_conds.append(sa.func.upper(col_inv) == inv.upper())
            if col_key is not None:
                # два популярных паттерна ключей
                if batch_id:
                    or_conds.append(sa.cast(col_key, String).ilike(f"%|B{batch_id}"))
                if sup and inv:
                    or_conds.append(sa.func.upper(col_key) == (sup + inv).upper())
            # JSON-поля
            for mc in META_CANDS:
                if mc in cols:
                    jcol = cols[mc]
                    or_conds.append(sa.cast(jcol, String).ilike(f'%\"batch_id\": {batch_id}%'))
                    or_conds.append(sa.cast(jcol, String).ilike(f'%\"batch_id\":\"{batch_id}\"%'))
                    if sup:
                        or_conds.append(sa.cast(jcol, String).ilike(f'%\"supplier\":\"{sup}\"%'))
                        or_conds.append(sa.cast(jcol, String).ilike(f'%\"supplier_name\":\"{sup}\"%'))
                    if inv:
                        for invk in ("invoice", "invoice_number", "invoice_no", "order_no"):
                            or_conds.append(sa.cast(jcol, String).ilike(f'%\"{invk}\":\"{inv}\"%'))

            if not or_conds:
                continue

            stmt = sa.delete(table).where(sa.or_(*or_conds))
            try:
                res = engine.execute(stmt)
                deleted += int(getattr(res, "rowcount", 0) or 0)
            except Exception:
                continue

        if deleted:
            db.session.commit()
    except Exception:
        pass

    return int(deleted or 0)

@inventory_bp.get("/api/job_reserve", endpoint="api_job_reserve")
@login_required
def api_job_reserve():
    from flask import request, jsonify
    from models import JobReservation, WorkOrder
    from extensions import db
    from datetime import timezone  # ✅ add

    role = (getattr(current_user, "role", "") or "").strip().lower()
    if role not in ("admin", "superadmin"):
        return jsonify({"ok": False, "error": "Access denied"}), 403

    job_numbers = (request.args.get("job_numbers") or "").strip()
    tokens = _job_tokens_from_text(job_numbers)
    if not tokens:
        return jsonify({"ok": True, "tokens": [], "status": "empty"})

    now = datetime.utcnow()
    ttl = timedelta(minutes=15)
    exp = now + ttl

    # ✅ helper: always return UTC with Z so frontend converts correctly
    def _iso_utc_z(dt):
        if not dt:
            return None
        return dt.replace(tzinfo=timezone.utc).isoformat().replace("+00:00", "Z")

    # clean expired
    JobReservation.query.filter(JobReservation.expires_at < now).delete(synchronize_session=False)

    # 1) if WO exists -> duplicate
    for t in tokens:
        wo = (
            WorkOrder.query
            .filter(WorkOrder.job_numbers.ilike(f"%{t}%"))
            .order_by(WorkOrder.id.desc())
            .first()
        )
        if wo:
            return jsonify({
                "ok": True,
                "status": "exists",
                "existing_id": wo.id,
                "token": t
            })

    blocked = []
    reserved = []
    for t in tokens:
        row = JobReservation.query.filter_by(job_token=t).first()
        if row and row.expires_at >= now:
            if row.holder_user_id == getattr(current_user, "id", None):
                row.expires_at = exp
                row.holder_username = getattr(current_user, "username", None)
                db.session.add(row)
                reserved.append(t)
            else:
                blocked.append({
                    "token": t,
                    "holder": row.holder_username or "unknown",
                    "expires_at": _iso_utc_z(row.expires_at)  # ✅ fixed
                })
        else:
            if not row:
                row = JobReservation(job_token=t)
            row.holder_user_id = getattr(current_user, "id", None)
            row.holder_username = getattr(current_user, "username", None)
            row.expires_at = exp
            db.session.add(row)
            reserved.append(t)

    db.session.commit()

    if blocked:
        return jsonify({
            "ok": True,
            "status": "locked",
            "reserved": reserved,
            "blocked": blocked,
            "ttl_seconds": 15 * 60
        })

    return jsonify({
        "ok": True,
        "status": "reserved",
        "reserved": reserved,
        "ttl_seconds": 15 * 60
    })

@inventory_bp.post("/api/job_release", endpoint="api_job_release")
@login_required
def api_job_release():
    from flask import request, jsonify
    from models import JobReservation
    from extensions import db

    role = (getattr(current_user, "role", "") or "").strip().lower()
    if role not in ("admin", "superadmin"):
        return jsonify({"ok": False, "error": "Access denied"}), 403

    job_numbers = (request.form.get("job_numbers") or "").strip()
    tokens = _job_tokens_from_text(job_numbers)
    if not tokens:
        return jsonify({"ok": True})

    uid = getattr(current_user, "id", None)
    for t in tokens:
        JobReservation.query.filter_by(job_token=t, holder_user_id=uid).delete(synchronize_session=False)

    db.session.commit()
    return jsonify({"ok": True})

@inventory_bp.post("/receiving/<int:batch_id>/delete", endpoint="receiving_delete")
@login_required
def receiving_delete(batch_id: int):
    from flask import current_app, flash, redirect, url_for
    from flask_login import current_user
    from extensions import db
    from sqlalchemy import func
    from models import ReceivingBatch
    from services.receiving import unpost_receiving_batch

    # --- доступ ---
    role = (getattr(current_user, "role", "") or "").lower()
    if not (
        role == "superadmin"
        or getattr(current_user, "is_superadmin", False)
        or getattr(current_user, "is_super_admin", False)
    ):
        flash("Only superadmin can delete receiving.", "danger")
        return redirect(url_for("inventory.receiving_list"))

    batch = db.session.get(ReceivingBatch, batch_id)
    if not batch:
        flash("Batch not found.", "warning")
        return redirect(url_for("inventory.receiving_list"))

    supplier = (getattr(batch, "supplier_name", "") or "").strip()
    invoice  = (getattr(batch, "invoice_number", "") or "").strip()
    status   = (getattr(batch, "status", "") or "").lower()

    # helper: вытащить строки batch-а (для логов, и чтобы потом не лазить в уже удалённый объект)
    def _get_lines_for_batch(b):
        try:
            return list(b.items or [])
        except Exception:
            pass
        try:
            from models import GoodsReceiptLine
            return list(GoodsReceiptLine.query.filter_by(goods_receipt_id=b.id).all())
        except Exception:
            return []

    lines = _get_lines_for_batch(batch)

    pns = {
        (getattr(it, "part_number", "") or "").strip().upper()
        for it in lines
        if (getattr(it, "part_number", "") or "").strip()
    }

    current_app.logger.info(
        "[DELETE RECEIVING] start: batch_id=%s, status=%s, supplier=%s, invoice=%s, pns=%s",
        batch_id, status, supplier, invoice, sorted(pns)
    )

    # 1. Если батч был posted → аккуратно откатываем СКЛАД через официальный сервис.
    #    ВАЖНО: это уже делает минус qty по каждой строке.
    #    После этого qty в инвентаре должен стать (старый - прихода этого батча).
    if status == "posted":
        try:
            unpost_receiving_batch(batch.id, getattr(current_user, "id", None))
            current_app.logger.info("[DELETE RECEIVING] unposted batch_id=%s", batch_id)
            # перечитаем batch после unpost (он теперь draft)
            batch = db.session.get(ReceivingBatch, batch_id)
        except Exception as e:
            current_app.logger.exception(
                "[DELETE RECEIVING] unpost failed for batch_id=%s",
                batch_id
            )
            flash(f"Unable to delete: auto-unpost failed: {e}", "danger")
            return redirect(url_for("inventory.receiving_detail", batch_id=batch_id))

    # 2. Если батч уже был draft (не posted), мы НИЧЕГО не трогаем в остатках.
    #    Почему? draft не должен быть в инвентаре. Если он как-то попал туда ранним багом —
    #    мы не будем делать второй минус. Это лучше, чем случайно обнулить сток.

    # 3. Теперь удаляем сам batch.
    try:
        batch = db.session.get(ReceivingBatch, batch_id)
        if batch:
            db.session.delete(batch)
        db.session.commit()
        current_app.logger.info("[DELETE RECEIVING] batch deleted batch_id=%s", batch_id)
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception("[DELETE RECEIVING] delete failed for batch_id=%s", batch_id)
        flash(f"Failed to delete batch: {e}", "danger")
        return redirect(url_for("inventory.receiving_detail", batch_id=batch_id))

    # 4. Чистим dedup-ключи (чтоб можно было снова импортнуть тот же invoice/file).
    keys_removed = 0
    try:
        fn = globals().get("_clear_dedup_keys_for_batch")
        if fn:
            try:
                removed = fn(batch_id, supplier=supplier, invoice=invoice)
            except TypeError:
                removed = fn(batch_id)
            keys_removed = int(removed or 0)
            current_app.logger.info(
                "[DELETE RECEIVING] dedup removed=%s for batch_id=%s",
                keys_removed, batch_id
            )
        else:
            current_app.logger.info(
                "[DELETE RECEIVING] no _clear_dedup_keys_for_batch; skip dedup cleanup for batch_id=%s",
                batch_id
            )
    except Exception:
        current_app.logger.debug(
            "[DELETE RECEIVING] dedup cleanup failed for batch_id=%s",
            batch_id,
            exc_info=True
        )

    # 5. ВАЖНО: мы НЕ удаляем Part, и мы НЕ делаем никаких "вручную минусни qty".
    #    Всё, что должно списаться, уже списал unpost_receiving_batch().
    #    Значит:
    #      было 2 → постнули +1 → стало 3
    #      unpost → 3-1 = 2
    #      delete → просто убрали сам batch, остаток остался 2
    #    Именно то, что ты хочешь.

    msg = f"Batch #{batch_id} deleted."
    if keys_removed:
        msg += f" Dedup keys removed: {keys_removed}."

    flash(msg, "success")

    current_app.logger.info(
        "[DELETE RECEIVING] done batch_id=%s, keys_removed=%s",
        batch_id, keys_removed
    )

    return redirect(url_for("inventory.receiving_list"))

@inventory_bp.post("/receiving/<int:batch_id>/clear-keys", endpoint="receiving_clear_keys")
@login_required
def receiving_clear_keys(batch_id: int):
    from flask import flash, redirect, url_for
    from flask_login import current_user
    from extensions import db
    from models import ReceivingBatch

    role = (getattr(current_user, "role", "") or "").lower()
    if not (role == "superadmin" or getattr(current_user, "is_superadmin", False) or getattr(current_user, "is_super_admin", False)):
        return ("Forbidden", 403)

    batch = db.session.get(ReceivingBatch, batch_id)
    supplier = (getattr(batch, "supplier_name", "") or "").strip() if batch else ""
    invoice  = (getattr(batch, "invoice_number", "") or "").strip() if batch else ""

    try:
        removed = _clear_dedup_keys_for_batch(batch_id, supplier=supplier, invoice=invoice)
        if removed:
            flash(f"Dedup keys cleared for Batch #{batch_id} (removed: {removed}).", "success")
        else:
            flash(f"No dedup keys found for Batch #{batch_id}.", "info")
    except Exception as e:
        flash(f"Failed to clear keys: {e}", "danger")

    return redirect(url_for("inventory.receiving_list"))

# def _ensure_norm_columns(df, default_loc: str, saved_path: str):
#     import pandas as pd
#     df = _coerce_norm_df(df).copy()
#
#     # Базовый набор
#     need_cols = [
#         "part_number", "part_name",
#         "qty", "unit_cost", "location",
#         "row_key", "source_file",
#         "supplier", "order_no", "invoice_no"
#     ]
#     for c in need_cols:
#         if c not in df.columns:
#             df[c] = None
#
#     # qty → int, отрицательные в 0
#     df["qty"] = pd.to_numeric(df["qty"], errors="coerce").fillna(0).astype(int)
#     df.loc[df["qty"] < 0, "qty"] = 0
#
#     # === ALIASES (ВАЖНО) =====================================================
#     # build_receive_movements ждёт 'quantity' → делаем алиас
#     df["quantity"] = df["qty"]
#
#     # На всякий случай поддержим разные имена цены
#     if "unit_cost" not in df.columns or df["unit_cost"].isna().all():
#         for alt in ["unitcost", "unit_price", "unitprice", "price", "cost", "last_cost"]:
#             if alt in df.columns and not df[alt].isna().all():
#                 df["unit_cost"] = df[alt]
#                 break
#     # ========================================================================
#
#     # unit_cost → float
#     df["unit_cost"] = pd.to_numeric(df["unit_cost"], errors="coerce")
#
#     # location: только пустые заполняем default_loc
#     empty_loc = df["location"].isna() | (df["location"].astype(str).str.strip() == "")
#     df.loc[empty_loc, "location"] = default_loc
#
#     # source_file: только пустые — путём загрузки
#     empty_sf = df["source_file"].isna() | (df["source_file"].astype(str).str.strip() == "")
#     df.loc[empty_sf, "source_file"] = saved_path
#
#     # row_key: генерим, если пусто
#     def _mk_key(row):
#         return f"{row.get('part_number','')}/{row.get('location','')}/{row.get('qty',0)}/{row.get('unit_cost','')}"
#     empty_rk = df["row_key"].isna() | (df["row_key"].astype(str).str.strip() == "")
#     df.loc[empty_rk, "row_key"] = df[empty_rk].apply(_mk_key, axis=1)
#
#     return df

# --- ЖЁСТКАЯ НОРМАЛИЗАЦИЯ ШАПОК ТАБЛИЦЫ ---
def _harden_preview_headers(df, default_loc: str, supplier_hint: str | None, source_file: str | None):
    """
    Гарантирует наличие и корректные типы колонок:
      part_number, part_name, quantity, unit_cost, location, supplier, source_file
    Понимает PDF-шапки: PART #, DESCRIPTION, QTY, UNIT COST, LOCATION.
    """
    import re
    import pandas as pd

    if df is None:
        return df

    # 1) точные заголовки из PDF
    hard_map = {
        "PART #": "part_number",
        "PART#": "part_number",
        "DESCRIPTION": "part_name",
        "QTY": "quantity",
        "UNIT COST": "unit_cost",
        "LOCATION": "location",
        "SUPPLIER": "supplier",
        "ORDER #": "order_no",
        "ORDER NO": "order_no",
        "PO": "order_no",
    }
    for col in list(df.columns):
        k = str(col).strip()
        if k.upper() in hard_map:
            df.rename(columns={col: hard_map[k.upper()]}, inplace=True)

    # 2) алиасы на всякий случай (вдруг normalize_table вернула иначе)
    alias = {
        "part_number": ["pn", "part", "number", "sku", "code", "item", "item #", "item#"],
        "part_name":   ["name", "descr", "description", "title"],
        "quantity":    ["qty", "count", "on_hand", "onhand", "q-ty"],
        "unit_cost":   ["cost", "price", "unitprice", "last_cost", "unit price", "unitprice"],
        "location":    ["loc", "bin", "shelf", "place"],
        "supplier":    ["vendor", "provider"],
        "order_no":    ["invoice", "invoice_no", "invoice number", "po"],
    }
    for canonical, cands in alias.items():
        if canonical not in df.columns:
            for c in cands:
                if c in df.columns:
                    df.rename(columns={c: canonical}, inplace=True)
                    break

    # 3) обязательные поля — если нет, создадим пустые
    for col in ("part_number", "part_name", "quantity", "unit_cost", "location", "supplier", "order_no", "source_file"):
        if col not in df.columns:
            df[col] = None

    # 4) если part_number пуст — попытка угадать по «похожести» на артикулы
    if df["part_number"].isna().all() or (df["part_number"].astype(str).str.strip() == "").all():
        candidate = None
        for c in df.columns:
            if c in ("part_number", "part_name", "quantity", "unit_cost", "location", "supplier", "order_no", "source_file"):
                continue
            s = df[c].astype(str).str.strip()
            mask = s.str.match(r"(?i)^(?=.*\d)[A-Z0-9\-\._/]{3,30}$")
            if mask.mean() >= 0.6:
                candidate = c
                break
        if candidate:
            df["part_number"] = df[candidate].astype(str).str.strip()

    # 5) типы
    df["quantity"] = pd.to_numeric(df["quantity"], errors="coerce").fillna(0).astype(int)
    df["unit_cost"] = pd.to_numeric(df["unit_cost"], errors="coerce").fillna(0.0)

    for col in ("part_number", "part_name", "location", "supplier", "order_no", "source_file"):
        df[col] = df[col].astype(str).replace({"None": ""}).fillna("").str.strip()

    # 6) значения по умолчанию
    if not default_loc:
        default_loc = "MAIN"
    df.loc[df["location"].eq("") | df["location"].isna(), "location"] = default_loc
    df["location"] = df["location"].str.upper()

    if supplier_hint:
        df.loc[df["supplier"].eq("") | df["supplier"].isna(), "supplier"] = supplier_hint

    if source_file:
        # заполним пустые/пробельные source_file
        empty_sf = (df["source_file"].isna()) | (df["source_file"].astype(str).str.strip() == "")
        df.loc[empty_sf, "source_file"] = source_file

    # 7) отфильтруем пустые строки, но НЕ слишком агрессивно
    keep = (
        (~df["part_number"].astype(str).str.strip().eq("")) |
        (~df["part_name"].astype(str).str.strip().eq("")) |
        (df["quantity"] > 0) |
        (df["unit_cost"] > 0)
    )
    return df[keep].copy()

def _job_token_match(col, token: str):
    t = (token or "").strip()
    if not t:
        return False

    return or_(
        func.trim(col) == t,
        col.ilike(f"{t},%"),
        col.ilike(f"%, {t},%"),
        col.ilike(f"%,{t},%"),
        col.ilike(f"%, {t}"),
        col.ilike(f"%,{t}"),
        col.ilike(f"% {t} %"),
        col.ilike(f"{t} %"),
        col.ilike(f"% {t}"),
    )


def _coerce_norm_df(obj) -> pd.DataFrame:
    """Гарантирует, что на выходе будет DataFrame. None/список/словарь → DF."""
    if isinstance(obj, pd.DataFrame):
        return obj
    if obj is None:
        return pd.DataFrame([])
    if isinstance(obj, list):
        return pd.DataFrame(obj)
    if isinstance(obj, dict):
        return pd.DataFrame([obj])
    return pd.DataFrame([])


def rows_to_norm_df(rows: list[dict], source_file: str):
    import pandas as pd
    out = []
    for i, r in enumerate(rows):
        pn = (r.get("part_number") or "").strip()
        if not pn:
            continue
        name = (r.get("part_name") or "").strip()
        loc  = (r.get("location") or "").strip()
        sup  = (r.get("supplier") or "").strip() or None

        # qty / unit_cost мягко
        q_raw = r.get("qty", r.get("quantity", 0))
        try:    qty = int(float(str(q_raw).strip() or 0))
        except: qty = 0

        u_raw = r.get("unit_cost", r.get("cost"))
        try:    unit_cost = float(str(u_raw).strip()) if u_raw not in (None, "") else None
        except: unit_cost = None

        out.append({
            "part_number": pn,
            "part_name": name,
            "qty": qty,
            "quantity": qty,
            "unit_cost": unit_cost,
            "location": loc,
            "supplier": sup,
            "source_file": source_file,
            "order_no": (r.get("order_no") or "").strip() or None,
            "invoice_no": (r.get("invoice_no") or "").strip() or None,
            "date": (r.get("date") or "").strip() or None,   # ← ДОБАВЛЕНО
            "row_key": f"{pn}|{loc or ''}|{i}",
        })
    return pd.DataFrame(out)


def fix_norm_records(records: list[dict], default_loc: str = "MAIN") -> list[dict]:
    """Правит записи уже ПОСЛЕ normalize_table: разлепляет PN/NAME, чистит мусор/итоги, дописывает LOCATION."""
    out = []
    for r in records:
        pn  = str(r.get("part_number", "") or "").strip()
        nm  = str(r.get("part_name", "") or "").strip()

        # 0) мусор и итоги (на всякий случай)
        if _MARKER_RX.match(pn) or _MARKER_RX.match(nm):
            continue
        if _TOTALS_RX.match(pn) or _TOTALS_RX.match(nm):
            continue

        # 1) PN выглядит как "PN rest..." → делим
        m = _PN_SPLIT_RX.match(pn)
        if m:
            tok, rest = m.group(1).strip(), m.group(2).strip()
            pn = tok
            if (not nm) or (nm.upper() == r.get("part_number","").upper()) or (nm.upper() == tok.upper()):
                nm = rest
            elif rest and rest not in nm:
                nm = f"{nm} {rest}".strip()

        # 2) PN пуст, а NAME = "PN rest..." → переносим
        elif not pn:
            m2 = _PN_SPLIT_RX.match(nm)
            if m2:
                pn, nm = m2.group(1).strip(), m2.group(2).strip()

        # 3) NAME дублирует PN → почистим
        if pn and nm and pn.upper() == nm.upper():
            nm = ""

        # 4) Локация по умолчанию
        loc = (r.get("location") or "").strip() or default_loc

        # 5) Записываем обратно
        r["part_number"] = pn
        r["part_name"]   = nm
        r["location"]    = loc

        # 6) Отсекаем полностью пустые
        if pn or nm or (r.get("quantity") or 0) or (r.get("unit_cost") or 0):
            out.append(r)
    return out

def fix_pn_and_description_in_df(df):
    if df is None or getattr(df, "empty", True):
        return df
    if "part_number" not in df.columns:
        df["part_number"] = ""
    if "part_name" not in df.columns:
        desc_col = None
        for k in df.columns:
            lk = k.lower().strip()
            if lk in ("description", "descr", "desc", "name", "part_name"):
                desc_col = k; break
        df["part_name"] = df[desc_col] if desc_col else ""

    pn = df["part_number"].astype(str); nm = df["part_name"].astype(str)
    new_pn, new_nm = pn.copy(), nm.copy()

    for i in df.index:
        cur_pn = (pn.at[i] or "").strip()
        cur_nm = (nm.at[i] or "").strip()

        m = _PN_FIRST_TOKEN_RX.match(cur_pn)
        if m:
            tok, rest = m.group(1).strip(), m.group(2).strip()
            new_pn.at[i] = tok
            if not cur_nm or cur_nm.upper() in (cur_pn.upper(), tok.upper()):
                new_nm.at[i] = rest
            elif rest and rest not in cur_nm:
                new_nm.at[i] = f"{cur_nm} {rest}"
            continue

        if not cur_pn:
            m2 = _PN_FIRST_TOKEN_RX.match(cur_nm)
            if m2:
                new_pn.at[i] = m2.group(1).strip()
                new_nm.at[i] = m2.group(2).strip()
                continue

        if cur_pn and cur_nm and cur_pn.upper() == cur_nm.upper():
            new_nm.at[i] = ""

    df["part_number"] = new_pn
    df["part_name"]   = new_nm
    return df

def _promote_split_pn_desc(df):
    """
    Если в исходной таблице PN склеен с описанием в одном поле
    (например: '60034 1/2FLRX1/2FIPANGG EZI'), создаём/заполняем
    явные колонки df['part_number'] и df['part_name'].
    Логика мягкая: применяем только если удаётся нормально сплитнуть
    заметную часть строк.
    """
    if df is None or getattr(df, "empty", True):
        return df

    # 1) найдём текстовые колонки-кандидаты
    text_cols = [c for c in df.columns if getattr(df[c], "dtype", None) == object]
    if not text_cols:
        return df

    # 2) если уже есть внятные part_number/part_name — уходим
    has_pn  = any(c.lower() == "part_number" for c in df.columns)
    has_name= any(c.lower() in ("part_name", "description", "descr", "desc", "name") for c in df.columns)
    if has_pn and has_name:
        return df  # не вмешиваемся

    # 3) пробуем найти колонку, где большинство строк выглядит как "TOKEN + пробел + текст"
    best_col = None
    best_score = 0
    for c in text_cols:
        series = df[c].astype(str).str.strip()
        m = series.str.match(_PN_TOKEN_RX, na=False)
        score = float(m.sum()) / max(1, len(series))
        if score > best_score:
            best_score, best_col = score, c

    # применяем только если >= 0.5 строк в колонке соответствует паттерну
    if not best_col or best_score < 0.5:
        return df

    # 4) делим выбранную колонку на PN + NAME
    pn_series  = []
    name_series= []
    for s in df[best_col].astype(str).tolist():
        s = s.strip()
        m = _PN_TOKEN_RX.match(s)
        if m:
            pn_series.append(m.group(1))
            name_series.append(m.group(2).strip())
        else:
            pn_series.append("")
            name_series.append(s)

    # 5) Создаём/заполняем явные колонки — так normalize_table возьмёт их без эвристик
    if "part_number" not in df.columns:
        df["part_number"] = pn_series
    else:
        # заполняем только пустые
        df["part_number"] = df["part_number"].astype(str).where(df["part_number"].astype(str).str.strip() != "", pn_series)

    # имя/описание
    # если есть 'part_name' — дополняем, иначе создаём
    if "part_name" in df.columns:
        cur = df["part_name"].astype(str)
        df["part_name"] = cur.where(cur.str.strip() != "", name_series)
    else:
        df["part_name"] = name_series

    return df

def drop_vendor_noise_rows(df):
    if df is None or getattr(df, "empty", True):
        return df
    text_cols = [c for c in df.columns if getattr(df[c], "dtype", None) == object]
    if not text_cols:
        return df
    has_marker = df[text_cols].apply(lambda s: s.astype(str).str.match(_MARKER_RX, na=False)).any(axis=1)
    return df[~has_marker].reset_index(drop=True)

def detect_supplier_hint(df, fname: str | None = None) -> str | None:
    fn = (fname or "").lower()
    if "reliable" in fn: return "Reliable"
    if "marcone"  in fn: return "Marcone"
    if df is not None and not df.empty:
        sample = " ".join(
            str(x).lower()
            for col in df.columns
            for x in df[col].head(50).astype(str).tolist()
        )
        if "reliable" in sample: return "Reliable"
        if "marcone"  in sample: return "Marcone"
    return None


def _coalesce_same_parts(rows: list[dict]) -> list[dict]:
    """
    Склеивает строки с одинаковым part_number (и одинаковой ценой),
    суммируя quantity. PN нормализуем в UPPER.
    """
    acc = {}
    for r in rows:
        pn = (r.get("part_number") or r.get("pn") or "").strip().upper()
        if not pn:
            continue
        price = float((r.get("unit_cost") or r.get("price") or 0) or 0)
        key = (pn, price)
        if key not in acc:
            acc[key] = {
                "part_number": pn,
                "part_name": (r.get("part_name") or r.get("description") or r.get("descr") or "").strip(),
                "quantity": 0,
                "unit_cost": price,
                "location": (r.get("location") or r.get("supplier") or "").strip(),
            }
        acc[key]["quantity"] += int((r.get("quantity") or r.get("qty") or 0) or 0)
    # отбрасываем пустые
    return [v for v in acc.values() if v["quantity"] > 0]



def _is_return_row(r) -> bool:
    """Возвратная строка — это строка с отрицательным количеством."""
    try:
        return (r.quantity or 0) < 0
    except Exception:
        return False

def _ensure_column(table: str, column: str, ddl_type: str):
    """Безопасно добавляет колонку, если её нет (SQLite).
       Если таблицы нет — просто логируем и выходим (без ошибок)."""
    try:
        # есть ли таблица?
        row = db.session.execute(
            db.text("SELECT name FROM sqlite_master WHERE type IN ('table','view') AND name=:t"),
            {"t": table}
        ).fetchone()
        if not row:
            logging.info("Skip ensure column %s.%s: table does not exist", table, column)
            return

        # есть ли колонка?
        rows = db.session.execute(db.text(f"PRAGMA table_info({table})")).fetchall()
        names = {r[1] for r in rows}  # name на индексе 1
        if column in names:
            return

        logging.info("Adding column %s to %s ...", column, table)
        db.session.execute(db.text(f"ALTER TABLE {table} ADD COLUMN {column} {ddl_type}"))
        db.session.commit()
        logging.info("Added column %s to %s.", column, table)
    except Exception as e:
        logging.exception("Failed to ensure column %s.%s: %s", table, column, e)


def _ensure_invoice_number_for_records(records, issued_to, issued_by, reference_job, issue_date, location):
    """
    Если у всех строк invoice_number == None — создаёт батч и закрепляет уникальный номер.
    Исторические инвойсы с уже заданным номером НЕ трогаем.
    """
    from extensions import db
    if not records:
        return

    if all(getattr(r, "invoice_number", None) is None for r in records):
        # Используем уже существующую у тебя функцию
        batch = _create_batch_for_records(
            records=records,
            issued_to=issued_to,
            issued_by=issued_by,
            reference_job=reference_job,
            issue_date=issue_date,
            location=location or None,
        )
        # номер закреплён в batch.invoice_number
        # flush/commit снаружи, в update_invoice


def _format_invoice_no(n: int | None) -> str:
    """UI/PDF формат: 6 цифр с ведущими нулями; '—' если None."""
    return f"{int(n):06d}" if n else "—"


def _tech_norm(s: str) -> str:
    return (s or '').strip().upper()

def _next_invoice_number() -> int:
    """
    Следующий invoice_number:
      max(IssuedBatch.invoice_number, IssuedPartRecord.invoice_number, INVOICE_START_AT-1) + 1
      (учитывает legacy-строки и гарантирует старт с 140)
    """
    # max по батчам (новая схема)
    max_batch = db.session.query(
        func.coalesce(func.max(IssuedBatch.invoice_number), 0)
    ).scalar()

    # max по строкам (legacy)
    max_line = db.session.query(
        func.coalesce(func.max(IssuedPartRecord.invoice_number), 0)
    ).scalar()

    try:
        mb = int(max_batch or 0)
    except Exception:
        mb = 0
    try:
        ml = int(max_line or 0)
    except Exception:
        ml = 0

    # базовый «семечко», чтобы первый новый = 140
    seed = INVOICE_START_AT - 1
    return max(mb, ml, seed) + 1

def _next_invoice_number() -> int:
    """Новый уникальный invoice_number на основе максимумов по Batch и Record."""
    mb = db.session.query(func.coalesce(func.max(IssuedBatch.invoice_number), 0)).scalar() or 0
    ml = db.session.query(func.coalesce(func.max(IssuedPartRecord.invoice_number), 0)).scalar() or 0
    return max(int(mb), int(ml)) + 1

def _create_batch_for_records(
    records: list,
    issued_to: str,
    issued_by: str,
    reference_job: str | None = None,
    issue_date: datetime | None = None,
    location: str | None = None,
    work_order_id: int | None = None,   # ✅ ADD
):
    """
    Создаёт IssuedBatch с уникальным invoice_number и привязывает все строки.
    Использует SAVEPOINT (begin_nested), чтобы коллизия unique не откатывала всю сессию.
    """
    if not records:
        raise ValueError("No records passed to _create_batch_for_records")

    from datetime import datetime, timezone
    from zoneinfo import ZoneInfo

    LA_TZ = ZoneInfo("America/Los_Angeles")

    if issue_date is None:
        issue_date = datetime.now(LA_TZ).astimezone(timezone.utc).replace(tzinfo=None)

    for _ in range(5):  # несколько попыток на случай гонки за номер
        inv_no = _next_invoice_number()
        try:
            with db.session.begin_nested():  # SAVEPOINT
                batch = IssuedBatch(
                    invoice_number=inv_no,
                    issued_to=issued_to,
                    issued_by=issued_by or "system",
                    reference_job=reference_job,
                    issue_date=issue_date,
                    location=(location or None),
                    work_order_id=work_order_id,   # ✅ ADD
                )
                db.session.add(batch)
                db.session.flush()  # резервируем уникальный номер (может кинуть IntegrityError)

                # привязать строки к батчу + синхронизировать «шапку»
                for r in records:
                    r.batch_id = batch.id
                    r.invoice_number = inv_no
                    r.issued_to = issued_to
                    r.issued_by = issued_by or "system"
                    r.reference_job = reference_job
                    if location:
                        r.location = location

                db.session.flush()

            return batch  # успех

        except IntegrityError:
            db.session.rollback()
            continue

    raise RuntimeError("Не удалось сгенерировать уникальный invoice_number после нескольких попыток")


def _parse_dt_flex(s: str):
    """Безопасный парсер даты/времени.
       Поддерживает:
         - YYYY-MM-DD HH:MM:SS.%f
         - YYYY-MM-DD HH:MM:SS
         - YYYY-MM-DD
       Возвращает datetime или None.
    """
    if not s:
        return None
    s = str(s).strip()
    for fmt in ('%Y-%m-%d %H:%M:%S.%f',
                '%Y-%m-%d %H:%M:%S',
                '%Y-%m-%d'):
        try:
            dt = datetime.strptime(s, fmt)
            if fmt == '%Y-%m-%d':
                # Если только дата → прикрепляем полночь
                return datetime.combine(dt.date(), time.min)
            return dt
        except Exception:
            continue
    return None

def _parse_units_form(form):
    """
    Превращает плоский request.form во вложенную структуру:
    [
      {
        "brand": "...", "model": "...", "serial": "...",
        "rows": [
          {"part_number": "...", "part_name": "...", "quantity": 1, "unit_cost": 12.34, ...},
          ...
        ]
      },
      ...
    ]
    """
    units = {}

    # мета-поля юнитов (brand/model/serial)
    for k in form.keys():
        m = _units_re.match(k)
        if not m:
            continue
        ui = int(m.group(1))
        key = m.group(2)
        units.setdefault(ui, {"brand": "", "model": "", "serial": "", "rows": {}})
        units[ui][key] = (form.get(k) or "").strip()

    # строки
    for k in form.keys():
        m = _rows_re.match(k)
        if not m:
            continue
        ui = int(m.group(1))
        ri = int(m.group(2))
        key = m.group(3)

        units.setdefault(ui, {"brand": "", "model": "", "serial": "", "rows": {}})
        units[ui]["rows"].setdefault(ri, {
            "part_number": "", "part_name": "", "quantity": 0,
            "alt_numbers": "", "supplier": "", "backorder_flag": False,
            "line_status": "search_ordered", "unit_cost": ""
        })

        val = form.get(k)

        if key == "backorder_flag":
            units[ui]["rows"][ri][key] = True  # наличие ключа = checked
        else:
            units[ui]["rows"][ri][key] = (val or "").strip()

    # привести словарь → упорядоченный список
    result = []
    for ui in sorted(units.keys()):
        u = units[ui]
        # rows по порядку
        rows = [u["rows"][ri] for ri in sorted(u["rows"].keys())]
        u["rows"] = rows
        result.append(u)

    return result

@inventory_bp.post("/api/issued/confirm_toggle", endpoint="issued_confirm_toggle")
@login_required
def issued_confirm_toggle():
    """Техник ставит подтверждение через fetch (one-way), админ может ставить/снимать через форму."""
    from flask import request, redirect, url_for, jsonify
    from datetime import datetime, timezone
    from flask_login import current_user
    from extensions import db
    from models import IssuedPartRecord, WorkOrder

    payload = request.get_json(silent=True) or {}

    def _get(k):
        v = request.form.get(k)
        if v is None:
            v = payload.get(k)
        if v is None:
            v = request.args.get(k)
        return v

    rec_id_raw = _get("record_id")
    wo_id_raw = _get("wo_id")
    state_raw = _get("state")

    try:
        rec_id = int(rec_id_raw or 0)
    except Exception:
        rec_id = 0

    requested_checked = str(state_raw or "").strip().lower() in ("1", "true", "on", "yes")

    role = (getattr(current_user, "role", "") or "").strip().lower()
    user_id = getattr(current_user, "id", None)
    username = (
        getattr(current_user, "username", "")
        or getattr(current_user, "email", "")
        or getattr(current_user, "name", "")
        or str(user_id)
    )

    is_adminlike = role in ("admin", "superadmin")

    if not wo_id_raw or not rec_id:
        if is_adminlike:
            return redirect(url_for("inventory.wo_detail", wo_id=wo_id_raw or 0))
        return jsonify({"ok": False, "error": "MISSING_PARAMS"}), 400

    wo = WorkOrder.query.get(int(wo_id_raw))
    if not wo:
        if is_adminlike:
            return redirect(url_for("inventory.wo_detail", wo_id=wo_id_raw))
        return jsonify({"ok": False, "error": "WO_NOT_FOUND"}), 404

    wo_tech_id = getattr(wo, "technician_id", None)
    wo_tech_name = (
        (getattr(wo, "technician_username", None) or "") or
        (getattr(wo, "technician_name", None) or "")
    ).strip().lower()

    me_name = (
        (getattr(current_user, "username", "") or "") or
        (getattr(current_user, "email", "") or "") or
        (getattr(current_user, "name", "") or "")
    ).strip().lower()

    is_my_wo = (
        (wo_tech_id is not None and user_id is not None and int(wo_tech_id) == int(user_id))
        or (wo_tech_name and me_name and wo_tech_name == me_name)
    )

    is_techlike = role in ("technician", "tech") and is_my_wo

    rec = IssuedPartRecord.query.filter_by(id=rec_id).first()
    if not rec:
        if is_adminlike:
            return redirect(url_for("inventory.wo_detail", wo_id=wo_id_raw))
        return jsonify({"ok": False, "error": "REC_NOT_FOUND"}), 404

    currently_confirmed = bool(rec.confirmed_by_tech)

    if is_adminlike:
        new_state = bool(requested_checked)
    elif is_techlike:
        # one-way for technician
        new_state = True if requested_checked else currently_confirmed
    else:
        return jsonify({"ok": False, "error": "FORBIDDEN"}), 403

    if new_state == currently_confirmed:
        if is_adminlike:
            return redirect(url_for("inventory.wo_detail", wo_id=wo_id_raw))
        return jsonify({"ok": True, "noop": True, "confirmed": currently_confirmed})

    rec.confirmed_by_tech = new_state
    if new_state:
        rec.confirmed_by = username or "tech"
        rec.confirmed_at = datetime.now(timezone.utc)
    else:
        rec.confirmed_by = None
        rec.confirmed_at = None

    db.session.commit()

    if is_adminlike:
        return redirect(url_for("inventory.wo_detail", wo_id=wo_id_raw))

    return jsonify({"ok": True, "record_id": rec_id, "confirmed": bool(rec.confirmed_by_tech)})


# ---------- STOCK HINT API ----------
from flask import jsonify, request
from sqlalchemy import func

# ...

@inventory_bp.get("/api/stock_hint", endpoint="api_stock_hint")
@login_required
def api_stock_hint():
    """
    Простой API для подсказки наличия:
    принимает pn, qty, wh и возвращает JSON с количеством и подсказкой.
    """
    pn = (request.args.get("pn") or "").strip().upper()
    qty_needed_raw = (request.args.get("qty") or "").strip()
    wh = (request.args.get("wh") or "").strip()

    try:
        qty_needed = int(qty_needed_raw or 0)
    except ValueError:
        qty_needed = 0

    if not pn:
        return jsonify({
            "ok": False,
            "error": "NO_PN",
            "hint": "WAIT",
            "qty_available": 0
        }), 400

    # базовый запрос по Part
    q = db.session.query(Part).filter(Part.part_number == pn)

    # если указан склад – фильтруем по location (регистр не важен)
    if wh:
        q = q.filter(func.lower(Part.location) == func.lower(wh))

    part = q.first()

    if not part:
        return jsonify({
            "ok": True,
            "part_number": pn,
            "qty_available": 0,
            "hint": "WAIT",
            "location": None,
            "unit_cost": None,
            "name": None,
        })

    qty_available = int(part.quantity or 0)
    is_stock = qty_available >= qty_needed if qty_needed > 0 else qty_available > 0
    hint = "STOCK" if is_stock else "WAIT"

    return jsonify({
        "ok": True,
        "part_number": pn,
        "qty_available": qty_available,
        "hint": hint,
        "location": part.location,
        "unit_cost": part.unit_cost,
        "name": part.name,
    })

@inventory_bp.get("/debug/db_objects")
def debug_db_objects():
    import os, sqlite3
    from flask import jsonify, current_app
    dbp = os.path.normpath(os.path.join(current_app.instance_path, "inventory.db"))
    data = {"db_path": dbp, "tables": [], "views": []}
    with sqlite3.connect(dbp) as con:
        data["tables"] = [r[0] for r in con.execute(
            "SELECT name FROM sqlite_master WHERE type='table' ORDER BY name").fetchall()]
        data["views"] = [r[0] for r in con.execute(
            "SELECT name FROM sqlite_master WHERE type='view' ORDER BY name").fetchall()]
    return jsonify(data)


def _db_path() -> str:
    # Используем instance-папку Flask, там у тебя и лежит inventory.db
    p = os.path.join(current_app.instance_path, "inventory.db")
    return os.path.normpath(p)


# >>> ADD HELPERS (ONCE)
def _distinct_suppliers(wo_id: int):
    sql = """
    SELECT DISTINCT supplier
    FROM v_work_order_parts_dto
    WHERE wo_id = ? AND supplier <> ''
    ORDER BY supplier COLLATE NOCASE
    """
    dbp = _db_path()
    with sqlite3.connect(dbp) as con:
        con.row_factory = sqlite3.Row
        try:
            return [r[0] for r in con.execute(sql, (wo_id,)).fetchall()]
        except sqlite3.OperationalError as e:
            current_app.logger.error("distinct_suppliers error: %s", e)
            current_app.logger.error("DB_PATH = %s", dbp)
            exists = con.execute("""
                SELECT name FROM sqlite_master
                WHERE type IN ('view','table') AND name='v_work_order_parts_dto'
            """).fetchone()
            current_app.logger.error("v_work_order_parts_dto exists? %s", bool(exists))
            return []

def _issued_batches_stub(wo_id: int):
    return [{
        "issued_at": "2025-09-16 14:23",
        "technician": "Jane Doe",
        "canonical_ref": str(wo_id),
        "items": 3,
        "report_id": f"RPT-{wo_id}",
    }]


def _parse_q(q: str):
    if not q: return {"type":"text","value":""}
    m = _PREFIX_RE.match(q)
    return {"type": m.group(1).lower(), "value": m.group(2).strip()} if m else {"type":"text","value":q.strip()}

def get_on_hand(part_number: str) -> int:
    """Возвращает суммарный остаток по PN (всех локаций) из твоей модели Part."""
    try:
        total = db.session.query(db.func.sum(Part.quantity)).filter(
            Part.part_number == part_number.upper()
        ).scalar()
        return int(total or 0)
    except Exception:
        return 0

def compute_availability(work_order: "WorkOrder"):
    """
    Возвращает список словарей по каждой строке WO:
    [{
      wop_id, part_number, part_name, requested, on_hand, issue_now, status_hint
    }, ...]

    Логика:
      - Только при status == "ordered" обращаемся к складу:
          on_hand = get_on_hand(PN)
          issue_now = min(requested, on_hand)
          hint = "STOCK" | "WAIT {need} (stock {on})"
      - При других статусах (например, "search_ordered"):
          on_hand = 0, issue_now = 0, hint = "WAIT {requested} (not ordered)"
    """
    rows = []
    # считаем «можно ли проверять склад»
    check_stock = (getattr(work_order, "status", "") or "").lower() == "ordered"

    for wop in (work_order.parts or []):
        # нормализуем PN
        pn = (wop.part_number or "").strip().upper()
        # количество, безопасно
        req = int(wop.quantity or 0)
        if req < 0:
            req = 0

        if check_stock and pn:
            on = int(get_on_hand(pn) or 0)
            issue = min(req, on)
            if req <= 0:
                hint = "WAIT 0"
            elif on >= req:
                hint = "STOCK"
            else:
                need = req - on
                hint = f"WAIT {need} (stock {on})"
        else:
            # статус ещё не ordered → по твоей логике не сверяем склад
            on = 0
            issue = 0
            hint = f"WAIT {req} (not ordered)" if req > 0 else "WAIT 0"

        rows.append({
            "wop_id": wop.id,
            "part_number": pn,
            "part_name": (wop.part_name or "").strip(),
            "requested": req,
            "on_hand": on,
            "issue_now": issue,
            "status_hint": hint,
        })

    return rows

def compute_availability_unit(unit: "WorkUnit", wo_status: str):
    """
    Возвращает список строк по одному unit:
    [{unit_id, unit_label, part_number, part_name, requested, on_hand, issue_now, status_hint}, ...]
    """
    rows = []
    label = f"{(unit.brand or '').strip()} {(unit.model or '').strip()} / S/N {(unit.serial or '').strip()}".strip()
    label = label or f"UNIT #{unit.id}"

    for wup in (unit.parts or []):
        req = int(wup.quantity or 0)
        on  = get_on_hand((wup.part_number or "").upper())
        issue = 0
        # как договаривались: пока WO не в "ordered" — не выдаём, только WAIT ...
        if wo_status == "ordered":
            issue = max(0, min(req, on))
            hint = "STOCK" if on >= req else f"WAIT {req - on} (stock {on})"
        else:
            hint = f"WAIT {req} (not ordered)"

        rows.append({
            "unit_id": unit.id,
            "unit_label": label,
            "part_number": wup.part_number,
            "part_name": wup.part_name or "",
            "requested": req,
            "on_hand": on,
            "issue_now": issue,
            "status_hint": hint,
        })
    return rows

def compute_availability_multi(wo: "WorkOrder"):
    """
    Объединяет:
      - строки по юнитам (новая схема)
      - и твои старые строки work_order.parts (чтобы legacy не сломать)
    """
    all_rows = []

    # Новые unit-строки
    for unit in getattr(wo, "units", []) or []:
        all_rows.extend(compute_availability_unit(unit, wo.status))

    # Legacy-строки (если у заказа ещё есть старые parts)
    for wop in getattr(wo, "parts", []) or []:
        req = int(wop.quantity or 0)
        on  = get_on_hand((wop.part_number or "").upper())
        if wo.status == "ordered":
            issue = max(0, min(req, on))
            hint  = "STOCK" if on >= req else f"WAIT {req - on} (stock {on})"
        else:
            issue = 0
            hint  = f"WAIT {req} (not ordered)"
        all_rows.append({
            "unit_id": None,
            "unit_label": "LEGACY",
            "part_number": wop.part_number,
            "part_name": wop.part_name or "",
            "requested": req,
            "on_hand": on,
            "issue_now": issue,
            "status_hint": hint,
        })

    return all_rows

def _issue_records_bulk(
    issued_to: str,
    reference_job: str,
    items: list,
    billed_price_per_item: float | None = None,
):
    """
    items: list of dicts:
      {
        "part_id": int,
        "qty": int,
        "unit_price": float | None,
        "inv_ref": str | None,   # INV# from WorkOrderPart.invoice_number
      }

    Returns:
      (issue_date: datetime, created_records: list[IssuedPartRecord])
    """
    if not items:
        raise ValueError("No items to issue")

    issue_date = datetime.utcnow()
    created_records: list[IssuedPartRecord] = []

    for it in items:
        part_id = int(it["part_id"])
        qty = max(0, int(it.get("qty") or 0))
        if qty <= 0:
            continue

        part = Part.query.get(part_id)
        if not part:
            continue

        on_hand = int(part.quantity or 0)
        issue_now = min(qty, on_hand)
        if issue_now <= 0:
            continue

        # price snapshot
        if "unit_price" in it and it["unit_price"] is not None:
            price_to_fix = float(it["unit_price"])
        elif billed_price_per_item is not None:
            price_to_fix = float(billed_price_per_item)
        else:
            price_to_fix = float(part.unit_cost or 0.0)

        # decrement stock (NO COMMIT here)
        part.quantity = on_hand - issue_now
        db.session.add(part)

        inv_ref = (str(it.get("inv_ref") or "").strip()[:32] or None)
        base_loc = (getattr(part, "location", "") or "").strip() or None

        rec = IssuedPartRecord(
            part_id=part.id,
            quantity=issue_now,
            issued_to=(issued_to or "").strip(),
            reference_job=(reference_job or "").strip(),
            issued_by=getattr(current_user, "username", "system"),
            issue_date=issue_date,
            unit_cost_at_issue=price_to_fix,
            location=base_loc,    # snapshot location
        )
        if hasattr(rec, "inv_ref"):
            rec.inv_ref = inv_ref  # сохраняем INV отдельно

        db.session.add(rec)
        created_records.append(rec)

    if not created_records:
        raise ValueError("Nothing available to issue")

    # ВАЖНО: только flush, чтобы появились id (но без commit)
    db.session.flush()
    return issue_date, created_records

def _is_return_record(record: IssuedPartRecord) -> bool:
    """Признак 'возвратной' строки — отрицательное количество или reference_job начинается с RETURN."""
    if record.quantity is not None and record.quantity < 0:
        return True
    ref = (record.reference_job or "").strip().upper()
    return ref.startswith("RETURN")

def _is_return_group(records: list[IssuedPartRecord]) -> bool:
    """Группа — возвратная, если есть хотя бы одна позиция с qty<0 или ref начинается с RETURN."""
    if not records:
        return False
    if any((r.quantity or 0) < 0 for r in records):
        return True
    ref = (records[0].reference_job or "").strip().upper()
    return ref.startswith("RETURN")

def _fetch_invoice_group(issued_to: str, reference_job: str|None, issued_by: str, issue_date):
    """Достаём ВСЕ записи накладной за конкретный день по ключу (issued_to, reference_job, issued_by, дата)."""
    from datetime import datetime
    start = datetime.combine(issue_date.date(), datetime.min.time())
    end   = datetime.combine(issue_date.date(), datetime.max.time())
    return IssuedPartRecord.query.filter(
        IssuedPartRecord.issued_to == issued_to,
        IssuedPartRecord.reference_job == reference_job,
        IssuedPartRecord.issued_by == issued_by,
        IssuedPartRecord.issue_date.between(start, end)
    ).all()
# ==== /RETURN HELPERS ====


# ---------- Helpers for editable preview ----------


def parse_preview_rows(form):
    """
    Parse rows[*][field] inputs from the preview form into a list of dicts.
    Accepts fields: part_number, part_name, quantity, unit_cost, supplier, location, order_no, date.
    Skips fully empty rows.
    """
    import re
    pattern = re.compile(r'^rows\[(\d+)\]\[(\w+)\]$')
    indexed = {}
    for key, val in form.items():
        m = pattern.match(key)
        if not m:
            continue
        idx = int(m.group(1))
        field = m.group(2)
        indexed.setdefault(idx, {})[field] = (val or "").strip()
    rows = []
    for i in sorted(indexed.keys()):
        r = indexed[i]
        # keep row if any meaningful cell is present
        if any((r.get(k) or "").strip() for k in ("part_number","part_name","quantity","unit_cost","supplier","location","order_no","date")):
            rows.append(r)
    return rows


def rows_to_dataframe(rows):
    """Turn list[dict] into a DataFrame with light type coercion."""
    import pandas as pd
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows)
    # Ensure all expected columns exist
    for c in ["part_number","part_name","supplier","location","order_no","date","quantity","unit_cost"]:
        if c not in df.columns:
            df[c] = ""
    # Types
    df["quantity"]  = pd.to_numeric(df["quantity"], errors="coerce").fillna(0).astype(int)
    df["unit_cost"] = pd.to_numeric(df["unit_cost"], errors="coerce")
    for c in ["part_number","part_name","supplier","location","order_no","date"]:
        df[c] = df[c].astype(str).fillna("").str.strip()
    return df


# --- PDF items table fixer ----------------------------------------------------
def coerce_invoice_items(df_raw):
    """
    Find an invoice line-items header inside a noisy DataFrame (from pdfplumber)
    and rebuild a clean items table with canonical column names that your
    normalize_table() already understands.

    Canonical headers we try to produce:
      "QTY", "PART #", "DESCRIPTION", "UNIT COST", "TOTAL", plus optional
      "ORDER #", "DATE", "LOCATION".
    If nothing recognizable is found, returns df_raw unchanged.
    """
    import pandas as pd
    import numpy as np
    import re

    if df_raw is None or getattr(df_raw, "empty", True):
        return df_raw

    # Make a string-only view for scanning and cleanup (keeps original untouched)
    str_df = df_raw.copy()
    for j in range(str_df.shape[1]):
        s = str_df.iloc[:, j].astype(str).str.replace("\xa0", " ", regex=False)
        str_df.iloc[:, j] = s.fillna("").str.strip()

    # Canonical buckets (keys are the final column names we want to emit)
    BUCKETS = {
        "QTY":         ["QTY", "QUANTITY", "Q-ty", "QTY ORDERED", "ORDERED QTY"],
        "PART #":      ["PART #", "PART", "PART NO", "PART NUMBER", "ITEM", "ITEM #", "ITEM NO", "SKU", "MODEL", "MODEL #"],
        "DESCRIPTION": ["DESCRIPTION", "DESCR.", "ITEM DESCRIPTION", "DESC", "PRODUCT", "NAME"],
        "UNIT COST":   ["UNIT COST", "UNIT PRICE", "PRICE", "COST", "PRICE $", "PRICE USD"],
        "TOTAL":       ["TOTAL", "EXT", "EXTENDED", "LINE TOTAL", "AMOUNT"],

        # Optional extras — if present we keep them, else no problem
        "ORDER #":     ["ORDER #", "ORDER", "PO #", "SO #", "WORK ORDER", "WO"],
        "DATE":        ["DATE", "ORDER DATE", "INVOICE DATE"],
        "LOCATION":    ["LOCATION", "BIN", "SHELF", "PLACE", "LOC"],
    }

    def norm_key(s: str) -> str:
        # Normalize for fuzzy match: remove non-alphanumerics and uppercase
        return re.sub(r"[\W_]+", "", s.upper())

    # Pre-map all tokens to their bucket (final canonical name)
    token2bucket = {}
    for bucket_name, tokens in BUCKETS.items():
        for t in tokens:
            token2bucket[norm_key(t)] = bucket_name

    # Choose the "best" header row: the one covering the largest number of buckets
    best_idx, best_score = None, -1
    for i in range(str_df.shape[0]):
        seen = set()
        for val in str_df.iloc[i, :].tolist():
            k = norm_key(str(val))
            for token, bucket in token2bucket.items():
                if token and token in k:
                    seen.add(bucket)
        score = len(seen)
        if score > best_score:
            best_score, best_idx = score, i

    # If we cannot find a reasonable header (at least 2 buckets), bail out
    if best_idx is None or best_score < 2:
        return df_raw

    header_row = str_df.iloc[best_idx, :].tolist()

    # Map each header cell to a canonical bucket if possible
    def map_header_cell(s):
        ks = norm_key(str(s))
        for token, bucket in token2bucket.items():
            if token and token in ks:
                return bucket
        return ""  # unknown/unused

    mapped = [map_header_cell(v) for v in header_row]

    # Ensure uniqueness and fill blanks as col_<index>
    seen = {}
    final_cols = []
    for idx, m in enumerate(mapped):
        name = m if m else f"col_{idx}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        final_cols.append(name)

    # Body = rows under the header
    body = str_df.iloc[best_idx + 1 : ].copy()
    if body.empty:
        return df_raw

    # Rectangularize to header width
    if body.shape[1] < len(final_cols):
        for _ in range(len(final_cols) - body.shape[1]):
            body[f"pad_{_}"] = ""
    body = body.iloc[:, :len(final_cols)]
    body.columns = final_cols

    # Light type coercion for filters
    def to_int_safe(x):
        try:
            sx = str(x).replace(",", "").strip()
            if sx == "": return np.nan
            return int(float(sx))
        except:
            return np.nan

    def to_float_safe(x):
        try:
            sx = str(x).replace("$", "").replace(",", "").strip()
            if sx == "": return np.nan
            return float(sx)
        except:
            return np.nan

    if "QTY" in body.columns:
        body["QTY_num"] = body["QTY"].apply(to_int_safe)
    if "UNIT COST" in body.columns:
        body["UNIT_COST_num"] = body["UNIT COST"].apply(to_float_safe)

    # Keep rows that look like real items:
    keep_mask = False
    if "QTY_num" in body.columns:
        keep_mask = (body["QTY_num"].notna()) | (body["QTY"].astype(str).str.len() > 0)
    if "UNIT_COST_num" in body.columns:
        keep_mask = keep_mask | (body["UNIT_COST_num"].notna())
    if "PART #" in body.columns:
        keep_mask = keep_mask | (body["PART #"].astype(str).str.strip() != "")
    if "DESCRIPTION" in body.columns:
        keep_mask = keep_mask | (body["DESCRIPTION"].astype(str).str.strip() != "")

    if hasattr(keep_mask, "any"):
        body = body[keep_mask]

    # Drop fully empty rows
    body = body.replace("", np.nan).dropna(how="all").fillna("")

    return body if not body.empty else df_raw



def dataframe_from_pdf(path, try_ocr: bool = False):
    """
    1) Try extracting tables from a text-based PDF (pdfplumber).
    2) If no tables and try_ocr=True: OCR with poppler+Tesseract and then:
       - try to parse Marcone-like rows into PART # / DESCR. / QTY / UNIT COST / SUPPLIER
       - fallback to a generic table (at least 'text' column) so preview is never empty.
    """
    import os, re
    import pdfplumber
    import pandas as pd
    import numpy as np
    from flask import current_app
    from config import Config

    def log(msg: str):
        try:
            current_app.logger.debug(msg)
        except Exception:
            pass

    def unique_headers(headers):
        seen = {}
        out = []
        for i, h in enumerate(headers):
            h = (h or "").strip() or f"col_{i}"
            if h in seen:
                seen[h] += 1
                h = f"{h}_{seen[h]}"
            else:
                seen[h] = 0
            out.append(h)
        return out

    # ---------- A) TEXT-PDF ----------
    frames = []
    try:
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables(table_settings={
                    "vertical_strategy": "lines",
                    "horizontal_strategy": "lines",
                    "intersection_tolerance": 5,
                }) or []
                if not tables:
                    tables = page.extract_tables(table_settings={
                        "vertical_strategy": "text",
                        "horizontal_strategy": "text",
                        "snap_tolerance": 3,
                        "intersection_tolerance": 3,
                        "join_tolerance": 3,
                    }) or []

                for t in tables:
                    rows = [r for r in t if any((c or "").strip() for c in r)]
                    if len(rows) < 2:
                        continue
                    w = max(len(r) for r in rows)
                    rows = [list(r) + [""] * (w - len(r)) for r in rows]

                    header = unique_headers([(h or "").strip() for h in rows[0]])
                    df = pd.DataFrame(rows[1:], columns=header)

                    for j in range(df.shape[1]):
                        s = df.iloc[:, j].astype(str)
                        s = s.str.replace("\xa0", " ", regex=False).str.strip()
                        s = s.mask(s.str.lower().isin(["nan", "none", "null"]), "")
                        df.iloc[:, j] = s

                    df = df.replace("", np.nan).dropna(how="all").fillna("")
                    if not df.empty:
                        frames.append(df)
    except Exception as e:
        log(f"[PDF] text-parse failed: {e}")

    if frames:
        out = pd.concat(frames, ignore_index=True)
        log(f"[PDF] text tables found, shape={out.shape}")
        return out

    # ---------- B) OCR ----------
    if not try_ocr:
        log("[OCR] try_ocr=False → skipping OCR")
        return pd.DataFrame()

    from pdf2image import convert_from_path
    import pytesseract

    poppler_bin = current_app.config.get("POPPLER_BIN") or getattr(
        Config, "POPPLER_BIN", r"C:\Program Files\poppler\bin"
    )
    tess_exe = current_app.config.get("TESSERACT_EXE") or getattr(
        Config, "TESSERACT_EXE", r"C:\Program Files\Tesseract-OCR\tesseract.exe"
    )
    pytesseract.pytesseract.tesseract_cmd = tess_exe

    try:
        images = convert_from_path(path, dpi=400, poppler_path=poppler_bin)
        log(f"[OCR] pages={len(images)} poppler_bin={poppler_bin}")
    except Exception as e:
        log(f"[OCR] convert_from_path failed: {e}")
        # вернём спец-таблицу, чтобы в превью видно было причину
        return pd.DataFrame({"error": [f"OCR init failed: {e}"]})

    tesseract_cfg = r"--oem 3 --psm 6"
    lines: list[str] = []
    for idx, img in enumerate(images, start=1):
        try:
            text = pytesseract.image_to_string(img, lang="eng", config=tesseract_cfg)
        except Exception as e:
            log(f"[OCR] page {idx} failed: {e}")
            continue
        for ln in text.splitlines():
            s = (ln or "").strip()
            if s:
                lines.append(s)

    # Dump raw OCR text (удобно проверять глазами)
    dump_path = os.path.splitext(path)[0] + ".ocr.txt"
    try:
        with open(dump_path, "w", encoding="utf-8") as fh:
            fh.write("\n".join(lines))
        log(f"[OCR] lines={len(lines)} dump={dump_path}")
    except Exception:
        pass

    # ---- C) Try parsing Marcone-style rows
    base = os.path.basename(path).lower()
    supplier_guess = "Marcone" if "marcone" in base else ("ReliableParts" if "reliable" in base else "")

    money = r"\$[\d,]+\.\d{2}"
    pat_marcone = re.compile(
        rf"""
        ^\s*
        (?P<qty>\d+)\s+                      # qty
        (?P<pn>[A-Za-z0-9\-\/\.]+)\s+        # part number
        (?P<descr>.+?)\s+                    # description
        (?P<unit>{money})\s+                 # unit price
        (?:{money}\s+)?                      # optional MSRP/total
        (?P<price>{money})\s+                # total (ignored)
        (?P<bo>\d+)\s+(?P<ship>\d+)          # backorder / shipped (ignored)
        \s*$""",
        re.VERBOSE,
    )

    rows = []
    for ln in lines:
        m = pat_marcone.match(ln)
        if not m:
            continue
        rows.append({
            "PART #": m.group("pn").strip(),
            "DESCR.": m.group("descr").strip(),
            "QTY": int(m.group("qty")),
            "UNIT COST": float(m.group("unit").replace("$", "").replace(",", "")),
            "SUPPLIER": supplier_guess or "Marcone",
        })

    if rows:
        df = pd.DataFrame(rows)
        for c in ["PART #", "DESCR.", "SUPPLIER"]:
            df[c] = df[c].astype(str).str.replace("\xa0", " ", regex=False).str.strip()
        log(f"[OCR-MARCONE] parsed rows={len(df)}")
        return df

    # ---- D) Generic heuristic: "<qty> <pn> <descr...> <price$>" ИЛИ "<pn> <descr...> <qty> <price$>"
    pat1 = re.compile(rf"^\s*(?P<qty>\d+)\s+(?P<pn>[A-Za-z0-9\-\/\.]+)\s+(?P<descr>.+?)\s+(?P<unit>{money})\s*$")
    pat2 = re.compile(rf"^\s*(?P<pn>[A-Za-z0-9\-\/\.]+)\s+(?P<descr>.+?)\s+(?P<qty>\d+)\s+(?P<unit>{money})\s*$")

    rows = []
    for ln in lines:
        m = pat1.match(ln) or pat2.match(ln)
        if not m:
            continue
        qty = int(m.group("qty"))
        pn = m.group("pn").strip()
        descr = m.group("descr").strip()
        unit = float(m.group("unit").replace("$", "").replace(",", ""))
        rows.append({
            "PART #": pn,
            "DESCR.": descr,
            "QTY": qty,
            "UNIT COST": unit,
            "SUPPLIER": supplier_guess or "OCR",
        })

    if rows:
        df = pd.DataFrame(rows)
        for c in ["PART #", "DESCR.", "SUPPLIER"]:
            df[c] = df[c].astype(str).str.replace("\xa0", " ", regex=False).str.strip()
        log(f"[OCR-GENERIC] parsed rows={len(df)}")
        return df

    # ---- E) Last resort: show raw text so preview is not empty
    if lines:
        log("[OCR] fallback to raw text table")
        return pd.DataFrame({"text": lines})

    log("[OCR] no lines recognized")
    return pd.DataFrame()

def _settings_path():
    return os.path.join(current_app.instance_path, "app_settings.json")

def get_setting(key, default=None):
    try:
        with open(_settings_path(), "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception:
        data = {}
    return data.get(key, default)

def set_setting(key, value):
    p = _settings_path()
    os.makedirs(os.path.dirname(p), exist_ok=True)
    try:
        with open(p, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception:
        data = {}
    data[key] = value
    with open(p, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def _recompute_batch_consumption(batch: "IssuedBatch") -> None:
    """
    Устанавливает batch.consumed_flag = True, если все строки в батче полностью списаны (или qty<=0),
    иначе False. Также обновляет consumed_at/by по последнему действию.
    """
    if not batch:
        return
    # Считаем сумму позитивных qty и сумму consumed по всем строкам
    total_qty = 0
    total_used = 0
    last_when = None
    last_who  = None

    for it in (batch.parts or []):
        q = int(it.quantity or 0)
        used = int(it.consumed_qty or 0)
        if q > 0:
            total_qty  += q
            total_used += min(q, used)
        # возьмём самое позднее consumed_at
        if it.consumed_at and (last_when is None or it.consumed_at > last_when):
            last_when = it.consumed_at
            last_who  = it.consumed_by

    fully = (total_qty > 0 and total_used >= total_qty)
    batch.consumed_flag = bool(fully)
    batch.consumed_at   = last_when
    batch.consumed_by   = last_who

@inventory_bp.post("/reports/consume/reset")
@login_required
def unconsume_invoice():
    """
    Частично снимает consumed_* у выбранных строк инвойса, либо у всего инвойса (по группе).
    Снимает НЕ всё, а ровно qty_<id> для каждой строки (как обратная операция к consume_invoice).
    Доступ: только superadmin.
    """
    from flask_login import current_user
    role = (getattr(current_user, "role", "") or "").strip().lower()
    # только супер-админ
    if role not in ("superadmin", "user"):
        flash("Access denied (unconsume is superadmin only).", "danger")
        return redirect(url_for("inventory.reports_grouped"))

    from sqlalchemy import func
    from extensions import db
    from models import IssuedPartRecord, IssuedBatch, IssuedConsumptionLog

    form = request.form

    # Группа (для режима "Entire invoice")
    issued_to     = (form.get("group_issued_to") or "").strip()
    reference_job = (form.get("group_reference_job") or "").strip()
    issued_by     = (form.get("group_issued_by") or "").strip()
    issue_date_s  = (form.get("group_issue_date") or "").strip()
    location      = (form.get("location") or "").strip()

    invoice_number = form.get("invoice_number", type=int)

    apply_scope = (form.get("apply_scope") or "all").strip().lower()  # all|selected
    selected_ids = [int(x) for x in form.getlist("record_ids[]") if str(x).isdigit()]

    # Базовый запрос
    q = IssuedPartRecord.query

    if invoice_number:
        q = q.filter(IssuedPartRecord.invoice_number == invoice_number)
    else:
        # супер-точное совпадение группы
        q = q.filter(
            IssuedPartRecord.issued_to == issued_to,
            IssuedPartRecord.issued_by == issued_by,
            IssuedPartRecord.reference_job == (reference_job or None),
            func.strftime("%Y-%m-%d %H:%M:%S", IssuedPartRecord.issue_date) == issue_date_s
        )
        if location:
            q = q.filter(func.coalesce(IssuedPartRecord.location, Part.location) == location)

    # Скоуп: только выбранные строки или весь инвойс
    rows = q.all()
    if apply_scope == "selected" and selected_ids:
        sel = set(selected_ids)
        rows = [r for r in rows if r.id in sel]

    if not rows:
        flash("No rows matched for unconsume.", "warning")
        return redirect(url_for("inventory.reports_grouped"))

    changed = 0
    touched_batches = set()

    for r in rows:
        # игнорируем возвраты
        if (r.quantity or 0) <= 0:
            continue

        used = int(r.consumed_qty or 0)
        if used <= 0:
            continue

        # сколько снять — аналогично consume_invoice: qty_<id>, по умолчанию 1
        try:
            dec_qty = int(form.get(f"qty_{r.id}", "1"))
        except Exception:
            dec_qty = 1

        if dec_qty <= 0:
            continue

        # не снимаем больше, чем уже списано
        dec_qty = min(dec_qty, used)
        new_used = used - dec_qty

        # обновляем consumed_* поля
        r.consumed_qty = new_used if new_used > 0 else None
        # если всё сняли — чистим метаданные
        if new_used <= 0:
            r.consumed_flag = False
            r.consumed_at = None
            r.consumed_by = None
            r.consumed_note = None
        else:
            # просто синхронизируем флаг
            r.consumed_flag = (new_used >= int(r.quantity or 0))

        # правим логи IssuedConsumptionLog (LIFO — последние списания снимаем первыми)
        logs = (IssuedConsumptionLog.query
                .filter_by(issued_part_id=r.id)
                .order_by(IssuedConsumptionLog.id.desc())
                .all())

        to_remove = dec_qty
        for log in logs:
            if to_remove <= 0:
                break

            log_qty = int(log.qty or 0)
            if log_qty <= 0:
                continue

            if log_qty <= to_remove:
                # полностью убираем лог
                to_remove -= log_qty
                db.session.delete(log)
            else:
                # уменьшаем qty в логе
                log.qty = log_qty - to_remove
                to_remove = 0
                db.session.add(log)

        changed += 1
        if r.batch_id:
            touched_batches.add(r.batch_id)
        db.session.add(r)

    # Пересчёт consumed_flag на батчах
    if touched_batches:
        batches = IssuedBatch.query.filter(IssuedBatch.id.in_(list(touched_batches))).all()
        for b in batches:
            # если у тебя есть _recompute_batch_consumption(b) — оставляем
            try:
                _recompute_batch_consumption(b)
            except NameError:
                # если его нет, но есть _sync_batches_consumed_flag — можно вызвать её вместо
                pass

    db.session.commit()
    flash(f"Unconsumed {changed} row(s).", "success")
    return redirect(url_for("inventory.reports_grouped"))

# ====== CONSUME (partial usage) ===============================================
@inventory_bp.post("/invoices/consume")
@login_required
def consume_invoice():
    """
    Увеличивает consumed_qty по выбранным строкам (или по всему инвойсу — по radio apply_scope).
    Использует значения qty_<id> из формы. Никогда не превышает исходный quantity.
    Игнорирует возвраты (quantity < 0).
    Доступно для admin/superadmin/user.
    """
    from flask_login import current_user
    from sqlalchemy import func
    from models import IssuedPartRecord, IssuedBatch, IssuedConsumptionLog

    role = (getattr(current_user, "role", "") or "").strip().lower()
    if role not in ("admin", "superadmin", "user"):
        flash("Access denied", "danger")
        return redirect(url_for("inventory.reports_grouped"))

    form = request.form

    # ключи группы (для режима 'Entire invoice')
    group_issued_to     = (form.get("group_issued_to") or "").strip()
    group_reference_job = (form.get("group_reference_job") or "").strip()
    group_issued_by     = (form.get("group_issued_by") or "").strip()
    group_issue_date_s  = (form.get("group_issue_date") or "").strip()
    location            = (form.get("location") or "").strip()

    # ⚠️ Job ОБЯЗАТЕЛЕН
    job_ref_raw = (form.get("job_ref") or "").strip()
    if not job_ref_raw:
        flash("Job number is required to mark parts as consumed.", "warning")
        return redirect(url_for("inventory.reports_grouped"))
    job_ref = job_ref_raw

    # радиокнопка области: all | selected
    apply_scope = (form.get("apply_scope") or "all").strip().lower()
    # список выбранных записей для режима 'selected'
    selected_ids = [int(x) for x in form.getlist("record_ids[]") if str(x).isdigit()]

    # Собираем кандидатные строки
    q = db.session.query(IssuedPartRecord).options(
        db.joinedload(IssuedPartRecord.part),
        db.joinedload(IssuedPartRecord.batch),
    )

    if apply_scope == "selected" and selected_ids:
        q = q.filter(IssuedPartRecord.id.in_(selected_ids))
    else:
        # Entire invoice (legacy-группа по ключам)
        try:
            grp_dt = datetime.strptime(group_issue_date_s, "%Y-%m-%d %H:%M:%S")
        except Exception:
            grp_dt = None

        if grp_dt is None:
            flash("Invalid group date for invoice consume.", "warning")
            return redirect(url_for("inventory.reports_grouped"))

        q = q.filter(
            func.trim(IssuedPartRecord.issued_to) == group_issued_to,
            func.trim(IssuedPartRecord.issued_by) == group_issued_by,
            func.date(IssuedPartRecord.issue_date) == grp_dt.date(),
        )
        # reference_job может быть пустым
        if group_reference_job:
            q = q.filter(func.trim(IssuedPartRecord.reference_job) == group_reference_job)
        else:
            q = q.filter(func.coalesce(IssuedPartRecord.reference_job, "") == "")

        # Если в карточке был location — берём его как scope.
        if location:
            q = q.filter(func.trim(IssuedPartRecord.location) == location)

    rows = q.all()
    if not rows:
        flash("No rows matched for consumption.", "warning")
        return redirect(url_for("inventory.reports_grouped"))

    changed = 0
    now_user = (getattr(current_user, "username", "") or "").strip()

    for r in rows:
        # Игнорируем возвраты
        if (r.quantity or 0) <= 0:
            continue

        # Сколько увеличить — берём qty_<id> из формы, по умолчанию 1
        try:
            add_qty = int(form.get(f"qty_{r.id}", "1"))
        except Exception:
            add_qty = 1
        if add_qty <= 0:
            continue

        # ограничим по оставшемуся
        q_total = int(r.quantity or 0)
        used    = int(r.consumed_qty or 0)
        remain  = max(0, q_total - used)
        if remain <= 0:
            continue

        add_qty = min(add_qty, remain)

        # Обновляем агрегат в IssuedPartRecord
        if r.apply_consume(add_qty, user=now_user, note="consume via reports_grouped"):
            changed += 1
            db.session.add(r)

            # Логируем это частичное списание в IssuedConsumptionLog
            log = IssuedConsumptionLog(
                issued_part_id=r.id,
                qty=add_qty,
                job_ref=job_ref,
                consumed_by=now_user,
                note="consume via reports_grouped",
            )
            db.session.add(log)

    if changed:
        # Пересчёт consumed_flag на батчах (как в unconsume_invoice)
        touched_batches = set()
        for r in rows:
            if r.batch_id:
                touched_batches.add(r.batch_id)

        if touched_batches:
            batches = IssuedBatch.query.filter(
                IssuedBatch.id.in_(list(touched_batches))
            ).all()
            for b in batches:
                _recompute_batch_consumption(b)

        db.session.commit()
        flash(f"Marked consumed on {changed} row(s).", "success")
    else:
        flash("Nothing to consume (all selected rows already fully consumed).", "info")

    return redirect(url_for("inventory.reports_grouped"))

# --- Confirm selected lines / whole invoice ---
@inventory_bp.post("/work_orders/<int:wo_id>/confirm")
@login_required
def wo_confirm_lines(wo_id: int):
    from flask import request, flash, redirect, url_for
    from datetime import datetime
    from sqlalchemy import func, or_
    from flask_login import current_user
    from extensions import db
    from models import WorkOrder, IssuedPartRecord, IssuedBatch

    wo = db.session.get(WorkOrder, wo_id)
    if not wo:
        flash(f"Work Order #{wo_id} not found.", "danger")
        return redirect(url_for("inventory.wo_list"))

    role = (current_user.role or "").lower()
    is_admin_like = role in ("admin", "superadmin")

    # проверка «свой ли WO» для техника
    if role == "technician":
        me_id = getattr(current_user, "id", None)
        me_name = (current_user.username or "").strip().lower()
        wo_tech_id = getattr(wo, "technician_id", None)
        wo_name = (wo.technician_username or wo.technician_name or "").strip().lower()
        is_my_wo = (
            (wo_tech_id and me_id and wo_tech_id == me_id)
            or (me_name and wo_name and me_name == wo_name)
        )
        if not is_my_wo:
            flash("Access denied", "danger")
            return redirect(url_for("inventory.wo_detail", wo_id=wo_id))

    # алиасы имени техника — чтобы совпало с issued_to
    me_aliases = {
        (current_user.username or "").strip().lower(),
        (wo.technician_username or "").strip().lower(),
        (wo.technician_name or "").strip().lower(),
    }
    me_aliases = {a for a in me_aliases if a}

    tech_display = (current_user.username or "").strip()
    if not tech_display:
        flash("Your account has no username; cannot confirm.", "danger")
        return redirect(url_for("inventory.wo_detail", wo_id=wo_id))

    canonical_job = (wo.canonical_job or "").strip()

    raw_ids = request.form.getlist("record_ids[]") or request.form.getlist("record_ids")
    sel_ids = [int(x) for x in raw_ids if str(x).isdigit()]
    inv_s = (request.form.get("invoice_number") or "").strip()
    inv_no = int(inv_s) if inv_s.isdigit() else None

    q = db.session.query(IssuedPartRecord).outerjoin(
        IssuedBatch, IssuedBatch.id == IssuedPartRecord.batch_id
    )

    if canonical_job:
        q = q.filter(
            or_(
                func.trim(IssuedPartRecord.reference_job) == canonical_job,
                func.trim(IssuedPartRecord.reference_job).like(f"%{canonical_job}%"),
                func.trim(IssuedBatch.reference_job) == canonical_job,
            )
        )

    if sel_ids:
        q = q.filter(IssuedPartRecord.id.in_(sel_ids))
    elif inv_no is not None:
        q = q.filter(IssuedPartRecord.invoice_number == inv_no)
    else:
        flash("Nothing to confirm (no IDs or invoice number).", "warning")
        return redirect(url_for("inventory.wo_detail", wo_id=wo_id))

    if not is_admin_like and me_aliases:
        # техник подтверждает только свои строки (без регистра)
        q = q.filter(func.lower(func.trim(IssuedPartRecord.issued_to)).in_(list(me_aliases)))

    rows = q.all()
    if not rows:
        flash("No matching lines found to confirm.", "warning")
        return redirect(url_for("inventory.wo_detail", wo_id=wo_id))

    now = datetime.utcnow()
    updated = 0
    for r in rows:
        if not getattr(r, "confirmed_by_tech", False):
            r.confirmed_by_tech = True
            r.confirmed_at = now
            r.confirmed_by = tech_display
            updated += 1

    try:
        db.session.commit()
        msg = f"Confirmed {updated} line(s)"
        if inv_no is not None:
            msg += f" in invoice #{inv_no:06d}"
        flash(msg + ("" if updated else " (no changes)."), "success" if updated else "info")
    except Exception as e:
        db.session.rollback()
        flash(f"Database error: {e}", "danger")

    return redirect(url_for("inventory.wo_detail", wo_id=wo_id))

@inventory_bp.get("/api/job_duplicate_check", endpoint="api_job_duplicate_check")
@login_required
def api_job_duplicate_check():
    """
    Проверяет, пересекается ли любой из переданных job numbers с уже существующими Work Orders.
    Параметры:
      - job_numbers: строка с номерами через запятую/пробел/; (например '984891, 989898')
      - current_wo_id (опц.): id текущей WO — чтобы не считать её же дублем при редактировании

    Ответ:
      { "duplicate": true, "existing_id": 123, "existing_jobs": "984891, 989898" }
      или
      { "duplicate": false }
    """
    from flask import request, jsonify
    from flask_login import current_user
    from sqlalchemy import or_
    from models import WorkOrder
    import re

    role = (getattr(current_user, "role", "") or "").strip().lower()
    if role not in ("admin", "superadmin"):
        return jsonify({"duplicate": False})

    raw_jobs = (request.args.get("job_numbers") or "").strip()
    current_wo_id = (request.args.get("current_wo_id") or "").strip()

    def _parse_jobs(raw: str) -> list[str]:
        s = (raw or "").upper()
        parts = re.split(r"[,\s;]+", s)
        cleaned, seen = [], set()
        for p in parts:
            p = p.strip()
            if not p:
                continue
            if p not in seen:
                seen.add(p)
                cleaned.append(p)
        return cleaned

    def _parse_jobs_from_wo(wo: WorkOrder) -> set[str]:
        return set(_parse_jobs(wo.job_numbers or ""))

    jobs_list = _parse_jobs(raw_jobs)
    if not jobs_list:
        return jsonify({"duplicate": False})
    jobs_set = set(jobs_list)

    # Узкое SQL-фильтрование по LIKE, чтобы не грузить все WOs
    like_filters = [WorkOrder.job_numbers.ilike(f"%{j}%") for j in jobs_list]
    q = WorkOrder.query.filter(or_(*like_filters))
    if current_wo_id.isdigit():
        q = q.filter(WorkOrder.id != int(current_wo_id))

    possibles = q.order_by(WorkOrder.created_at.desc()).limit(50).all()

    for wo in possibles:
        exist_set = _parse_jobs_from_wo(wo)
        overlap = sorted(exist_set.intersection(jobs_set))
        if overlap:
            return jsonify({
                "duplicate": True,
                "existing_id": wo.id,
                "existing_jobs": wo.job_numbers or ", ".join(overlap),
            })

    return jsonify({"duplicate": False})

@inventory_bp.get("/work_orders/new", endpoint="wo_new")
@login_required
def wo_new():
    role = (getattr(current_user, "role", "") or "").strip().lower()
    if role not in ("admin", "superadmin"):
        flash("Access denied", "danger")
        return redirect(url_for("inventory.wo_list"))

    # минимальный "пустой" объект, чтобы шаблон мог отрисовать поля
    class _WO:
        id = None
        technician_id = None
        technician_name = ""
        brand = ""
        model = ""
        serial = ""
        job_numbers = ""
        job_type = "BASE"
        delivery_fee = 0.0
        markup_percent = 0.0
        status = "search_ordered"

    wo = _WO()

    technicians = _query_technicians()
    recent_suppliers = session.get("recent_suppliers", []) or []

    # один пустой unit с одной пустой строкой parts для формы
    units = [{
        "brand":  "",
        "model":  "",
        "serial": "",
        "rows": [{
            "id": None,
            "part_number": "",
            "part_name": "",
            "quantity": 1,
            "alt_numbers": "",
            "warehouse": "",
            "supplier": "",
            "backorder_flag": False,
            "line_status": "search_ordered",
            "unit_cost": 0.0,
        }],
    }]

    return render_template(
        "wo_form_units.html",
        wo=wo,
        units=units,
        recent_suppliers=recent_suppliers,
        readonly=False,
        technicians=technicians,
        selected_tech_id=None,
        selected_tech_username=None,
    )

@inventory_bp.get("/api/technicians")
@login_required
def api_technicians():
    from models import WorkOrder
    q = (request.args.get("q") or "").strip().upper()

    qry = (db.session.query(WorkOrder.technician_name)
           .filter(WorkOrder.technician_name.isnot(None),
                   WorkOrder.technician_name != ""))

    if q:
        qry = qry.filter(db.func.upper(WorkOrder.technician_name).like(q + "%"))

    names = sorted({(r[0] or "").strip().upper() for r in qry.limit(25).all() if r[0]})
    return jsonify({"items": names}), 200


@inventory_bp.get("/debug/db")
@login_required
def debug_db():
    import sqlite3, json, os
    from flask import jsonify, current_app
    dbp = os.path.join(current_app.instance_path, "inventory.db")
    dbp = os.path.normpath(dbp)
    views = []
    try:
        with sqlite3.connect(dbp) as con:
            views = [r[0] for r in con.execute("SELECT name FROM sqlite_master WHERE type IN ('view','table')").fetchall()]
    except Exception as e:
        return jsonify({"db_path": dbp, "error": str(e)})
    return jsonify({"db_path": dbp, "objects": views})



@inventory_bp.get("/search")
@login_required
def search():
    q = (request.args.get("q") or "").strip()
    status = (request.args.get("status") or "").strip()
    supplier = (request.args.get("supplier") or "").strip()
    scope = (request.args.get("scope") or "global").strip()
    wo_id = request.args.get("wo_id", type=int)
    limit = min(max(request.args.get("limit", type=int) or 50, 1), 200)
    offset = max(0, request.args.get("offset", type=int) or 0)

    if scope == "order" and not wo_id:
        abort(400, "wo_id required when scope=order")

    crit = _parse_q(q)
    where, params = [], {"limit":limit, "offset":offset}

    if scope == "order": where.append("wo_id=:wo_id"); params["wo_id"]=wo_id
    if status:           where.append("status=:status"); params["status"]=status
    if supplier:         where.append("supplier=:supplier COLLATE NOCASE"); params["supplier"]=supplier

    if crit["type"]=="text" and crit["value"]:
        params["v"]=f"%{crit['value']}%"
        where.append("(pn LIKE :v OR part_name LIKE :v OR brand LIKE :v OR model LIKE :v OR serial LIKE :v)")
    elif crit["type"]=="pn":
        params["pn"]=crit["value"].replace("-","").upper()+"%"
        where.append("UPPER(REPLACE(pn,'-','')) LIKE :pn")
    elif crit["type"] in {"brand","model","serial"}:
        params["v"]=f"%{crit['value']}%"; where.append(f"{crit['type']} LIKE :v")
    elif crit["type"] in {"wo","job"}:
        try: params["qwo"]=int(crit["value"]); where.append("wo_id=:qwo")
        except ValueError: pass

    where_sql=" WHERE "+ " AND ".join(where) if where else ""
    sql=f"""
      SELECT id, wo_id, pn, part_name, brand, model, serial,
             qty_needed, qty_issued, to_issue, unit_cost,
             on_hand, ordered_qty, eta, location, supplier, status
      FROM v_work_order_parts_dto
      {where_sql}
      ORDER BY wo_id DESC, brand, model, serial, pn
      LIMIT :limit OFFSET :offset
    """

    with sqlite3.connect(_db_path()) as con:
        con.row_factory = sqlite3.Row
        rows = con.execute(sql, params).fetchall()

    return jsonify({"items":[dict(r) for r in rows],
                    "limit":limit,"offset":offset,"scope":scope,"wo_id":wo_id})

@inventory_bp.get("/inventory/search", endpoint="search_alias")
@login_required
def search_alias():
    # просто проксируем на новый обработчик, сохраняя query params
    return search()

# --- toggle "Ordered" for a single WorkOrderPart ---
@inventory_bp.post("/work_orders/<int:wo_id>/parts/<int:wop_id>/toggle_ordered", endpoint="wo_toggle_ordered")
@login_required
def wo_toggle_ordered(wo_id: int, wop_id: int):
    from flask import request, flash, redirect, url_for
    from datetime import date
    from extensions import db
    from models import WorkOrderPart, WorkOrder
    from flask_login import current_user

    wop = db.session.get(WorkOrderPart, wop_id)
    if not wop or getattr(wop, "work_order_id", None) != wo_id:
        flash("Part row not found for this Work Order.", "danger")
        return redirect(url_for("inventory.wo_detail", wo_id=wo_id))

    wo = db.session.get(WorkOrder, wo_id)
    me   = (getattr(current_user, "username", "") or "").strip().lower()
    role = (getattr(current_user, "role", "") or "").strip().lower()
    tech = (getattr(wo, "technician_name", "") or "").strip().lower()

    if not (role in ("admin", "superadmin") or (me and me == tech)):
        flash("You are not allowed to change ordered status for this WO.", "danger")
        return redirect(url_for("inventory.wo_detail", wo_id=wo_id))

    new_state = (request.form.get("state") or "0").strip().lower() in ("1","true","on","yes")

    try:
        # sync флаги/статусы
        if hasattr(wop, "ordered_flag"):
            wop.ordered_flag = bool(new_state)
        if hasattr(wop, "status"):
            wop.status = "ordered" if new_state else "search_ordered"
        if hasattr(wop, "line_status"):
            wop.line_status = "ordered" if new_state else "search_ordered"

        # дата:
        if hasattr(wop, "ordered_date"):
            if new_state:
                # если даты нет — поставим сегодня; если есть — не трогаем
                if not wop.ordered_date:
                    wop.ordered_date = date.today()
            else:
                # снятие чекбокса — чистим дату
                wop.ordered_date = None

        db.session.commit()
        flash(("Ordered set" if new_state else "Ordered cleared") + " for the part.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Failed to update: {e}", "danger")

    return redirect(url_for("inventory.wo_detail", wo_id=wo_id))

def _auto_assign_invoice_for_wo(wo, current_user):
    """
    Находит «свежие» строки выдачи для этого WO без invoice_number/batch
    и группами присваивает им номер через _create_batch_for_records(...).
    Возвращает: количество созданных батчей.
    """
    from extensions import db
    from models import IssuedPartRecord

    canon = (wo.canonical_job or "").strip()
    if not canon:
        return 0

    # Нормализуем имя техника (куда выдавали)
    issued_to = (wo.technician_username or wo.technician_name or "").strip()
    if not issued_to:
        return 0

    # Подскоп «свежести»: за последний день (можно сузить до сегодняшней даты)
    now = datetime.utcnow()
    start = datetime.combine(now.date(), _time.min)
    end   = datetime.combine(now.date(), _time.max)

    # Кандидаты: строки по этому рефу, без номера и без батча, за сегодня
    q = (
        db.session.query(IssuedPartRecord)
        .filter(
            or_(
                func.trim(IssuedPartRecord.reference_job) == canon,
                func.trim(IssuedPartRecord.reference_job).like(f"%{canon}%"),
            ),
            or_(IssuedPartRecord.invoice_number.is_(None),
                IssuedPartRecord.invoice_number == 0),
            IssuedPartRecord.batch_id.is_(None),
            IssuedPartRecord.issue_date.between(start, end),
        )
        .order_by(IssuedPartRecord.id.asc())
    )
    cand = q.all()
    if not cand:
        return 0

    # Фильтруем реально «наши» строки (на случай общих рефов)
    # Чаще всего issued_to совпадает с техником WO:
    same_to = [r for r in cand if (r.issued_to or "").strip().lower() == issued_to.strip().lower()]
    if not same_to:
        # если ничего не нашли по issued_to — не трогаем (чтобы не схватить чужие)
        return 0

    # Группировка по (issued_to, issued_by, reference_job, дата)
    def _key(r):
        d = r.issue_date.date() if r.issue_date else now.date()
        return (
            (r.issued_to or issued_to).strip(),
            (r.issued_by or (getattr(current_user, "username", "") or "system")).strip(),
            (r.reference_job or canon).strip(),
            d,
            (r.location or None),
        )

    groups = {}
    for r in same_to:
        groups.setdefault(_key(r), []).append(r)

    created = 0
    for (issued_to_k, issued_by_k, ref_k, day_k, location_k), records in groups.items():
        try:
            issue_dt = datetime.combine(day_k, _time.min)
            _create_batch_for_records(
                records=records,
                issued_to=issued_to_k,
                issued_by=issued_by_k or "system",
                reference_job=ref_k,
                issue_date=issue_dt,
                location=location_k,
            )
            created += 1
        except Exception:
            db.session.rollback()
            continue

    if created:
        from extensions import db
        try:
            db.session.commit()
        except Exception:
            db.session.rollback()
            created = 0

    return created

# --- helper: build 'batches' payload for wo_detail (adds issued_at_dt) ---
from datetime import datetime

def _serialize_batches_for_wo_detail(db_batches, wo):
    """
    db_batches: iterable IssuedBatch (или твоя структура), у которой есть:
      - id, technician (имя/username), canonical_ref/reference_job, invoice_number, issued_at
      - items/records: список IssuedPartRecord (или эквивалент) с полями:
          id, part (obj) или part_id, quantity, unit_cost_at_issue, issue_date, confirmed
    wo: текущий WorkOrder (для фолбэков на canonical_job/technician_name)

    Возвращает список словарей для Jinja:
      keys: is_return, items, total_value, technician, canonical_ref, reference_job,
            invoice_number, issued_at_dt  (+ оставим issued_at для бэк-совместимости)
    """
    result = []
    for b in db_batches:
        # --- безопасно собираем список записей (IssuedPartRecord)
        recs = getattr(b, "records", None) or getattr(b, "items", None) or []

        items = []
        total_value = 0.0
        any_confirmed = False
        all_confirmed = True if recs else False
        is_return_batch = False

        # Определяем issued_at_dt: приоритет — поле батча; иначе по первой записи
        issued_at_dt = getattr(b, "issued_at", None)
        if not issued_at_dt and recs:
            # возьмём минимальную дату записи как «время батча»
            issued_at_dt = min((getattr(r, "issue_date", None) for r in recs if getattr(r, "issue_date", None)), default=None)

        # Собираем айтемы
        for r in recs:
            qty = getattr(r, "quantity", 0) or 0
            unit = getattr(r, "unit_cost_at_issue", None)
            name = None
            pn = None

            part_obj = getattr(r, "part", None)
            if part_obj is not None:
                pn = getattr(part_obj, "part_number", None) or getattr(part_obj, "id", None) or getattr(r, "part_id", None)
                name = getattr(part_obj, "name", None) or ""
            else:
                pn = getattr(r, "part_id", None)
                name = ""

            line_val = (unit or 0.0) * qty
            total_value += line_val

            if qty < 0:
                is_return_batch = True

            confirmed = bool(getattr(r, "confirmed", False))
            any_confirmed = any_confirmed or confirmed
            if not confirmed:
                all_confirmed = False

            items.append({
                "id": getattr(r, "id", None),
                "pn": pn or "",
                "name": name or "",
                "qty": qty,
                "unit_price": unit,
                "confirmed": confirmed,
                "negative": qty < 0,
            })

        # Фолбэки по технику/референсу
        tech = getattr(b, "technician", None) or getattr(wo, "technician_name", None) or ""
        canonical_ref = getattr(b, "canonical_ref", None) or getattr(b, "reference_job", None) or getattr(wo, "canonical_job", None) or ""
        reference_job = getattr(b, "reference_job", None) or getattr(wo, "canonical_job", None) or ""

        result.append({
            "is_return": is_return_batch,
            "items": items,
            "total_value": float(total_value),
            "technician": tech,
            "canonical_ref": canonical_ref,
            "reference_job": reference_job,
            "invoice_number": getattr(b, "invoice_number", None),

            # ключ для нового шаблона (будет отрендерен через |local_dt):
            "issued_at_dt": issued_at_dt,

            # оставим старый, если где-то ещё используется:
            "issued_at": getattr(b, "issued_at", None) or (
                issued_at_dt.strftime("%Y-%m-%d %H:%M") if isinstance(issued_at_dt, datetime) else None
            ),
            # для совместимости можно вернуть и «issued_by», если нужно в legacy-печатях
            "issued_by": getattr(b, "issued_by", None),
        })
    return result


# --- Work Order details ---
@inventory_bp.get("/work_orders/<int:wo_id>", endpoint="wo_detail")
@login_required
def wo_detail(wo_id):
    from flask import render_template, flash, redirect, url_for, session
    from sqlalchemy import func, or_, and_, case
    from sqlalchemy.orm import selectinload, joinedload
    from flask_login import current_user
    from extensions import db
    from models import WorkOrder, WorkUnit, WorkOrderPart, Part, IssuedPartRecord, IssuedBatch
    from collections import defaultdict, defaultdict as _dd
    import re

    # 1) load WO
    wo = (
        db.session.query(WorkOrder)
        .options(
            selectinload(WorkOrder.parts),
            selectinload(WorkOrder.units).options(selectinload(WorkUnit.parts)),
        )
        .get(wo_id)
    )
    if not wo:
        flash(f"Work Order #{wo_id} not found.", "danger")
        return redirect(url_for("inventory.wo_list"))

    # 2) access
    role_raw = (getattr(current_user, "role", "") or "").strip().lower()
    me_id = getattr(current_user, "id", None)
    me_name = (getattr(current_user, "username", "") or "").strip().lower()

    wo_tech_id = getattr(wo, "technician_id", None)
    wo_tech_name = (wo.technician_username or wo.technician_name or "").strip().lower()

    is_admin_like = role_raw in ("admin", "superadmin")
    is_technician = role_raw in ("technician", "tech")
    is_user = (role_raw == "user")

    is_my_wo = False
    if wo_tech_id and me_id and wo_tech_id == me_id:
        is_my_wo = True
    elif wo_tech_name and me_name and wo_tech_name == me_name:
        is_my_wo = True

    if is_technician and not is_my_wo:
        flash("You don't have access to this Work Order.", "danger")
        return redirect(url_for("inventory.wo_list"))

    can_confirm_any = (role_raw == "superadmin")
    can_view_docs = is_admin_like or is_my_wo or is_user

    # 3) suppliers
    suppliers = [
        r[0]
        for r in (
            db.session.query(func.trim(WorkOrderPart.supplier))
            .filter(
                WorkOrderPart.work_order_id == wo.id,
                WorkOrderPart.supplier.isnot(None),
                func.trim(WorkOrderPart.supplier) != "",
            )
            .distinct()
            .order_by(func.trim(WorkOrderPart.supplier).asc())
            .all()
        )
    ]

    # ------------------------------------------------------------
    # 4) Issued / Batches — FIXED (work_order_id + safe fallback)
    # ------------------------------------------------------------
    canon = (wo.canonical_job or "").strip()
    raw_jobs = (getattr(wo, "job_numbers", "") or "").strip()

    job_tokens = []
    for t in re.findall(r"\d+[A-Za-z]?", raw_jobs):
        tt = (t or "").strip().upper()
        if tt:
            job_tokens.append(tt)

    if canon:
        cu = canon.strip().upper()
        if cu and cu not in job_tokens:
            job_tokens.insert(0, cu)

    seen = set()
    job_tokens = [x for x in job_tokens if not (x in seen or seen.add(x))]

    def _token_match(col, tok: str):
        tok = (tok or "").strip().upper()
        if not tok:
            return False
        c = func.upper(func.coalesce(func.trim(col), ""))
        return or_(
            c == tok,
            c.op("GLOB")(f"{tok}[^0-9A-Za-z]*"),
            c.op("GLOB")(f"*[^0-9A-Za-z]{tok}[^0-9A-Za-z]*"),
            c.op("GLOB")(f"*[^0-9A-Za-z]{tok}"),
        )

    base_q = (
        db.session.query(IssuedPartRecord)
        .options(joinedload(IssuedPartRecord.part), joinedload(IssuedPartRecord.batch))
        .outerjoin(IssuedBatch, IssuedBatch.id == IssuedPartRecord.batch_id)
    )

    token_ors = []
    for tok in job_tokens:
        token_ors.append(_token_match(IssuedPartRecord.reference_job, tok))
        token_ors.append(_token_match(IssuedBatch.reference_job, tok))

    if "work_order_id" in IssuedBatch.__table__.c:
        if token_ors:
            base_q = base_q.filter(
                or_(
                    IssuedBatch.work_order_id == wo.id,
                    and_(IssuedBatch.work_order_id.is_(None), or_(*token_ors)),
                )
            )
        else:
            base_q = base_q.filter(IssuedBatch.work_order_id == wo.id)
    else:
        if token_ors:
            base_q = base_q.filter(or_(*token_ors))

    # ------------------------------------------------------------
    # 4b) technician safety (оставлено как у тебя)
    # ------------------------------------------------------------
    tech_aliases = set()
    if wo.technician_name:
        tech_aliases.add(wo.technician_name.strip().lower())
    if wo.technician_username:
        tech_aliases.add(wo.technician_username.strip().lower())
    if wo.technician and wo.technician.username:
        tech_aliases.add(wo.technician.username.strip().lower())
    tech_aliases = {x for x in tech_aliases if x}

    if tech_aliases:
        issued_to_norm = func.lower(func.trim(func.coalesce(IssuedPartRecord.issued_to, "")))
        ref_norm = func.upper(func.trim(func.coalesce(IssuedPartRecord.reference_job, "")))

        base_q = base_q.filter(
            or_(
                issued_to_norm.in_(list(tech_aliases)),
                issued_to_norm == "",
                ref_norm.op("GLOB")("STOCK*"),
                ref_norm.op("GLOB")("RETURN*"),
            )
        )

    issued_items = base_q.order_by(
        IssuedPartRecord.issue_date.asc(),
        IssuedPartRecord.id.asc()
    ).all()


    money_issued = func.coalesce(
        func.sum(
            case(
                (
                    IssuedPartRecord.quantity > 0,
                    IssuedPartRecord.quantity * IssuedPartRecord.unit_cost_at_issue,
                ),
                else_=0,
            )
        ),
        0.0,
    )
    money_returned = func.coalesce(
        func.sum(
            case(
                (
                    IssuedPartRecord.quantity < 0,
                    -IssuedPartRecord.quantity * IssuedPartRecord.unit_cost_at_issue,
                ),
                else_=0,
            )
        ),
        0.0,
    )
    qty_issued = func.coalesce(
        func.sum(
            case(
                (IssuedPartRecord.quantity > 0, IssuedPartRecord.quantity),
                else_=0,
            )
        ),
        0,
    )
    qty_returned = func.coalesce(
        func.sum(
            case(
                (IssuedPartRecord.quantity < 0, -IssuedPartRecord.quantity),
                else_=0,
            )
        ),
        0,
    )

    agg = (
        db.session.query(money_issued, money_returned, qty_issued, qty_returned)
        .select_from(IssuedPartRecord)
        .outerjoin(IssuedBatch, IssuedBatch.id == IssuedPartRecord.batch_id)
        .filter(base_q.whereclause if base_q.whereclause is not None else True)
        .one()
    )

    issued_total = float(agg[0] or 0.0)
    returned_total = float(agg[1] or 0.0)
    net_total = issued_total - returned_total
    issued_qty = int(agg[2] or 0)
    returned_qty = int(agg[3] or 0)
    net_qty = issued_qty - returned_qty

    def _fmt(dt):
        return dt.strftime("%Y-%m-%d %H:%M") if dt else "—"

    def _extract(rec):
        return (
            rec.id,
            (getattr(rec.part, "part_number", None) or "").strip().upper(),
            getattr(rec.part, "name", None) or "",
            int(rec.quantity or 0),
            float(rec.unit_cost_at_issue or 0.0),
            bool(getattr(rec, "confirmed_by_tech", False)),
            getattr(rec, "confirmed_by", None) or None,
            _fmt(getattr(rec, "confirmed_at", None)),
        )

    grouped = defaultdict(list)
    for r in issued_items:
        if getattr(r, "batch_id", None):
            key = ("batch", r.batch_id)
        elif getattr(r, "invoice_number", None):
            key = ("inv", r.invoice_number)
        else:
            key = (
                "ungrouped",
                f"{r.issue_date:%Y%m%d%H%M%S}-{r.issued_to}-{r.reference_job or ''}",
            )
        grouped[key].append(r)

    net_by_pn = _dd(int)
    batches = []
    for _, recs in grouped.items():
        b = recs[0].batch if recs and getattr(recs[0], "batch", None) else None
        issued_at = (b.issue_date if b else recs[0].issue_date)
        issued_at_str = _fmt(issued_at)
        tech = recs[0].issued_to
        ref = (b.reference_job if b else (recs[0].reference_job or "")) or ""
        report_id = (b.invoice_number if b else recs[0].invoice_number)
        location = getattr(b, "location", None)

        ref_is_return = "RETURN" in ref.upper()

        items = []
        total_value_raw = 0.0
        unconfirmed_count = 0

        for rec in recs:
            rid, pn, name, qty, price, confirmed, who, when_s = _extract(rec)
            if not pn or qty == 0:
                continue
            line_total = qty * price
            total_value_raw += line_total

            is_item_return = ref_is_return or (qty < 0)
            eff_sign = -1 if is_item_return else 1
            net_by_pn[pn] += eff_sign * abs(qty)

            if (not confirmed) and qty > 0:
                unconfirmed_count += 1

            items.append(
                {
                    "id": rid,
                    "pn": pn,
                    "name": name or "—",
                    "qty": abs(qty),
                    "unit_price": price,
                    "negative": (line_total < 0) or is_item_return,
                    "confirmed": confirmed,
                    "confirmed_by": who,
                    "confirmed_at": when_s,
                }
            )

        is_group_return = (total_value_raw < 0) or ref_is_return
        all_conf = (unconfirmed_count == 0) and bool(items)
        any_conf = any(it.get("confirmed") for it in items) if items else False

        batches.append(
            {
                "issued_at_dt": issued_at,
                "issued_at": issued_at_str,
                "technician": tech,
                "canonical_ref": canon,
                "reference_job": ref,
                "location": location,
                "report_id": report_id,
                "invoice_number": report_id,
                "is_return": is_group_return,
                "total_value": total_value_raw,
                "items": items,
                "all_confirmed": all_conf,
                "any_confirmed": any_conf,
                "unconfirmed_count": unconfirmed_count,
            }
        )

    invoiced_pns = sorted([pn for pn, net in net_by_pn.items() if net > 0])

    # 5) BLENDED PRICING + INS-строки
    all_rows = []
    if wo.units:
        for u in wo.units:
            if getattr(u, "parts", None):
                all_rows.extend(u.parts)
    if (not all_rows) and getattr(wo, "parts", None):
        all_rows.extend(wo.parts)

    issued_info_by_pn = {}
    for rec in issued_items:
        pn = (getattr(rec.part, "part_number", "") or "").strip().upper()
        if not pn:
            continue
        q = int(rec.quantity or 0)
        cost_at_issue = float(rec.unit_cost_at_issue or 0.0)
        val = q * cost_at_issue

        info = issued_info_by_pn.get(pn)
        if not info:
            info = {"qty": 0, "value": 0.0}
            issued_info_by_pn[pn] = info
        info["qty"] += q
        info["value"] += val

    pn_list = []
    for r in all_rows:
        pn = (getattr(r, "part_number", "") or "").strip().upper()
        if pn and pn not in pn_list:
            pn_list.append(pn)

    inv_cost_map = {}
    if pn_list:
        part_rows = (
            db.session.query(Part)
            .filter(func.upper(Part.part_number).in_(pn_list))
            .all()
        )
        for pr in part_rows:
            key = (pr.part_number or "").strip().upper()
            inv_cost_map[key] = float(pr.unit_cost or 0.0)

    display_parts = []
    grand_total_display = 0.0

    for r in all_rows:
        pn = (getattr(r, "part_number", "") or "").strip().upper()
        qty_planned = int(getattr(r, "quantity", 0) or 0)

        is_ins = bool(getattr(r, "is_insurance_supplied", False))

        ii = issued_info_by_pn.get(pn)
        issued_qty_raw = int(ii["qty"] or 0) if ii else 0
        issued_val_raw = float(ii["value"] or 0.0) if ii else 0.0

        issued_qty_eff = issued_qty_raw if issued_qty_raw > 0 else 0
        issued_val_eff = issued_val_raw if issued_qty_raw > 0 else 0.0

        inv_cost = inv_cost_map.get(pn)
        if inv_cost is None:
            inv_cost = float(getattr(r, "unit_cost", 0.0) or 0.0)

        if qty_planned > 0:
            not_issued_qty = max(0, qty_planned - issued_qty_eff)
            not_issued_val = not_issued_qty * inv_cost
            blended_total_val = issued_val_eff + not_issued_val
            blended_unit_price = blended_total_val / qty_planned
        else:
            blended_total_val = 0.0
            blended_unit_price = 0.0

        # страховые строки в страховой работе — клиент = 0$
        if wo.job_type == "INSURANCE" and is_ins:
            blended_total_val = 0.0
            blended_unit_price = 0.0

        grand_total_display += blended_total_val

        raw_oflag = getattr(r, "ordered_flag", None)
        ordered_flag_truthy = raw_oflag not in (None, False, 0, "0", "", "false", "False")
        is_ordered = (
            ordered_flag_truthy
            or (str(getattr(r, "status", "")).strip().lower() == "ordered")
            or (str(getattr(r, "line_status", "")).strip().lower() == "ordered")
        )
        is_invoiced = pn in invoiced_pns

        display_parts.append(
            {
                "part_number": getattr(r, "part_number", "") or "",
                "alt_part_numbers": (
                    getattr(r, "alt_part_numbers", "") or
                    getattr(r, "alt_numbers", "") or
                    getattr(r, "alt_pn", "") or
                    ""
                ),
                "part_name": getattr(r, "part_name", "") or getattr(r, "name", "") or "—",
                "qty": qty_planned,
                "unit_price_display": blended_unit_price,
                "total_display": blended_total_val,
                "supplier": getattr(r, "supplier", "") or "",
                "is_ordered": is_ordered,
                "ordered_date": getattr(r, "ordered_date", None) or getattr(r, "ordered_on", None),
                "backorder_flag": bool(getattr(r, "backorder_flag", False)),
                "is_invoiced": is_invoiced,
                "is_insurance_supplied": is_ins,

                # ✅ INV# from WorkOrderPart.invoice_number
                "invoice_number": (getattr(r, "invoice_number", "") or "").strip(),
                "inv_ref": (getattr(r, "invoice_number", "") or "").strip(),
            }
        )

    # 6) отдаём в шаблон
    return render_template(
        "wo_detail.html",
        wo=wo,
        display_parts=display_parts,
        grand_total_display=grand_total_display,
        issued_items=issued_items,
        issued_total=issued_total,
        returned_total=returned_total,
        net_total=net_total,
        issued_qty=issued_qty,
        returned_qty=returned_qty,
        net_qty=net_qty,
        batches=batches,
        suppliers=suppliers,
        invoiced_pns=invoiced_pns,
        is_my_wo=is_my_wo,
        can_confirm=can_confirm_any,
        can_confirm_any=can_confirm_any,
        can_view_docs=can_view_docs,
        current_user=current_user,
    )
@inventory_bp.get("/reports_grouped/xlsx", endpoint="download_report_xlsx")
@login_required
def download_report_xlsx():
    from flask import request, send_file
    from datetime import datetime
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    from extensions import db
    from models import IssuedPartRecord, Part

    # --------- фильтры ---------
    recipient = (request.args.get("recipient") or "").strip() or None
    reference_job = (request.args.get("reference_job") or "").strip() or None

    invoice_number_s = (request.args.get("invoice_number") or "").strip()
    try:
        invoice_number = int(invoice_number_s) if invoice_number_s else None
    except ValueError:
        invoice_number = None

    inv_ref = (request.args.get("inv_ref") or "").strip() or None
    start_date_str = request.args.get("start_date") or None
    end_date_str = request.args.get("end_date") or None

    start_dt = None
    end_dt = None
    if start_date_str:
        try:
            start_dt = datetime.strptime(start_date_str, "%Y-%m-%d")
        except ValueError:
            start_dt = None
    if end_date_str:
        try:
            end_dt = datetime.strptime(end_date_str, "%Y-%m-%d")
        except ValueError:
            end_dt = None

    # --------- запрос ---------
    q = (
        db.session.query(IssuedPartRecord)
        .join(Part, IssuedPartRecord.part_id == Part.id)
    )

    if start_dt:
        q = q.filter(IssuedPartRecord.issue_date >= start_dt)
    if end_dt:
        q = q.filter(IssuedPartRecord.issue_date < end_dt.replace(hour=23, minute=59, second=59))

    if recipient:
        q = q.filter(IssuedPartRecord.issued_to.ilike(f"%{recipient}%"))
    if reference_job:
        q = q.filter(IssuedPartRecord.reference_job.ilike(f"%{reference_job}%"))
    if invoice_number is not None:
        q = q.filter(IssuedPartRecord.invoice_number == invoice_number)
    if inv_ref:
        q = q.filter(IssuedPartRecord.inv_ref.ilike(f"%{inv_ref}%"))

    q = q.order_by(
        IssuedPartRecord.invoice_number.asc().nullsfirst(),
        IssuedPartRecord.issue_date.asc(),
        IssuedPartRecord.id.asc(),
    )

    rows = q.all()

    # --------- Excel ---------
    wb = Workbook()
    ws_grouped = wb.active
    ws_grouped.title = "Grouped by Invoice"
    ws_flat = wb.create_sheet("Flat Data")

    header_fill = PatternFill("solid", fgColor="FFD9D9D9")
    invoice_header_fill = PatternFill("solid", fgColor="FFEBF1DE")
    subtotal_fill = PatternFill("solid", fgColor="FFFCE4D6")
    grand_total_fill = PatternFill("solid", fgColor="FFE2EFDA")

    return_fill = PatternFill("solid", fgColor="FFF8CBAD")
    stock_fill = PatternFill("solid", fgColor="FFDEEBF7")

    bold_font = Font(bold=True)
    title_font = Font(bold=True, size=14)
    filters_font = Font(italic=True, size=11)

    right_align = Alignment(horizontal="right")
    center_align = Alignment(horizontal="center")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # >>> итоговый набор колонок (13) <<<
    header = [
        "Date",             # A
        "Invoice #",        # B
        "Part #",           # C
        "Name",             # D
        "Qty",              # E
        "Unit Cost",        # F
        "Total",            # G
        "Issued To",        # H
        "Job Ref.",         # I
        "Location",         # J
        "INV# (vendor)",    # K
        "Return To",        # L
        "Return Company",   # M
    ]

    # строка фильтров
    filters_parts = []
    if start_date_str or end_date_str:
        filters_parts.append(f"from {start_date_str or '…'} to {end_date_str or '…'}")
    if recipient:
        filters_parts.append(f"Issued To contains '{recipient}'")
    if reference_job:
        filters_parts.append(f"Job Ref contains '{reference_job}'")
    if invoice_number is not None:
        filters_parts.append(f"Invoice # = {invoice_number}")
    if inv_ref:
        filters_parts.append(f"INV# contains '{inv_ref}'")

    filters_line = "Filters: " + ("; ".join(filters_parts) if filters_parts else "(none)")

    def fmt_invoice_num(inv):
        if inv is None:
            return ""
        try:
            return f"{int(inv):06d}"
        except (TypeError, ValueError):
            return str(inv)

    # ===== 1) GROUPED =====
    ws_grouped["A1"] = "Issued Parts Report (Grouped by Invoice)"
    ws_grouped.merge_cells("A1:M1")
    ws_grouped["A1"].font = title_font
    ws_grouped["A1"].alignment = center_align

    ws_grouped["A2"] = filters_line
    ws_grouped.merge_cells("A2:M2")
    ws_grouped["A2"].font = filters_font

    ws_grouped.append([])
    ws_grouped.append(header)
    header_row_grouped = ws_grouped.max_row

    for cell in ws_grouped[header_row_grouped]:
        cell.fill = header_fill
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border

    ws_grouped.freeze_panes = f"A{header_row_grouped + 1}"

    # ===== 2) FLAT =====
    ws_flat["A1"] = "Issued Parts Report (Flat Data)"
    ws_flat.merge_cells("A1:M1")
    ws_flat["A1"].font = title_font
    ws_flat["A1"].alignment = center_align

    ws_flat["A2"] = filters_line
    ws_flat.merge_cells("A2:M2")
    ws_flat["A2"].font = filters_font

    ws_flat.append([])
    ws_flat.append(header)
    header_row_flat = ws_flat.max_row

    for cell in ws_flat[header_row_flat]:
        cell.fill = header_fill
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border

    ws_flat.freeze_panes = f"A{header_row_flat + 1}"

    # ===== 3) DATA =====
    total_sum = 0.0
    current_invoice = None
    current_invoice_total = 0.0

    for rec in rows:
        part = rec.part
        qty = rec.quantity or 0
        unit_cost = rec.unit_cost_at_issue or 0
        line_total = qty * unit_cost
        total_sum += float(line_total)

        inv_val = rec.invoice_number
        inv_str = fmt_invoice_num(inv_val)

        # GROUPED: новая шапка инвойса
        if inv_val != current_invoice:
            # subtotal предыдущего инвойса
            if current_invoice is not None:
                ws_grouped.append([
                    "", "", "", "Invoice subtotal:",
                    "", "", round(current_invoice_total, 2),
                    "", "", "", "", "", ""
                ])
                for cell in ws_grouped[ws_grouped.max_row]:
                    cell.fill = subtotal_fill
                    cell.font = bold_font
                    cell.border = thin_border
                ws_grouped[ws_grouped.max_row][6].alignment = right_align  # G (Total)

                ws_grouped.append([])

            header_text = f"Invoice {inv_str or '(no number)'}"
            ws_grouped.append([header_text] + [""] * 12)  # 13 колонок
            for cell in ws_grouped[ws_grouped.max_row]:
                cell.fill = invoice_header_fill
                cell.font = bold_font
                cell.border = thin_border

            current_invoice = inv_val
            current_invoice_total = 0.0

        current_invoice_total += float(line_total)

        row_common = [
            rec.issue_date.date().isoformat() if rec.issue_date else "",
            inv_str,
            part.part_number if part else "",
            part.name if part else "",
            qty,
            float(f"{unit_cost:.2f}"),
            float(f"{line_total:.2f}"),
            rec.issued_to or "",
            rec.reference_job or "",
            rec.location or (part.location if part else ""),
            rec.inv_ref or "",
            (getattr(rec, "return_to", None) or ""),
            (rec.return_destination.name if getattr(rec, "return_destination", None) else ""),
        ]

        # GROUPED append
        ws_grouped.append(row_common)
        row_idx_grouped = ws_grouped.max_row

        # FLAT append
        ws_flat.append(list(row_common))
        row_idx_flat = ws_flat.max_row

        # return/stock flags
        is_return = (qty or 0) < 0 or ((rec.reference_job or "").upper().startswith("RETURN"))
        loc_upper = (rec.location or "").upper()
        ref_upper = (rec.reference_job or "").upper()
        is_stock = loc_upper.startswith("STOCK") or ref_upper.startswith("STOCK")

        # styles grouped
        for cell in ws_grouped[row_idx_grouped]:
            cell.border = thin_border
            if is_return:
                cell.fill = return_fill
            elif is_stock:
                cell.fill = stock_fill

        # align numeric (Qty E, UnitCost F, Total G)
        ws_grouped[row_idx_grouped][4].alignment = right_align  # E Qty
        ws_grouped[row_idx_grouped][5].alignment = right_align  # F Unit Cost
        ws_grouped[row_idx_grouped][6].alignment = right_align  # G Total

        # styles flat
        for cell in ws_flat[row_idx_flat]:
            cell.border = thin_border
            if is_return:
                cell.fill = return_fill
            elif is_stock:
                cell.fill = stock_fill

        ws_flat[row_idx_flat][4].alignment = right_align
        ws_flat[row_idx_flat][5].alignment = right_align
        ws_flat[row_idx_flat][6].alignment = right_align

    # subtotal последнего инвойса
    if current_invoice is not None:
        ws_grouped.append([
            "", "", "", "Invoice subtotal:",
            "", "", round(current_invoice_total, 2),
            "", "", "", "", "", ""
        ])
        for cell in ws_grouped[ws_grouped.max_row]:
            cell.fill = subtotal_fill
            cell.font = bold_font
            cell.border = thin_border
        ws_grouped[ws_grouped.max_row][6].alignment = right_align

    # GRAND TOTAL
    if rows:
        ws_grouped.append([])
        ws_grouped.append([
            "", "", "", "GRAND TOTAL:",
            "", "", round(total_sum, 2),
            "", "", "", "", "", ""
        ])
        for cell in ws_grouped[ws_grouped.max_row]:
            cell.fill = grand_total_fill
            cell.font = bold_font
            cell.border = thin_border
        ws_grouped[ws_grouped.max_row][6].alignment = right_align

    # filters
    ws_grouped.auto_filter.ref = f"A{header_row_grouped}:M{ws_grouped.max_row}"
    ws_flat.auto_filter.ref = f"A{header_row_flat}:M{ws_flat.max_row}"

    # widths A..M
    for sheet in (ws_grouped, ws_flat):
        sheet.column_dimensions["A"].width = 12  # Date
        sheet.column_dimensions["B"].width = 10  # Invoice #
        sheet.column_dimensions["C"].width = 14  # Part #
        sheet.column_dimensions["D"].width = 32  # Name
        sheet.column_dimensions["E"].width = 8   # Qty
        sheet.column_dimensions["F"].width = 10  # Unit Cost
        sheet.column_dimensions["G"].width = 12  # Total
        sheet.column_dimensions["H"].width = 18  # Issued To
        sheet.column_dimensions["I"].width = 16  # Job Ref.
        sheet.column_dimensions["J"].width = 14  # Location
        sheet.column_dimensions["K"].width = 16  # INV# (vendor)
        sheet.column_dimensions["L"].width = 12  # Return To
        sheet.column_dimensions["M"].width = 22  # Return Company

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    if start_date_str or end_date_str:
        fname = f"issued_parts_{start_date_str or ''}_{end_date_str or ''}.xlsx"
    else:
        fname = "issued_parts_report.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@inventory_bp.get("/reports_grouped/returns_xlsx", endpoint="download_returns_xlsx")
@login_required
def download_returns_xlsx():
    from flask import request, send_file
    from datetime import datetime
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    from extensions import db
    from models import IssuedPartRecord, Part

    recipient = (request.args.get("recipient") or "").strip() or None
    reference_job = (request.args.get("reference_job") or "").strip() or None

    invoice_number_s = (request.args.get("invoice_number") or "").strip()
    try:
        invoice_number = int(invoice_number_s) if invoice_number_s else None
    except ValueError:
        invoice_number = None

    inv_ref = (request.args.get("inv_ref") or "").strip() or None
    start_date_str = request.args.get("start_date") or None
    end_date_str = request.args.get("end_date") or None

    start_dt = None
    end_dt = None
    if start_date_str:
        try:
            start_dt = datetime.strptime(start_date_str, "%Y-%m-%d")
        except ValueError:
            start_dt = None
    if end_date_str:
        try:
            end_dt = datetime.strptime(end_date_str, "%Y-%m-%d")
        except ValueError:
            end_dt = None

    q = (
        db.session.query(IssuedPartRecord)
        .join(Part, IssuedPartRecord.part_id == Part.id)
    )

    if start_dt:
        q = q.filter(IssuedPartRecord.issue_date >= start_dt)
    if end_dt:
        q = q.filter(IssuedPartRecord.issue_date < end_dt.replace(hour=23, minute=59, second=59))

    if recipient:
        q = q.filter(IssuedPartRecord.issued_to.ilike(f"%{recipient}%"))
    if reference_job:
        q = q.filter(IssuedPartRecord.reference_job.ilike(f"%{reference_job}%"))
    if invoice_number is not None:
        q = q.filter(IssuedPartRecord.invoice_number == invoice_number)
    if inv_ref:
        q = q.filter(IssuedPartRecord.inv_ref.ilike(f"%{inv_ref}%"))

    # только возвраты
    q = q.filter(
        (IssuedPartRecord.quantity < 0) |
        (IssuedPartRecord.reference_job.ilike("RETURN%"))
    )

    q = q.order_by(
        IssuedPartRecord.issue_date.asc(),
        IssuedPartRecord.invoice_number.asc().nullsfirst(),
        IssuedPartRecord.id.asc(),
    )

    rows = q.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Returns"

    header_fill = PatternFill("solid", fgColor="FFD9D9D9")
    return_fill = PatternFill("solid", fgColor="FFF8CBAD")

    bold_font = Font(bold=True)
    title_font = Font(bold=True, size=14)
    filters_font = Font(italic=True, size=11)

    right_align = Alignment(horizontal="right")
    center_align = Alignment(horizontal="center")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # >>> 13 колонок (как и в Issued) <<<
    header = [
        "Date",             # A
        "Invoice #",        # B
        "Part #",           # C
        "Name",             # D
        "Qty",              # E
        "Unit Cost",        # F
        "Total",            # G
        "Issued To",        # H
        "Job Ref.",         # I
        "Location",         # J
        "INV# (vendor)",    # K
        "Return To",        # L
        "Return Company",   # M
    ]

    # фильтры
    filters_parts = []
    if start_date_str or end_date_str:
        filters_parts.append(f"from {start_date_str or '…'} to {end_date_str or '…'}")
    if recipient:
        filters_parts.append(f"Issued To contains '{recipient}'")
    if reference_job:
        filters_parts.append(f"Job Ref contains '{reference_job}'")
    if invoice_number is not None:
        filters_parts.append(f"Invoice # = {invoice_number}")
    if inv_ref:
        filters_parts.append(f"INV# contains '{inv_ref}'")

    filters_line = "Filters: " + ("; ".join(filters_parts) if filters_parts else "(none)")

    ws["A1"] = "Returns Report"
    ws.merge_cells("A1:M1")
    ws["A1"].font = title_font
    ws["A1"].alignment = center_align

    ws["A2"] = filters_line
    ws.merge_cells("A2:M2")
    ws["A2"].font = filters_font

    ws.append([])
    ws.append(header)
    header_row = ws.max_row

    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border

    ws.freeze_panes = f"A{header_row + 1}"

    total_sum = 0.0

    def fmt_invoice_num(inv):
        if inv is None:
            return ""
        try:
            return f"{int(inv):06d}"
        except (TypeError, ValueError):
            return str(inv)

    for rec in rows:
        part = rec.part
        qty = rec.quantity or 0
        unit_cost = rec.unit_cost_at_issue or 0
        line_total = qty * unit_cost
        total_sum += float(line_total)

        inv_str = fmt_invoice_num(rec.invoice_number)

        ws.append([
            rec.issue_date.date().isoformat() if rec.issue_date else "",
            inv_str,
            part.part_number if part else "",
            part.name if part else "",
            qty,
            float(f"{unit_cost:.2f}"),
            float(f"{line_total:.2f}"),
            rec.issued_to or "",
            rec.reference_job or "",
            rec.location or (part.location if part else ""),
            rec.inv_ref or "",
            (getattr(rec, "return_to", None) or ""),
            (rec.return_destination.name if getattr(rec, "return_destination", None) else ""),
        ])

        row_idx = ws.max_row
        for cell in ws[row_idx]:
            cell.border = thin_border
            cell.fill = return_fill

        # align numeric (Qty E, UnitCost F, Total G)
        ws[row_idx][4].alignment = right_align
        ws[row_idx][5].alignment = right_align
        ws[row_idx][6].alignment = right_align

    if rows:
        ws.append([])
        ws.append([
            "", "", "", "TOTAL RETURNS:",
            "", "", round(total_sum, 2),
            "", "", "", "", "", ""
        ])

        for cell in ws[ws.max_row]:
            cell.font = bold_font
            cell.border = thin_border
        ws[ws.max_row][6].alignment = right_align  # G total

    ws.auto_filter.ref = f"A{header_row}:M{ws.max_row}"

    # widths A..M
    widths = {
        "A": 12,  # Date
        "B": 10,  # Invoice #
        "C": 14,  # Part #
        "D": 32,  # Name
        "E": 8,   # Qty
        "F": 10,  # Unit Cost
        "G": 12,  # Total
        "H": 18,  # Issued To
        "I": 16,  # Job Ref.
        "J": 14,  # Location
        "K": 16,  # INV# (vendor)
        "L": 12,  # Return To
        "M": 22,  # Return Company
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="returns_report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
@inventory_bp.get("/reports_grouped/stock_xlsx", endpoint="download_stock_xlsx")
@login_required
def download_stock_xlsx():
    from flask import request, send_file
    from datetime import datetime
    from io import BytesIO

    from sqlalchemy import or_
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    from extensions import db
    from models import IssuedPartRecord, Part

    recipient = (request.args.get("recipient") or "").strip() or None
    reference_job = (request.args.get("reference_job") or "").strip() or None

    invoice_number_s = (request.args.get("invoice_number") or "").strip()
    try:
        invoice_number = int(invoice_number_s) if invoice_number_s else None
    except ValueError:
        invoice_number = None

    # (опционально, чтобы было одинаково с Issued/Returns)
    inv_ref = (request.args.get("inv_ref") or "").strip() or None

    start_date_str = request.args.get("start_date") or None
    end_date_str = request.args.get("end_date") or None

    start_dt = None
    end_dt = None
    if start_date_str:
        try:
            start_dt = datetime.strptime(start_date_str, "%Y-%m-%d")
        except ValueError:
            start_dt = None
    if end_date_str:
        try:
            end_dt = datetime.strptime(end_date_str, "%Y-%m-%d")
        except ValueError:
            end_dt = None

    q = (
        db.session.query(IssuedPartRecord)
        .join(Part, IssuedPartRecord.part_id == Part.id)
    )

    if start_dt:
        q = q.filter(IssuedPartRecord.issue_date >= start_dt)
    if end_dt:
        q = q.filter(IssuedPartRecord.issue_date < end_dt.replace(hour=23, minute=59, second=59))

    if recipient:
        q = q.filter(IssuedPartRecord.issued_to.ilike(f"%{recipient}%"))
    if reference_job:
        q = q.filter(IssuedPartRecord.reference_job.ilike(f"%{reference_job}%"))
    if invoice_number is not None:
        q = q.filter(IssuedPartRecord.invoice_number == invoice_number)
    if inv_ref:
        q = q.filter(IssuedPartRecord.inv_ref.ilike(f"%{inv_ref}%"))

    # STOCK: is_stock=True или STOCK* в reference_job/location
    is_stock_col = getattr(IssuedPartRecord, "is_stock", None)
    if is_stock_col is not None:
        q = q.filter(
            or_(
                is_stock_col.is_(True),
                IssuedPartRecord.reference_job.ilike("STOCK%"),
                IssuedPartRecord.location.ilike("STOCK%"),
            )
        )
    else:
        q = q.filter(
            or_(
                IssuedPartRecord.reference_job.ilike("STOCK%"),
                IssuedPartRecord.location.ilike("STOCK%"),
            )
        )

    q = q.order_by(
        IssuedPartRecord.location.asc().nullsfirst(),
        IssuedPartRecord.issue_date.asc(),
        IssuedPartRecord.id.asc(),
    )

    rows = q.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Movements"

    header_fill = PatternFill("solid", fgColor="FFD9D9D9")
    stock_fill = PatternFill("solid", fgColor="FFDEEBF7")

    bold_font = Font(bold=True)
    title_font = Font(bold=True, size=14)
    filters_font = Font(italic=True, size=11)

    right_align = Alignment(horizontal="right")
    center_align = Alignment(horizontal="center")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    header = [
        "Date",         # A
        "Location",     # B
        "Invoice #",    # C
        "Part #",       # D
        "Name",         # E
        "Qty",          # F
        "Unit Cost",    # G
        "Total",        # H
        "Issued To",    # I
        "Job Ref.",     # J
    ]

    filters_parts = []
    if start_date_str or end_date_str:
        filters_parts.append(f"from {start_date_str or '…'} to {end_date_str or '…'}")
    if recipient:
        filters_parts.append(f"Issued To contains '{recipient}'")
    if reference_job:
        filters_parts.append(f"Job Ref contains '{reference_job}'")
    if invoice_number is not None:
        filters_parts.append(f"Invoice # = {invoice_number}")
    if inv_ref:
        filters_parts.append(f"INV# contains '{inv_ref}'")

    filters_line = "Filters: " + ("; ".join(filters_parts) if filters_parts else "(none)")

    ws["A1"] = "Stock Movements Report"
    ws.merge_cells("A1:J1")
    ws["A1"].font = title_font
    ws["A1"].alignment = center_align

    ws["A2"] = filters_line
    ws.merge_cells("A2:J2")
    ws["A2"].font = filters_font

    ws.append([])
    ws.append(header)

    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border

    ws.freeze_panes = f"A{header_row + 1}"

    total_sum = 0.0

    def fmt_invoice_num(inv):
        if inv is None:
            return ""
        try:
            return f"{int(inv):06d}"
        except (TypeError, ValueError):
            return str(inv)

    for rec in rows:
        part = rec.part
        qty = rec.quantity or 0
        unit_cost = rec.unit_cost_at_issue or 0
        line_total = qty * unit_cost
        total_sum += float(line_total)

        inv_str = fmt_invoice_num(rec.invoice_number)

        ws.append([
            rec.issue_date.date().isoformat() if rec.issue_date else "",
            rec.location or (part.location if part else ""),
            inv_str,
            part.part_number if part else "",
            part.name if part else "",
            qty,
            float(f"{unit_cost:.2f}"),
            float(f"{line_total:.2f}"),
            rec.issued_to or "",
            rec.reference_job or "",
        ])

        row_idx = ws.max_row

        for cell in ws[row_idx]:
            cell.border = thin_border
            cell.fill = stock_fill

        # align numeric columns: Qty(F), Unit Cost(G), Total(H)
        ws[row_idx][5].alignment = right_align
        ws[row_idx][6].alignment = right_align
        ws[row_idx][7].alignment = right_align

    if rows:
        ws.append([])
        ws.append([
            "", "", "", "", "TOTAL STOCK MOVEMENTS:",
            "", "", round(total_sum, 2),
            "", ""
        ])
        for cell in ws[ws.max_row]:
            cell.font = bold_font
            cell.border = thin_border
        ws[ws.max_row][7].alignment = right_align  # H

    ws.auto_filter.ref = f"A{header_row}:J{ws.max_row}"

    widths = {
        "A": 12,  # Date
        "B": 14,  # Location
        "C": 10,  # Invoice #
        "D": 14,  # Part #
        "E": 32,  # Name
        "F": 8,   # Qty
        "G": 10,  # Unit Cost
        "H": 12,  # Total
        "I": 18,  # Issued To
        "J": 16,  # Job Ref
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="stock_movements_report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@inventory_bp.post("/work_orders/new", endpoint="wo_create")
@login_required
def wo_create():
    role = (getattr(current_user, "role", "") or "").strip().lower()
    if role not in ("admin", "superadmin"):
        flash("Access denied", "danger")
        return redirect(url_for("inventory.wo_list"))

    f = request.form

    # NEW: prefer technician_id; fallback to technician_name
    tech_id_raw = (f.get("technician_id") or "").strip()
    technician_id = int(tech_id_raw) if tech_id_raw.isdigit() else None
    technician_name = (f.get("technician_name") or "").strip()

    if technician_id:
        tech_obj = db.session.get(User, technician_id)
        if not tech_obj or (tech_obj.role or "").lower() != "technician":
            flash("Selected technician is invalid.", "danger")
            return redirect(url_for("inventory.wo_new"))
        technician_name = tech_obj.username  # normalize to username
    else:
        # require at least a name when no id provided
        if not technician_name:
            flash("Technician is required.", "danger")
            return redirect(url_for("inventory.wo_new"))

    job_numbers   = (f.get("job_numbers")   or "").strip()
    brand         = (f.get("brand")         or "").strip()
    model         = (f.get("model")         or "").strip()
    serial        = (f.get("serial")        or "").strip()
    job_type      = (f.get("job_type")      or "BASE").strip().upper()

    def _f(x, default=0.0):
        try: return float(x)
        except Exception: return float(default)

    delivery_fee   = _f(f.get("delivery_fee"), 0)
    markup_percent = _f(f.get("markup_percent"), 0)

    if not job_numbers:
        flash("Job(s) are required.", "danger")
        return redirect(url_for("inventory.wo_new"))

    # normalize job list
    jobs = [j.strip() for j in job_numbers.replace(",", " ").split() if j.strip()]

    # create WO
    wo = WorkOrder(
        technician_id=technician_id,       # NEW
        technician_name=technician_name,   # keep legacy compatibility
        job_numbers=", ".join(jobs),
        brand=brand, model=model, serial=serial,
        job_type=job_type if job_type in ("BASE", "INSURANCE") else "BASE",
        delivery_fee=delivery_fee,
        markup_percent=markup_percent,
        status="search_ordered",
    )
    db.session.add(wo)
    db.session.flush()

    # parse part rows
    import re
    pat = re.compile(r"^rows\[(\d+)\]\[(\w+)\]$")
    tmp = {}
    for k, v in f.items():
        m = pat.match(k)
        if not m: continue
        i = int(m.group(1)); field = m.group(2)
        tmp.setdefault(i, {})[field] = (v or "").strip()

    def _clip20(s: str) -> str: return (s or "").strip()[:20]
    def _clip6(s: str)  -> str: return (s or "").strip()[:6]

    for i in sorted(tmp.keys()):
        row = tmp[i]
        pn = _clip20((row.get("part_number") or "").upper())
        if not pn:
            continue

        alt_raw = (row.get("alt_numbers") or row.get("alt_part_numbers") or "")
        alt_tokens = [_clip20(t) for t in alt_raw.split(",") if t is not None]
        alt_joined = ",".join(alt_tokens)

        try:
            qty = int(row.get("quantity") or 0)
        except Exception:
            qty = 0

        part = WorkOrderPart(
            work_order_id=wo.id,
            part_number=pn,
            alt_numbers=alt_joined,
            part_name=(row.get("part_name") or "").strip(),
            quantity=qty,
            supplier=_clip6(row.get("supplier") or ""),
            backorder_flag=bool(row.get("backorder_flag")),
            status="search_ordered",
        )
        db.session.add(part)

    db.session.commit()
    flash("Work order created.", "success")
    return redirect(url_for("inventory.wo_detail", wo_id=wo.id))

def _render_new_wo_form(prefill=None, units_prefill=None, flash_msg=None):
    """
    Рендер той самой multi-appliance формы /work_orders/new
    без редиректа, с уже введёнными данными.
    """
    from flask import session, render_template, flash
    from flask_login import current_user
    from models import User

    # recent suppliers (чтоб datalist и т.п. остались как раньше)
    recent_suppliers = session.get("recent_suppliers", [])

    if flash_msg:
        flash(flash_msg, "warning")

    # ---------- technician dropdown ----------
    # у тебя в шаблоне "This list includes only users with the technician role."
    tech_users = User.query.filter(
        (User.role == "technician") | (User.role == "TECHNICIAN")
    ).order_by(User.username.asc()).all()

    # ---------- шапка ордера ----------
    # prefill может быть WorkOrder (несохранённый) или dict
    wo_hdr = {
        "technician_id":    "",
        "technician_name":  "",
        "job_numbers":      "",
        "job_type":         "BASE",
        "delivery_fee":     0.0,
        "markup_percent":   0.0,
        "status":           "search_ordered",
    }
    if prefill:
        # аккуратно копируем поля, если есть
        wo_hdr["technician_id"]   = getattr(prefill, "technician_id",   "") or ""
        wo_hdr["technician_name"] = getattr(prefill, "technician_name", "") or ""
        wo_hdr["job_numbers"]     = getattr(prefill, "job_numbers",     "") or ""
        wo_hdr["job_type"]        = getattr(prefill, "job_type",        "BASE") or "BASE"
        wo_hdr["delivery_fee"]    = getattr(prefill, "delivery_fee",    0.0) or 0.0
        wo_hdr["markup_percent"]  = getattr(prefill, "markup_percent",  0.0) or 0.0
        wo_hdr["status"]          = getattr(prefill, "status",          "search_ordered") or "search_ordered"

    # ---------- блоки appliances / units / rows ----------
    # Это то, что ты собираешь внутри wo_save в units_payload.
    # Если у нас уже есть units_prefill (спарсили из формы перед валидацией),
    # то отдаём его; иначе отдаём дефолт с одним appliance и одной пустой строкой.
    if units_prefill:
        units_for_template = units_prefill
    else:
        units_for_template = [{
            "brand": "",
            "model": "",
            "serial": "",
            "rows": [{
                "part_number": "",
                "part_name": "",
                "quantity": 1,
                "alt_numbers": "",
                "warehouse": "",
                "supplier": "",
                "backorder_flag": False,
                "ordered_flag": False,
                "unit_cost": "",
                "line_status": "search_ordered",
                "stock_hint": "",
            }],
        }]

    # ВАЖНО: этот шаблон должен быть ТОЧНО той страницей, которая у тебя на /work_orders/new сейчас.
    # Если он у тебя называется иначе – подставь правильный путь.
    return render_template(
        "work_orders/new.html",
        wo=wo_hdr,
        units=units_for_template,
        tech_users=tech_users,
        recent_suppliers=recent_suppliers,
        readonly=False,
    )

@inventory_bp.get("/work_orders/newx", endpoint="wo_newx")
@login_required
def wo_newx(prefill_wo=None, prefill_units=None):
    from flask import session, render_template
    recent_suppliers = session.get("recent_suppliers", [])

    role_low = (getattr(current_user, "role", "") or "").lower()
    if role_low not in ("admin", "superadmin"):
        flash("Access denied", "danger")
        return redirect(url_for("inventory.wo_list"))

    if prefill_wo is not None:
        wo = prefill_wo
    else:
        wo = None

    if prefill_units is not None:
        units = prefill_units
    else:
        units = [{
            "brand": "",
            "model": "",
            "serial": "",
            "rows": [{
                "part_number": "",
                "part_name": "",
                "quantity": 1,
                "alt_numbers": "",
                "supplier": "",
                "backorder_flag": False,
                "line_status": "search_ordered",
                "warehouse": "",
                "unit_cost": "",
            }],
        }]

    return render_template(
        "wo_form.html",
        wo=wo,
        units=units,
        readonly=False,
        recent_suppliers=recent_suppliers,
    )

@inventory_bp.post("/work_orders/save", endpoint="wo_save")
@login_required
def wo_save():
    from flask import request, session, redirect, url_for, flash, current_app, render_template
    from models import WorkOrder, WorkUnit, WorkOrderPart, User, JobReservation
    from extensions import db
    from datetime import datetime, timedelta
    from flask_login import current_user
    import re
    from datetime import date
    from sqlalchemy import or_
    from flask import request

    # ---------- helpers ----------
    def _clip(s, n):
        return (s or "").strip()[:n]

    def _i(x, default=0):
        try:
            return int(x)
        except Exception:
            return int(default)

    def _f(x, default=None):
        if x is None or x == "":
            return default
        try:
            return float(x)
        except Exception:
            return default

    def _b(x):
        return str(x).strip().lower() in ("1", "true", "on", "yes", "y")

    def _safe_detail_redirect(work_order):
        wo_id_val = getattr(work_order, "id", None)
        if wo_id_val:
            return redirect(url_for("inventory.wo_detail", wo_id=wo_id_val))
        return redirect(url_for("inventory.wo_list"))

    def _parse_jobs(raw: str) -> list[str]:
        s = (raw or "").upper()
        parts = re.split(r"[,\s;]+", s)
        cleaned, seen = [], set()
        for p in parts:
            p = p.strip()
            if not p:
                continue
            if p not in seen:
                seen.add(p)
                cleaned.append(p)
        return cleaned

    def _rerender_same_screen(msg_text: str, errors=None):
        db.session.rollback()

        # ✅ restore audit after rollback (so UI shows correct actor)
        from datetime import datetime as _dt
        actor_id = getattr(current_user, "id", None)
        now = _dt.utcnow()

        actor_username = (getattr(current_user, "username", None) or "").strip()
        actor_role = (getattr(current_user, "role", None) or "").strip()

        with open(dbg_path, "a", encoding="utf-8") as fp:
            fp.write(f"ACTOR: id={actor_id!r} user={actor_username!r} role={actor_role!r}\n")
            fp.write(
                "WO_BEFORE: "
                f"wo_id={getattr(wo, 'id', None)!r} "
                f"tech_id={getattr(wo, 'technician_id', None)!r} "
                f"created_by_id={getattr(wo, 'created_by_id', None)!r} "
                f"updated_by_id={getattr(wo, 'updated_by_id', None)!r}\n"
            )

        try:
            if is_new and not getattr(wo, "created_by_id", None):
                wo.created_by_id = actor_id
                if not getattr(wo, "created_at", None):
                    wo.created_at = now
            wo.updated_by_id = actor_id
            wo.updated_at = now

            with open(dbg_path, "a", encoding="utf-8") as fp:
                fp.write(
                    "WO_AFTER_SET: "
                    f"created_by_id={getattr(wo, 'created_by_id', None)!r} "
                    f"updated_by_id={getattr(wo, 'updated_by_id', None)!r}\n"
                )

        except Exception:
            pass

        flash(msg_text, "warning")

        technicians = _query_technicians()
        recent_suppliers = session.get("recent_suppliers", []) or []

        sel_tid = wo.technician_id
        sel_tname = wo.technician_name or None

        safe_units = units_payload if units_payload else [{
            "brand": wo.brand or "",
            "model": wo.model or "",
            "serial": wo.serial or "",
            "rows": [{
                "id": None,
                "part_number": "",
                "part_name": "",
                "quantity": 1,
                "alt_numbers": "",
                "warehouse": "",
                "supplier": "",
                "backorder_flag": False,
                "line_status": "search_ordered",
                "unit_cost": 0.0,
                "ordered_flag": False,
                "is_insurance_supplied": False,
                "invoice_number": "",
            }],
        }]

        return render_template(
            "wo_form_units.html",
            wo=wo,
            units=safe_units,
            recent_suppliers=recent_suppliers,
            readonly=False,
            technicians=technicians,
            selected_tech_id=sel_tid,
            selected_tech_username=sel_tname,
            errors=(errors or {}),
        ), 400


    def _jobs_set_from_wo(wo_obj) -> set[str]:
        return set(_parse_jobs(getattr(wo_obj, "job_numbers", "") or ""))

    # ---------- access control ----------
    role_low = (getattr(current_user, "role", "") or "").strip().lower()
    if role_low not in ("admin", "superadmin"):
        flash("Access denied", "danger")
        return redirect(url_for("inventory.wo_list"))

    f = request.form

    # --- DEBUG FILE (form keys with invoice_number) ---
    import os
    from datetime import datetime
    dbg_path = os.path.join(current_app.instance_path, "wo_save_debug.txt")
    inv_keys = [k for k in f.keys() if "invoice_number" in k]
    with open(dbg_path, "a", encoding="utf-8") as fp:
        fp.write("\n" + "=" * 60 + "\n")
        fp.write("TS: " + datetime.now().isoformat() + "\n")
        fp.write("PATH: /work_orders/save hit\n")
        fp.write("INV_KEYS_COUNT: " + str(len(inv_keys)) + "\n")
        fp.write("INV_KEYS_SAMPLE:\n")
        for k in inv_keys[:20]:
            fp.write(f"  {k} = {f.get(k)!r}\n")

        pn_keys = [k for k in f.keys() if k.endswith("[part_number]")]
        fp.write("PN_KEYS_COUNT: " + str(len(pn_keys)) + "\n")
        fp.write("PN+INV SAMPLE:\n")
        for k in pn_keys[:20]:
            inv_k = k.replace("[part_number]", "[invoice_number]")
            fp.write(f"  {k} = {f.get(k)!r} | {inv_k} = {f.get(inv_k)!r}\n")

    log_keys = sorted(f.keys())
    current_app.logger.debug("WO_SAVE keys (%s): %s", len(log_keys), log_keys[:200])

    # new vs edit
    wo_id  = (f.get("wo_id") or "").strip()
    is_new = not wo_id

    # взять WorkOrder
    if is_new:
        wo = WorkOrder()
    else:
        wo = WorkOrder.query.get_or_404(int(wo_id))

    from datetime import datetime as _dt

    # ---------- заголовок / шапка ----------
    tech_id_raw   = (f.get("technician_id")   or f.get("technician") or "").strip()
    tech_name_raw = (f.get("technician_name") or f.get("technician") or "").strip()

    tech_id_val   = None
    tech_name_val = None

    if tech_id_raw.isdigit():
        try:
            tid = int(tech_id_raw)
            u = User.query.get(tid)
            if u:
                tech_id_val   = tid
                tech_name_val = (u.username or "").strip().upper()
        except Exception:
            pass

    if not tech_name_val:
        if tech_name_raw:
            tech_name_val = tech_name_raw.strip().upper()
        else:
            tech_name_val = None

    wo.technician_id   = tech_id_val
    wo.technician_name = (tech_name_val or "").strip().upper() if tech_name_val else ""

    # job_numbers всегда в UPPER
    wo.job_numbers     = (f.get("job_numbers") or "").upper().strip()
    wo.job_type        = (f.get("job_type") or "BASE").strip().upper()
    wo.delivery_fee    = _f(f.get("delivery_fee"), 0) or 0.0
    wo.markup_percent  = _f(f.get("markup_percent"), 0) or 0.0

    st_field = (f.get("status") or "search_ordered").strip()
    wo.status = st_field if st_field in ("search_ordered", "ordered", "done") else "search_ordered"

    # Customer PO (только для INSURANCE, иначе чистим)
    po = (f.get("customer_po") or "").strip().upper()
    if (wo.job_type or "").upper() == "INSURANCE":
        wo.customer_po = po or None
    else:
        wo.customer_po = None

    # Эти поля сейчас мало используются в multi-appliance, но приводим к UPPER на всякий случай
    brand_hdr_raw  = (f.get("brand")  or "").strip().upper()
    model_hdr_raw  = _clip(f.get("model"), 25).upper() if f.get("model") else ""
    serial_hdr_raw = _clip(f.get("serial"), 25).upper() if f.get("serial") else ""
    if brand_hdr_raw:
        wo.brand = brand_hdr_raw
    if model_hdr_raw:
        wo.model = model_hdr_raw
    if serial_hdr_raw:
        wo.serial = serial_hdr_raw

    # ---------- ВАЛИДАЦИЯ заголовка ----------
    if not tech_name_val:
        if is_new:
            pass
        else:
            db.session.rollback()
            flash("Technician is required before saving Work Order.", "warning")
            return redirect(url_for("inventory.wo_detail", wo_id=wo.id))

    if not wo.job_numbers:
        if is_new:
            pass
        else:
            db.session.rollback()
            flash("Job number is required.", "warning")
            return redirect(url_for("inventory.wo_detail", wo_id=wo.id))

    # =========================
    # HARD LOCK CHECK (atomic)
    # =========================
    if is_new:
        now = datetime.utcnow()
        uid = getattr(current_user, "id", None)
        uname = (getattr(current_user, "username", None) or "").strip()

        tokens = _job_tokens(wo.job_numbers)

        if tokens:
            JobReservation.query.filter(JobReservation.expires_at < now).delete(synchronize_session=False)

            locked = (
                JobReservation.query
                .filter(
                    JobReservation.job_token.in_(tokens),
                    JobReservation.expires_at >= now,
                    JobReservation.holder_user_id != uid
                )
                .first()
            )
            if locked:
                return _rerender_same_screen(
                    f"Job {locked.job_token} is locked by {locked.holder_username or 'another user'}. Try again later.",
                    errors={"job_numbers": "This job is currently locked by another user."}
                )

            exp = now + timedelta(minutes=15)
            for t in tokens:
                row = JobReservation.query.filter_by(job_token=t).first()
                if not row:
                    row = JobReservation(job_token=t)
                row.holder_user_id = uid
                row.holder_username = uname or None
                row.expires_at = exp
                db.session.add(row)

            db.session.flush()

    # =========================
    # /HARD LOCK CHECK
    # =========================

    # === DUP GUARD ===
    input_jobs = _parse_jobs(wo.job_numbers)

    # ✅ ВАЖНО: если job_numbers пустой — НЕ делаем dup-guard сейчас.
    # Пусть дальше соберётся units_payload и сработает _rerender_same_screen(),
    # чтобы НЕ потерять введённые данные.
    if input_jobs:
        input_set = set(input_jobs)
        like_filters = [WorkOrder.job_numbers.ilike(f"%{j}%") for j in input_jobs]
        q = WorkOrder.query.filter(or_(*like_filters))
        if not is_new:
            q = q.filter(WorkOrder.id != wo.id)

        possibles = q.order_by(WorkOrder.created_at.desc()).limit(50).all()
        for existing in possibles:
            exist_set = _jobs_set_from_wo(existing)
            if exist_set & input_set:
                flash(
                    f"Work Order for job(s) {', '.join(sorted(exist_set & input_set))} "
                    f"already exists (#{existing.id}).",
                    "warning",
                )
                return _rerender_same_screen(
                    f"Duplicate: Work Order #{existing.id} already exists for job(s) ...",
                    errors={"job_numbers": "Duplicate job number(s)."}
                )

    # ---------- собрать units[...] и их rows[...] ----------
    re_unit = re.compile(r"^units\[(\d+)\]\[(brand|model|serial)\]$")
    re_row = re.compile(
        r"^units\[(\d+)\]\[rows\]\[(\d+)\]\[(part_number|part_name|quantity|"
        r"alt_numbers|alt_part_numbers|warehouse|unit_label|supplier|supplier_name|"
        r"invoice_number|"
        r"backorder_flag|status|unit_cost|ordered_flag|is_insurance_supplied)\]$"
    )

    units_map = {}
    for key in f.keys():
        m = re_unit.match(key)
        if m:
            ui, name = int(m.group(1)), m.group(2)
            units_map.setdefault(ui, {"rows": {}})
            units_map[ui][name] = f.get(key)
            continue
        m = re_row.match(key)
        if m:
            ui, ri, name = int(m.group(1)), int(m.group(2)), m.group(3)
            units_map.setdefault(ui, {"rows": {}})
            units_map[ui]["rows"].setdefault(ri, {})
            if name in ("ordered_flag", "backorder_flag", "is_insurance_supplied"):
                vals = request.form.getlist(key)
                val  = vals[-1] if vals else f.get(key)
            else:
                val  = f.get(key)
            units_map[ui]["rows"][ri][name] = val

    # превратить units_map -> units_payload
    units_payload  = []
    new_rows_count = 0

    for ui in sorted(units_map.keys()):
        u_blk = units_map[ui]
        rows_payload = []

        for ri in sorted(u_blk.get("rows", {}).keys()):
            r = u_blk["rows"][ri]

            pn = (r.get("part_number") or "").strip().upper()
            qty = _i(r.get("quantity") or 0, 0)

            alt_raw = (r.get("alt_numbers") or r.get("alt_part_numbers") or "").strip()
            wh_raw  = (r.get("warehouse")   or r.get("unit_label")        or "").strip()
            sup_raw = (r.get("supplier")    or r.get("supplier_name")     or "").strip()

            # ВСЕ PN-подобные поля в UPPER
            alt_raw = alt_raw.upper()
            wh_raw  = wh_raw.upper()
            sup_raw = sup_raw.upper()

            inv_raw = (r.get("invoice_number") or "").strip().upper()
            inv_raw = inv_raw[:32] if inv_raw else ""

            ucost   = _f(r.get("unit_cost"), None)
            bo_flag = _b(r.get("backorder_flag"))
            lstatus = (r.get("status") or "search_ordered").strip()
            ord_flag = _b(r.get("ordered_flag"))
            ins_flag = _b(r.get("is_insurance_supplied"))

            # PART NAME всегда в верхнем регистре
            part_name_raw  = (r.get("part_name") or "").strip().upper()
            part_name_clip = _clip(part_name_raw, 120)

            qty_eff = qty if qty else 1

            row_dict = {
                "id": None,
                "part_number": _clip(pn, 80),
                "part_name": part_name_clip,
                "quantity": qty_eff,
                "alt_numbers": _clip(alt_raw, 200),
                "warehouse": _clip(wh_raw, 120),
                "supplier": _clip(sup_raw, 80),
                "invoice_number": inv_raw,
                "backorder_flag": bo_flag,
                "line_status": lstatus if lstatus in ("search_ordered", "ordered", "done") else "search_ordered",
                "unit_cost": (ucost if (ucost is not None) else 0.0),
                "ordered_flag": ord_flag,
                "is_insurance_supplied": ins_flag,
            }

            if (pn or part_name_clip) and qty_eff > 0:
                new_rows_count += 1
                rows_payload.append(row_dict)
                with open(dbg_path, "a", encoding="utf-8") as fp:
                    fp.write(f"ROWS_PARSED: ui={ui} ri={ri} pn={pn!r} inv={inv_raw!r}\n")

        units_payload.append({
            "brand":  (u_blk.get("brand")  or "").strip().upper(),
            "model":  _clip(u_blk.get("model"), 25).upper(),
            "serial": _clip(u_blk.get("serial"), 25).upper(),
            "rows":   rows_payload,
        })

    current_app.logger.debug(
        "WO_SAVE parsed units=%s rows_total=%s",
        len(units_payload),
        sum(len(u.get('rows') or []) for u in units_payload)
    )

    # ---------- ВАЛИДАЦИЯ (продолжение) ----------
    if not tech_name_val:
        if is_new:
            return _rerender_same_screen("Technician is required before saving Work Order.")
        else:
            db.session.rollback()
            flash("Technician is required before saving Work Order.", "warning")
            return redirect(url_for("inventory.wo_detail", wo_id=wo.id))

    if not wo.job_numbers:
        if is_new:
            return _rerender_same_screen(
                "Job number is required.",
                errors={"job_numbers": "Job number is required."}
            )
        else:
            db.session.rollback()
            flash("Job number is required.", "warning")
            return redirect(url_for("inventory.wo_detail", wo_id=wo.id))


    if is_new:
        db.session.add(wo)

    # --- allow header-only save ---
    has_any_rows = any(u.get("rows") for u in units_payload)
    has_any_unit_header = any(
        (u.get("brand") or u.get("model") or u.get("serial"))
        for u in units_payload
    )
    if not has_any_rows and not has_any_unit_header:
        current_app.logger.debug(
            "WO_SAVE: no unit rows parsed; saving header only for WO #%s",
            getattr(wo, "id", None) or "(new)"
        )

    # ---------- сохранить прежние ordered_flag / ordered_date ----------
    def _norm_supplier(s):
        s = (s or "").strip()
        return " ".join(s.split()).lower()

    def _unit_key(b, m, s):
        return (
            _clip(b, 80).lower(),
            _clip(m, 25).lower(),
            _clip(s, 25).lower()
        )

    old_index = {}
    if not is_new:
        for old_u in (wo.units or []):
            uk = _unit_key(old_u.brand or "", old_u.model or "", old_u.serial or "")
            for old_p in (old_u.parts or []):
                pn   = ((old_p.part_number or "").strip().upper())
                supn = _norm_supplier(old_p.supplier)
                was_ordered = (
                    bool(getattr(old_p, "ordered_flag", False)) or
                    ((getattr(old_p, "status", "") or "").lower() == "ordered") or
                    ((getattr(old_p, "line_status", "") or "").lower() == "ordered")
                )
                old_index[(uk, pn, supn)] = (
                    was_ordered,
                    getattr(old_p, "ordered_date", None)
                )

    # ---------- удалить старые юниты/parts и пересоздать ----------
    for u in list(getattr(wo, "units", []) or []):
        for p in list(getattr(u, "parts", []) or []):
            db.session.delete(p)
        db.session.delete(u)

    try:
        db.session.flush()
    except Exception as e:
        db.session.rollback()
        flash(f"Failed to refresh units on Work Order: {e}", "danger")
        return _safe_detail_redirect(wo)

    if units_payload:
        first_unit = units_payload[0]
        wo.brand  = (first_unit.get("brand")  or "").strip() or None
        wo.model  = _clip(first_unit.get("model"), 25) or None
        wo.serial = _clip(first_unit.get("serial"), 25) or None

    suppliers_seen = []

    # ---------- создаём заново WorkUnit и WorkOrderPart ----------
    for up in units_payload:
        if not (up.get("brand") or up.get("model") or up.get("serial")) and not up.get("rows"):
            continue

        unit = WorkUnit(
            work_order=wo,
            brand=(up.get("brand") or "").strip().upper(),
            model=_clip(up.get("model"), 25).upper(),
            serial=_clip(up.get("serial"), 25).upper(),
        )
        db.session.add(unit)

        try:
            db.session.flush()
        except Exception as e:
            db.session.rollback()
            flash(f"Failed to add unit: {e}", "danger")
            return _safe_detail_redirect(wo)

        uk = _unit_key(unit.brand or "", unit.model or "", unit.serial or "")

        for r in (up.get("rows") or []):
            sup = r.get("supplier") or ""
            if sup:
                norm = " ".join(sup.split())
                if norm and norm.lower() not in [x.lower() for x in suppliers_seen]:
                    suppliers_seen.append(norm)

            pn_upper = (r.get("part_number") or "").strip().upper()
            sup_norm = _norm_supplier(sup)

            ord_in   = bool(r.get("ordered_flag"))
            ins_flag = bool(r.get("is_insurance_supplied"))

            prev_state = old_index.get((uk, pn_upper, sup_norm))
            prev_was_ordered, prev_date = (prev_state if prev_state else (False, None))

            if ord_in:
                if prev_was_ordered and prev_date:
                    ord_date = prev_date
                else:
                    ord_date = date.today()
            else:
                ord_date = None
                ord_in   = False

            # ✅ invoice_number: нормализуем ОДИН РАЗ
            inv = (r.get("invoice_number") or "").strip().upper()
            inv = inv[:32] if inv else None

            wop = WorkOrderPart(
                work_order=wo,
                unit=unit,
                part_number=pn_upper,
                part_name=(r.get("part_name") or None),
                quantity=int(r.get("quantity") or 0),
                alt_part_numbers=(r.get("alt_numbers") or None),
                supplier=(sup or None),
                backorder_flag=bool(r.get("backorder_flag")),
                status=("ordered" if ord_in else "search_ordered"),
                is_insurance_supplied=ins_flag,
                invoice_number=inv,
            )

            if hasattr(wop, "warehouse"):
                wop.warehouse = (r.get("warehouse") or "")[:120]
            wop.unit_label = (r.get("warehouse") or "")[:120] or None

            if hasattr(wop, "unit_cost"):
                uc = r.get("unit_cost")
                if uc is not None and uc != "":
                    try:
                        wop.unit_cost = float(uc)
                    except Exception:
                        wop.unit_cost = None

            if hasattr(wop, "ordered_flag"):
                wop.ordered_flag = ord_in
            if hasattr(wop, "ordered_date"):
                wop.ordered_date = ord_date
            if hasattr(wop, "line_status"):
                wop.line_status = "ordered" if ord_in else "search_ordered"

            db.session.add(wop)

            # --- DEBUG write (no flush inside loop) ---
            with open(dbg_path, "a", encoding="utf-8") as fp:
                fp.write(
                    f"BEFORE COMMIT: wo={getattr(wo,'id',None)} pn={pn_upper} inv={inv!r} obj_inv={getattr(wop,'invoice_number',None)!r}\n"
                )

    # ---------- commit ----------
    # ---------- FINAL AUDIT (MUST BE RIGHT BEFORE COMMIT) ----------
    from datetime import datetime as _dt
    actor_id = getattr(current_user, "id", None)
    now = _dt.utcnow()

    if is_new and not getattr(wo, "created_by_id", None):
        wo.created_by_id = actor_id
        if not getattr(wo, "created_at", None):
            wo.created_at = now

    wo.updated_by_id = actor_id
    wo.updated_at = now
    # ---------- /FINAL AUDIT ----------

    try:
        db.session.commit()
        fresh = WorkOrder.query.get(wo.id)
        with open(dbg_path, "a", encoding="utf-8") as fp:
            fp.write(
                "WO_AFTER_COMMIT: "
                f"wo_id={wo.id} "
                f"tech_id={getattr(fresh, 'technician_id', None)!r} "
                f"created_by_id={getattr(fresh, 'created_by_id', None)!r} "
                f"updated_by_id={getattr(fresh, 'updated_by_id', None)!r}\n"
            )

        flash("Work Order saved.", "success")

        # --- DEBUG: verify WO audit in DB right after commit ---
        fresh = WorkOrder.query.get(wo.id)
        with open(dbg_path, "a", encoding="utf-8") as fp:
            fp.write("AFTER COMMIT WO AUDIT:\n")
            fp.write(
                f"  current_user.id={getattr(current_user, 'id', None)} username={getattr(current_user, 'username', None)!r}\n")
            fp.write(f"  wo.id={wo.id}\n")
            fp.write(
                f"  db.created_by_id={getattr(fresh, 'created_by_id', None)} created_by_user={(fresh.created_by_user.username if fresh.created_by_user else None)!r}\n")
            fp.write(
                f"  db.updated_by_id={getattr(fresh, 'updated_by_id', None)} updated_by_user={(fresh.updated_by_user.username if fresh.updated_by_user else None)!r}\n")

        rows = (WorkOrderPart.query
                .filter_by(work_order_id=wo.id)
                .with_entities(WorkOrderPart.part_number, WorkOrderPart.invoice_number)
                .all())

        with open(dbg_path, "a", encoding="utf-8") as fp:
            fp.write("AFTER COMMIT DB:\n")
            for pn, inv in rows:
                fp.write(f"  pn={pn} inv={inv!r}\n")

    except Exception as e:
        db.session.rollback()
        try:
            db.session.commit()

            # ✅ RELEASE reservation right after successful NEW WO save
            if is_new:
                try:
                    now2 = datetime.utcnow()
                    tokens2 = _job_tokens(wo.job_numbers)
                    if tokens2:
                        JobReservation.query.filter(
                            JobReservation.job_token.in_(tokens2),
                            JobReservation.holder_user_id == actor_id
                        ).delete(synchronize_session=False)
                        db.session.commit()  # отдельный commit на удаление резерва — ок
                except Exception:
                    db.session.rollback()

            flash("Work Order saved.", "success")

        except Exception as e:
            db.session.rollback()
            flash(f"Failed to save Work Order: {e}", "danger")
            return redirect(url_for("inventory.wo_list"))

        return redirect(url_for("inventory.wo_list"))

    # ---------- remember suppliers ----------
    if suppliers_seen:
        cur = session.get("recent_suppliers", [])
        merged, seen = [], set()
        for x in suppliers_seen + list(cur):
            xl = x.lower()
            if xl in seen:
                continue
            seen.add(xl)
            merged.append(x)
        session["recent_suppliers"] = merged[:20]
        session.modified = True

    return redirect(url_for("inventory.wo_detail", wo_id=wo.id))

@inventory_bp.get("/work_orders")
@login_required
def wo_list():
    """
    Unified search for Work Orders.

    GET params:
      q       - free text (tech name, job#, brand, model, part#, alt part#, part name)
      from    - date (YYYY-MM-DD)
      to      - date (YYYY-MM-DD)
      tech_id - technician id (admin/superadmin only; technicians ignored)
      type    - BASE | INSURANCE
      status  - search_ordered | ordered | done
    """
    from datetime import datetime, timedelta
    from sqlalchemy import and_, or_, func, String
    from models import IssuedBatch, IssuedPartRecord
    import re

    # -----------------------------
    # helper: token-safe job match
    # -----------------------------
    def _job_token_match(col, tok: str):
        """
        SQLite-safe token match for WorkOrder.job_numbers.

        - digits: matches as standalone token, not as substring inside longer digits
          handles: '991472', '991472 991929', 'RETURN 991472 991929'
        - alnum: space/edge boundaries
        """
        tok = (tok or "").strip()
        if not tok:
            return False

        c = func.coalesce(func.trim(col), "")

        if tok.isdigit():
            return or_(
                c == tok,
                c.op("GLOB")(f"{tok}[^0-9]*"),            # starts with tok then non-digit boundary
                c.op("GLOB")(f"*[^0-9]{tok}[^0-9]*"),     # middle token
                c.op("GLOB")(f"*[^0-9]{tok}"),            # ends with tok
            )

        return or_(
            c == tok,
            c.like(f"{tok} %"),
            c.like(f"% {tok} %"),
            c.like(f"% {tok}"),
            c.like(f"%{tok},%"),
        )

    # ---- incoming params ----
    qtext = (request.args.get("q") or "").strip()
    dfrom = (request.args.get("from") or "").strip()
    dto   = (request.args.get("to")   or "").strip()

    tech_id_raw = (request.args.get("tech_id") or "").strip()
    type_raw    = (request.args.get("type")    or "").strip().upper()
    status_raw  = (request.args.get("status")  or "").strip()

    tech_id = int(tech_id_raw) if tech_id_raw.isdigit() else None
    job_type = type_raw if type_raw in ("BASE", "INSURANCE") else ""
    status   = status_raw if status_raw in ("search_ordered", "ordered", "done") else ""

    # --- UI markers (so you can show: "Matched by Invoice #1005") ---
    matched_by_invoice = False
    matched_invoice_number = ""
    invoice_matched_wo_ids = set()

    # ---- base query ----
    q = db.session.query(WorkOrder)
    filters = []
    joined_parts = False

    # ---- technician visibility restriction ----
    if is_technician():
        me_id   = getattr(current_user, "id", None)
        me_name = (getattr(current_user, "username", "") or "").strip()
        filters.append(or_(
            WorkOrder.technician_id == me_id,
            func.trim(WorkOrder.technician_name) == me_name,
        ))
        tech_id = None  # технику нельзя выбирать чужого техника

    # ---- explicit filters (Apply) ----
    if tech_id:
        filters.append(WorkOrder.technician_id == tech_id)

    if job_type:
        filters.append(WorkOrder.job_type == job_type)

    if status:
        filters.append(WorkOrder.status == status)

    # ---- free-text search ----
    if qtext:
        like = f"%{qtext}%"

        # join parts
        q = q.outerjoin(WorkOrderPart, WorkOrderPart.work_order_id == WorkOrder.id)
        joined_parts = True

        conds = [
            WorkOrder.technician_name.ilike(like),
            WorkOrder.job_numbers.ilike(like),
            WorkOrder.brand.ilike(like),
            WorkOrder.model.ilike(like),

            WorkOrderPart.part_number.ilike(like),
            WorkOrderPart.alt_part_numbers.ilike(like),
            WorkOrderPart.part_name.ilike(like),

            # Customer PO
            func.coalesce(WorkOrder.customer_po, "").ilike(like),

            # Supplier/Receiving INV# shown in WO detail table (usually stored on WorkOrderPart)
            func.coalesce(WorkOrderPart.invoice_number, "").ilike(like),
        ]

        qnorm = qtext.strip()

        # ---- Issued INV# search (handles "000918" and "918") ----
        if qnorm.isdigit():
            matched_by_invoice = True
            matched_invoice_number = qnorm

            inv_raw = qnorm
            inv_trim = inv_raw.lstrip("0") or "0"
            inv_z6 = inv_trim.zfill(6)
            inv_variants = {inv_raw, inv_trim, inv_z6}
            inv_int = int(inv_trim) if inv_trim.isdigit() else None

            ref_jobs = set()

            # IssuedBatch.invoice_number can be int OR string -> compare as string + (optional) as int
            qb = db.session.query(IssuedBatch.reference_job).filter(
                or_(
                    func.trim(func.cast(IssuedBatch.invoice_number, String)).in_(list(inv_variants)),
                    (IssuedBatch.invoice_number == inv_int) if inv_int is not None else False,
                )
            )
            for (ref,) in qb.all():
                if ref:
                    ref_jobs.add(str(ref).strip())

            # fallback: IssuedPartRecord.invoice_number (old rows)
            qr = db.session.query(IssuedPartRecord.reference_job).filter(
                or_(
                    func.trim(func.cast(IssuedPartRecord.invoice_number, String)).in_(list(inv_variants)),
                    (IssuedPartRecord.invoice_number == inv_int) if inv_int is not None else False,
                )
            )
            for (ref,) in qr.all():
                if ref:
                    ref_jobs.add(str(ref).strip())

            # Turn ref_jobs into TOKENS, because ref might be "991472 991929"
            invoice_job_tokens = []
            for ref in ref_jobs:
                for t in re.findall(r"[A-Za-z0-9]+", ref or ""):
                    tt = (t or "").strip()
                    if tt:
                        invoice_job_tokens.append(tt)

            # de-dup keep order
            seen = set()
            invoice_job_tokens = [x for x in invoice_job_tokens if not (x in seen or seen.add(x))]

            # Add invoice-derived job tokens into WO search (NO canonical_job property!)
            for tok in invoice_job_tokens:
                conds.append(_job_token_match(WorkOrder.job_numbers, tok))

            # Also compute wo_ids to mark rows in UI as "matched by invoice"
            if invoice_job_tokens:
                id_q = db.session.query(WorkOrder.id).filter(
                    or_(*[_job_token_match(WorkOrder.job_numbers, tok) for tok in invoice_job_tokens])
                )
                # Respect same visibility/explicit filters (technician filter etc.)
                if filters:
                    id_q = id_q.filter(and_(*filters))
                invoice_matched_wo_ids = {int(x[0]) for x in id_q.all()}

        filters.append(or_(*conds))

    # ---- created_at date range (inclusive) ----
    if dfrom:
        try:
            start_dt = datetime.strptime(dfrom, "%Y-%m-%d")
            filters.append(WorkOrder.created_at >= start_dt)
        except ValueError:
            pass

    if dto:
        try:
            end_dt = datetime.strptime(dto, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
            filters.append(WorkOrder.created_at <= end_dt)
        except ValueError:
            pass

    # ---- apply filters ----
    if filters:
        q = q.filter(and_(*filters))

    # ---- avoid duplicate WO rows if multiple parts match
    if joined_parts:
        q = q.distinct(WorkOrder.id)

    # ---- fetch rows ----
    MAX_LIMIT = 200

    base_q = q.order_by(WorkOrder.created_at.desc())

    has_filters = any([
        bool(qtext),
        bool(dfrom),
        bool(dto),
        bool(tech_id),
        bool(job_type),
        bool(status),
    ])

    # IMPORTANT: лимит только когда нет фильтров
    if not has_filters:
        base_q = base_q.limit(MAX_LIMIT)

    items = base_q.all()
    count_items = len(items)

    # ---- datalist hints ----
    tech_suggestions = (
        db.session.query(WorkOrder.technician_name)
        .distinct().order_by(WorkOrder.technician_name.asc()).limit(50).all()
    )
    brand_model_suggestions = (
        db.session.query(WorkOrder.brand, WorkOrder.model)
        .distinct().limit(100).all()
    )
    recent_parts = (
        db.session.query(WorkOrderPart.part_number, WorkOrderPart.part_name)
        .distinct().limit(100).all()
    )
    hint_values_set = set()
    for (t_name,) in tech_suggestions:
        if t_name: hint_values_set.add(t_name)
    for (b, m) in brand_model_suggestions:
        combo = ((b or '') + ' ' + (m or '')).strip()
        if combo: hint_values_set.add(combo)
    for (pn, pname) in recent_parts:
        if pn: hint_values_set.add(pn)
        if pname: hint_values_set.add(pname)
    hint_values = sorted(hint_values_set)

    # ---- technicians for dropdown ----
    try:
        technicians = (
            db.session.query(User.id, User.username)
            .filter(func.lower(User.role) == "technician")
            .order_by(User.username.asc())
            .all()
        )
    except Exception:
        technicians = []

    # ---- filters context for template ----
    filters_ctx = {
        "q": qtext,
        "from": dfrom,
        "to": dto,
        "tech_id": tech_id,
        "type": job_type,
        "status": status,
    }

    return render_template(
        "wo_list.html",
        items=items,
        count_items=count_items,
        hint_values=hint_values,
        technicians=technicians,
        filters=filters_ctx,

        # ✅ NEW: UI markers
        matched_by_invoice=matched_by_invoice,
        matched_invoice_number=matched_invoice_number,
        invoice_matched_wo_ids=invoice_matched_wo_ids,
    )

@inventory_bp.post("/work_orders/<int:wo_id>/issue_instock", endpoint="wo_issue_instock")
@login_required
def wo_issue_instock(wo_id):
    # доступ
    if getattr(current_user, "role", "") not in ("admin", "superadmin"):
        flash("Access denied", "danger")
        return redirect(url_for("inventory.wo_detail", wo_id=wo_id))

    import re
    from urllib.parse import urlencode
    from markupsafe import Markup
    from sqlalchemy import func
    from datetime import datetime
    from flask import request, session

    from extensions import db
    from models import WorkOrder, WorkOrderPart, Part, IssuedPartRecord, IssuedBatch

    wo = WorkOrder.query.get_or_404(wo_id)
    is_ins_job = ((wo.job_type or "").upper() == "INSURANCE")

    # --- audit helper: who changed WO ---
    def _touch_wo():
        # safe even if columns differ
        try:
            if hasattr(wo, "updated_by_id"):
                wo.updated_by_id = getattr(current_user, "id", None)
            elif hasattr(wo, "updated_by"):
                wo.updated_by = getattr(current_user, "username", None)
        except Exception:
            pass
        try:
            wo.updated_at = datetime.utcnow()
        except Exception:
            pass
        db.session.add(wo)


    set_status = (request.form.get("set_status") or "").strip().lower()

    # === availability (для greedy) ===
    try:
        avail_rows = compute_availability(wo) or []
    except Exception:
        avail_rows = []

    stock_map: dict[str, int] = {}
    hint_map: dict[str, str] = {}
    for r in avail_rows:
        pn = (r.get("part_number") or "").strip().upper()
        if not pn:
            continue
        on_hand = int(r.get("on_hand") or 0)
        stock_map[pn] = stock_map.get(pn, 0) + on_hand
        hint_map[pn] = (
            r.get("status_hint")
            or r.get("hint")
            or ("STOCK" if on_hand > 0 else "WAIT")
        )

    def can_issue(pn: str, qty: int) -> bool:
        pn = (pn or "").strip().upper()
        if not pn or qty <= 0:
            return False
        left = int(stock_map.get(pn, 0))
        if left >= qty:
            stock_map[pn] = left - qty
            return True
        return False

    # --- хелпер: следующий номер инвойса ---
    def _reserve_invoice_number() -> int:
        mb = (
            db.session.query(func.coalesce(func.max(IssuedBatch.invoice_number), 0))
            .scalar()
            or 0
        )
        ml = (
            db.session.query(func.coalesce(func.max(IssuedPartRecord.invoice_number), 0))
            .scalar()
            or 0
        )
        return max(int(mb), int(ml)) + 1

    # ===== INV from UI: inv_map (preferred) + fallback inv_<id> =====
    inv_by_id: dict[int, str] = {}
    inv_map_raw = (request.form.get("inv_map") or "").strip()

    # format: "123=ABC;124=ZX9"
    if inv_map_raw:
        for pair in inv_map_raw.split(";"):
            pair = pair.strip()
            if not pair or "=" not in pair:
                continue
            k, v = pair.split("=", 1)
            k = (k or "").strip()
            v = (v or "").strip().upper()
            if k.isdigit() and v:
                inv_by_id[int(k)] = v[:32]

    def _read_inv_for_line(line_id: int) -> str:
        lid = int(line_id)

        # 1) inv_map has priority (works even if inputs are outside the submitted form)
        v = (inv_by_id.get(lid) or "").strip().upper()
        if v:
            return v[:32]

        # 2) fallback: classic input names inside the SAME form
        candidates = [
            f"inv_{lid}",
            f"invoice_{lid}",
            f"invoice_number_{lid}",
            f"invoice_number[{lid}]",
            f"wop[{lid}][invoice_number]",
            f"rows_by_id[{lid}][invoice_number]",
            f"parts[{lid}][invoice_number]",
        ]
        for name in candidates:
            vv = (request.form.get(name) or "").strip().upper()
            if vv:
                return vv[:32]

        patt = re.compile(rf".*\[{lid}\].*\[invoice_number\]$", re.IGNORECASE)
        for k in request.form.keys():
            if patt.match(k):
                vv = (request.form.get(k) or "").strip().upper()
                if vv:
                    return vv[:32]
        return ""

    # === 2) Selected mode ===
    raw_ids = (request.form.get("part_ids") or "").strip()

    items_to_issue = []
    issued_row_ids: list[int] = []
    skipped_rows = []
    new_records: list[IssuedPartRecord] = []
    issue_date = datetime.utcnow()

    raw_ins_ids = (request.form.get("ins_ids") or "").strip()
    ins_ids: set[int] = set()
    if raw_ins_ids:
        for tok in raw_ins_ids.split(","):
            tok = tok.strip()
            if tok.isdigit():
                ins_ids.add(int(tok))

    if raw_ids:
        ids = [int(tok) for tok in raw_ids.split(",") if tok.strip().isdigit()]
        if not ids:
            flash("Nothing selected to issue.", "warning")
            return redirect(url_for("inventory.wo_detail", wo_id=wo_id))

        wops: list[WorkOrderPart] = (
            WorkOrderPart.query
            .filter(
                WorkOrderPart.work_order_id == wo_id,
                WorkOrderPart.id.in_(ids)
            )
            .all()
        )
        if not wops:
            flash("Selected parts not found.", "warning")
            return redirect(url_for("inventory.wo_detail", wo_id=wo_id))

        part_has_location = hasattr(Part, "location")

        # precompute INV per selected line from UI snapshot
        inv_now_by_line_id: dict[int, str] = {}
        for line in wops:
            inv_now_by_line_id[int(line.id)] = _read_inv_for_line(int(line.id))

        for line in wops:
            pn = (line.part_number or "").strip().upper()
            qty_req = int(line.quantity or 0)
            if not pn or qty_req <= 0:
                continue

            # ✅ INV# snapshot from UI (even without Save)
            inv_now = (inv_now_by_line_id.get(int(line.id)) or "").strip().upper()[:32]

            # ✅ if user typed INV# but didn't Save — save it now
            # (safe: just one field, minimal risk)
            if inv_now and hasattr(line, "invoice_number"):
                try:
                    line.invoice_number = inv_now
                    db.session.add(line)
                except Exception:
                    pass

            is_ins_line = (
                (line.id in ins_ids)
                or bool(getattr(line, "is_insurance_supplied", False))
            )

            q_base = Part.query.filter(func.upper(Part.part_number) == pn)
            part = None
            if part_has_location and getattr(line, "warehouse", None):
                part = q_base.filter(
                    func.coalesce(Part.location, "") == (line.warehouse or "")
                ).first()
            if not part:
                part = q_base.first()

            hint_norm = (hint_map.get(pn) or "STOCK").upper()
            if hasattr(line, "stock_hint"):
                try:
                    line.stock_hint = hint_norm
                except Exception:
                    pass

            # ========== A) INS ==========
            if is_ins_job and is_ins_line:
                if not part:
                    part = Part(part_number=pn)
                    if hasattr(part, "name"):
                        part.name = getattr(line, "part_name", "") or pn
                    if hasattr(Part, "location"):
                        setattr(part, "location", "0.00INS")
                    if hasattr(Part, "quantity"):
                        setattr(part, "quantity", 0)
                    if hasattr(Part, "unit_cost"):
                        setattr(part, "unit_cost", 0.0)
                    db.session.add(part)
                    db.session.flush()

                rec = IssuedPartRecord(
                    part_id=part.id,
                    quantity=qty_req,
                    issued_to=(wo.technician_name or "").strip(),
                    reference_job=(wo.canonical_job or "").strip(),
                    issued_by=getattr(current_user, "username", "system"),
                    issue_date=issue_date,
                    unit_cost_at_issue=0.0,
                    is_insurance_supplied=True,
                    location=getattr(part, "location", "INS"),
                )
                # store INV# separately if field exists
                if hasattr(rec, "inv_ref"):
                    rec.inv_ref = inv_now or None

                if hasattr(rec, "part_number"):
                    rec.part_number = pn
                if hasattr(rec, "name_at_issue"):
                    rec.name_at_issue = getattr(line, "part_name", "") or "—"

                db.session.add(rec)
                new_records.append(rec)
                issued_row_ids.append(line.id)

                issued_so_far = int(getattr(line, "issued_qty", 0) or 0)
                line.issued_qty = issued_so_far + qty_req
                try:
                    line.status = "done"
                except Exception:
                    pass
                if hasattr(line, "last_issued_at"):
                    line.last_issued_at = issue_date
                db.session.add(line)
                continue

            # ========== B) STOCK ==========
            if not part:
                skipped_rows.append({
                    "id": line.id,
                    "pn": pn,
                    "name": getattr(line, "part_name", "") or "—",
                    "qty": qty_req,
                    "hint": hint_norm,
                })
                continue

            ok = can_issue(pn, qty_req)

            if not ok and part:
                try:
                    real_left = int(getattr(part, "quantity", 0) or 0)
                except Exception:
                    real_left = 0
                if real_left >= qty_req:
                    stock_map[pn] = int(stock_map.get(pn, 0))
                    stock_map[pn] = max(0, stock_map[pn] - qty_req)
                    ok = True

            hint_norm = (hint_map.get(pn) or ("STOCK" if ok else "WAIT"))
            if hasattr(line, "stock_hint"):
                try:
                    line.stock_hint = hint_norm
                except Exception:
                    pass

            if not ok:
                skipped_rows.append({
                    "id": line.id,
                    "pn": pn,
                    "name": getattr(line, "part_name", "") or "—",
                    "qty": qty_req,
                    "hint": hint_norm,
                })
                continue

            try:
                real_cost = float(part.unit_cost or 0.0)
            except Exception:
                real_cost = 0.0

            items_to_issue.append({
                "part_id": part.id,
                "qty": qty_req,
                "unit_price": real_cost,
                "inv_ref": inv_now,  # ✅ IMPORTANT: goes into IssuedPartRecord.inv_ref via helper
            })
            issued_row_ids.append(line.id)

        # --- issue stock items in bulk ---
        if items_to_issue:
            try:
                issue_date_stock, created_records = _issue_records_bulk(
                    issued_to=wo.technician_name,
                    reference_job=wo.canonical_job,
                    items=items_to_issue,
                )
                if issue_date_stock:
                    issue_date = issue_date_stock
                if created_records:
                    new_records.extend(created_records)
            except Exception as e:
                db.session.rollback()
                flash(f"Error issuing stock items: {e}", "danger")
                return redirect(url_for("inventory.wo_detail", wo_id=wo_id))

            # update issued_qty/status/last_issued_at for selected lines
            now = datetime.utcnow()
            for line in WorkOrderPart.query.filter(
                WorkOrderPart.id.in_(issued_row_ids)
            ).all():
                qty_needed = int(line.quantity or 0)
                issued_so_far = int(getattr(line, "issued_qty", 0) or 0)
                issue_now = max(qty_needed - issued_so_far, 0)
                if issue_now > 0:
                    line.issued_qty = issued_so_far + issue_now
                    if line.issued_qty >= qty_needed:
                        try:
                            line.status = "done"
                        except Exception:
                            pass
                    if hasattr(line, "last_issued_at"):
                        line.last_issued_at = now
                    db.session.add(line)

        if issued_row_ids or new_records or items_to_issue:
            _touch_wo()
            db.session.commit()

        # autostatus
        if set_status == "done":
            try:
                wo.status = "done"
                _touch_wo()
                db.session.commit()
            except Exception:
                db.session.rollback()

        # create invoice/batch if there are new records
        if new_records:
            try:
                inv_no = _reserve_invoice_number()
                batch = IssuedBatch(
                    invoice_number=inv_no,
                    issued_to=wo.technician_name,
                    issued_by=getattr(current_user, "username", "system"),
                    reference_job=wo.canonical_job,
                    issue_date=issue_date,
                    location=None,
                    work_order_id=wo.id,
                )
                db.session.add(batch)
                db.session.flush()

                for r in new_records:
                    r.batch_id = batch.id
                    r.invoice_number = inv_no

                _touch_wo()
                db.session.commit()

                params = urlencode({
                    "invoice_number": inv_no,
                    "ref_job": (wo.canonical_job or "").strip(),
                })
                session["last_invoice_url"] = f"/invoice/pdf?{params}"

                return redirect(url_for(
                    "inventory.wo_detail",
                    wo_id=wo.id,
                    issued_ids=",".join(map(str, issued_row_ids))
                ))
            except Exception:
                db.session.rollback()
                # fallback below

        # grouped fallback
        d = (issue_date or datetime.utcnow()).date().isoformat()
        params = urlencode({
            "start_date": d,
            "end_date": d,
            "recipient": wo.technician_name,
            "reference_job": wo.canonical_job,
        })
        link = f"/reports_grouped?{params}"

        if skipped_rows:
            session["wo_skip_info"] = skipped_rows

        flash(Markup(
            f'Issued {len(issued_row_ids)} line(s). '
            f'<a href="{link}" target="_blank" rel="noopener">Open invoice group</a> to print.'
        ), "success")

        return redirect(url_for(
            "inventory.wo_detail",
            wo_id=wo.id,
            issued_ids=",".join(map(str, issued_row_ids))
        ))

    # === 3) Not selected mode (mass in-stock) — unchanged ===
    pn_issue_map = {}
    items_to_issue = []

    for r in avail_rows:
        issue_now = int(r.get("issue_now") or 0)
        if issue_now <= 0:
            continue

        pn = (r.get("part_number") or "").strip().upper()
        if not pn:
            continue

        pn_issue_map[pn] = pn_issue_map.get(pn, 0) + issue_now

        part = Part.query.filter(func.upper(Part.part_number) == pn).first()
        if not part:
            continue

        try:
            real_cost = float(part.unit_cost or 0.0)
        except Exception:
            real_cost = 0.0

        items_to_issue.append({
            "part_id": part.id,
            "qty": issue_now,
            "unit_price": real_cost,
        })

    if not items_to_issue:
        flash("Nothing available to issue (all WAIT).", "warning")
        return redirect(url_for("inventory.wo_detail", wo_id=wo_id))

    issue_date, created_records = _issue_records_bulk(
        issued_to=wo.technician_name,
        reference_job=wo.canonical_job,
        items=items_to_issue,
    )

    now = datetime.utcnow()
    for pn, need_to_apply in pn_issue_map.items():
        if need_to_apply <= 0:
            continue

        rows = (
            WorkOrderPart.query
            .filter(
                WorkOrderPart.work_order_id == wo.id,
                func.upper(WorkOrderPart.part_number) == pn
            )
            .order_by(WorkOrderPart.id.asc())
            .all()
        )

        for line in rows:
            if need_to_apply <= 0:
                break

            qty_needed = int(line.quantity or 0)
            issued_so_far = int(getattr(line, "issued_qty", 0) or 0)
            remaining = max(qty_needed - issued_so_far, 0)
            if remaining <= 0:
                continue

            delta = min(remaining, need_to_apply)
            line.issued_qty = issued_so_far + delta

            if line.issued_qty >= qty_needed:
                try:
                    line.status = "done"
                except Exception:
                    pass

            if hasattr(line, "last_issued_at"):
                line.last_issued_at = now

            db.session.add(line)
            need_to_apply -= delta

    _touch_wo()
    inv_no = _reserve_invoice_number()

    batch = IssuedBatch(
        invoice_number=inv_no,
        issued_to=wo.technician_name,
        issued_by=current_user.username,
        reference_job=wo.canonical_job,
        issue_date=issue_date,
        location=None,
        work_order_id=wo.id,  # 🔑 КЛЮЧ
    )
    db.session.add(batch)
    db.session.flush()

    for r in created_records:
        r.batch_id = batch.id
        r.invoice_number = inv_no
    db.session.commit()

    if set_status == "done":
        try:
            wo.status = "done"
            _touch_wo()
            db.session.commit()
        except Exception:
            db.session.rollback()

    else:
        try:
            still_wait = any(
                (row.get("on_hand", 0) < row.get("requested", 0))
                for row in (compute_availability(wo) or [])
            )
        except Exception:
            still_wait = True

        if not still_wait:
            try:
                wo.status = "done"
                _touch_wo()
                db.session.commit()
            except Exception:
                db.session.rollback()

    d = (issue_date or datetime.utcnow()).date().isoformat()
    params = urlencode({
        "start_date": d,
        "end_date": d,
        "recipient": wo.technician_name,
        "reference_job": wo.canonical_job,
    })
    link = f"/reports_grouped?{params}"

    flash(Markup(
        f'Issued in-stock items. '
        f'<a href="{link}" target="_blank" rel="noopener">Open invoice group</a> to print.'
    ), "success")

    return redirect(url_for("inventory.wo_detail", wo_id=wo.id))

@inventory_bp.get("/work_orders/<int:wo_id>/edit", endpoint="wo_edit")
@login_required
def wo_edit(wo_id: int):
    from flask import render_template, flash, redirect, url_for, session, request
    from flask_login import current_user
    from sqlalchemy import func, or_
    from collections import defaultdict

    from extensions import db
    from models import WorkOrder, WorkUnit, WorkOrderPart, IssuedPartRecord, IssuedBatch

    role = (getattr(current_user, "role", "") or "").strip().lower()
    readonly_param = request.args.get("readonly", type=int) == 1
    readonly = (role not in ("admin", "superadmin")) or readonly_param
    if not readonly and role not in ("admin", "superadmin"):
        flash("Access denied", "danger")
        return redirect(url_for("inventory.wo_list"))

    wo = WorkOrder.query.get_or_404(wo_id)

    technicians = _query_technicians()

    # --- preselect technician ---
    selected_tech_id = None
    selected_tech_username = None
    if getattr(wo, "technician_id", None):
        selected_tech_id = int(wo.technician_id)
        for tid, uname in technicians:
            if tid == selected_tech_id:
                selected_tech_username = uname
                break
    if selected_tech_id is None:
        wo_name = (wo.technician_name or "").strip().lower()
        if wo_name:
            for tid, uname in technicians:
                if (uname or "").strip().lower() == wo_name:
                    selected_tech_id = tid
                    selected_tech_username = uname
                    break

    # ==================================================
    # 1) Фактическая выдача по этому WO (как в wo_detail)
    # ==================================================
    canon = (wo.canonical_job or "").strip()
    issued_items = []
    if canon:
        base_q = (
            db.session.query(IssuedPartRecord)
            .outerjoin(IssuedBatch, IssuedBatch.id == IssuedPartRecord.batch_id)
        ).filter(
            or_(
                func.trim(IssuedPartRecord.reference_job) == canon,
                func.trim(IssuedPartRecord.reference_job).like(f"%{canon}%"),
                func.trim(IssuedBatch.reference_job) == canon,
            )
        )
        issued_items = base_q.all()

    issued_qty_by_pn = defaultdict(int)
    for rec in issued_items:
        pn = (getattr(rec.part, "part_number", "") or "").strip().upper()
        if not pn:
            continue
        q = int(rec.quantity or 0)
        if q == 0:
            continue
        issued_qty_by_pn[pn] += q

    remaining_issued = dict(issued_qty_by_pn)

    # ==================================================
    # 2) units payload c issued_qty + ALT PN + INS-флагом
    # ==================================================
    units = []
    for u in (getattr(wo, "units", []) or []):
        rows = []
        for p in (getattr(u, "parts", []) or []):
            alt_val = (
                getattr(p, "alt_numbers", "") or
                getattr(p, "alt_part_numbers", "") or
                ""
            )

            pn_upper = (getattr(p, "part_number", "") or "").strip().upper()
            qty_plan = int(getattr(p, "quantity", 0) or 0)

            issued_left = int(remaining_issued.get(pn_upper, 0) or 0)
            assigned_issued = 0
            if qty_plan > 0 and issued_left > 0:
                assigned_issued = min(qty_plan, issued_left)
                remaining_issued[pn_upper] = issued_left - assigned_issued

            is_ordered_flag = getattr(p, "ordered_flag", None)
            is_ordered = bool(
                is_ordered_flag
                or (str(getattr(p, "status", "")).strip().lower() == "ordered")
                or (str(getattr(p, "line_status", "")).strip().lower() == "ordered")
            )

            rows.append({
                "id": getattr(p, "id", None),
                "part_number": getattr(p, "part_number", "") or "",
                "part_name": getattr(p, "part_name", "") or "",
                "quantity": qty_plan,

                "alt_numbers": alt_val,

                "warehouse": getattr(p, "warehouse", "") or "",
                "supplier": getattr(p, "supplier", "") or "",
                "backorder_flag": bool(getattr(p, "backorder_flag", False)),
                "line_status": getattr(p, "line_status", "") or "search_ordered",
                "unit_cost": (
                    float(getattr(p, "unit_cost"))
                    if getattr(p, "unit_cost") is not None else None
                ),

                "issued_qty": assigned_issued,
                "is_ordered": is_ordered,

                "is_insurance_supplied": bool(
                    getattr(p, "is_insurance_supplied", False)
                ),
                "invoice_number": (getattr(p, "invoice_number", "") or ""),

            })

        if not rows:
            rows = [{
                "id": None,
                "part_number": "", "part_name": "", "quantity": 1,
                "alt_numbers": "",
                "warehouse": "", "supplier": "",
                "backorder_flag": False, "line_status": "search_ordered",
                "unit_cost": 0.0,
                "issued_qty": 0,
                "is_ordered": False,
                "is_insurance_supplied": False,
            }]

        units.append({
            "id": getattr(u, "id", None),
            "brand": getattr(u, "brand", "") or "",
            "model": getattr(u, "model", "") or "",
            "serial": getattr(u, "serial", "") or "",
            "rows": rows,
        })

    if not units:
        units = [{
            "brand":  wo.brand or "",
            "model":  wo.model or "",
            "serial": wo.serial or "",
            "rows": [{
                "id": None,
                "part_number": "", "part_name": "", "quantity": 1,
                "alt_numbers": "",
                "warehouse": "", "supplier": "",
                "backorder_flag": False, "line_status": "search_ordered",
                "unit_cost": 0.0,
                "issued_qty": 0,
                "is_ordered": False,
                "is_insurance_supplied": False,
            }],
        }]

    recent_suppliers = session.get("recent_suppliers", []) or []

    return render_template(
        "wo_form_units.html",
        wo=wo,
        units=units,
        recent_suppliers=recent_suppliers,
        readonly=readonly,
        technicians=technicians,
        selected_tech_id=selected_tech_id,
        selected_tech_username=selected_tech_username,
    )

@inventory_bp.post("/work_orders/<int:wo_id>/delete", endpoint="wo_delete")
@login_required
def wo_delete(wo_id: int):
    if getattr(current_user, "role", "") != "superadmin":
        flash("Access denied", "danger")
        return redirect(url_for("inventory.wo_detail", wo_id=wo_id))

    def table_exists(name: str) -> bool:
        try:
            row = db.session.execute(
                text("SELECT name FROM sqlite_master WHERE type='table' AND name=:n"),
                {"n": name}
            ).fetchone()
            return bool(row)
        except Exception:
            return False

    try:
        # 1) Пробуем через модели (если у тебя они есть)
        try:
            from models import WorkOrder  # подстрой, если модуль другой
            wo = WorkOrder.query.get_or_404(wo_id)
            db.session.delete(wo)
            db.session.commit()
        except Exception as model_err:
            current_app.logger.debug("Model delete failed, fallback to SQL: %s", model_err)

            # 2) Фоллбэк на SQL — удаляем дочерние записи, но только если таблицы существуют
            for t in ("issued_part_records", "work_order_parts", "issued_batches", "job_index"):
                if table_exists(t):
                    db.session.execute(text(f"DELETE FROM {t} WHERE work_order_id=:id"), {"id": wo_id})

            if table_exists("work_orders"):
                db.session.execute(text("DELETE FROM work_orders WHERE id=:id"), {"id": wo_id})

            db.session.commit()

        flash(f"Work Order #{wo_id} deleted.", "success")
        return redirect(url_for("inventory.wo_list"))

    except Exception as e:
        db.session.rollback()
        current_app.logger.exception("Failed to delete work order %s", wo_id)
        flash("Failed to delete work order.", "danger")
        return redirect(url_for("inventory.wo_detail", wo_id=wo_id))

def _parse_units_form(form):
    """
    Ожидаем схему имён:
      units[U][brand], units[U][model], units[U][serial]
      units[U][rows][I][part_number], ... quantity, alt_numbers, supplier, backorder_flag, line_status
    """
    import re
    units = {}
    for k, v in form.items():
        m = re.match(r"^units\[(\d+)\]\[(brand|model|serial)\]$", k)
        if m:
            ui = int(m.group(1))
            units.setdefault(ui, {"rows": []})
            units[ui][m.group(2)] = v.strip()
            continue

        m = re.match(r"^units\[(\d+)\]\[rows\]\[(\d+)\]\[(\w+)\]$", k)
        if m:
            ui = int(m.group(1)); ri = int(m.group(2)); field = m.group(3)
            while len(units.setdefault(ui, {"rows": []})["rows"]) <= ri:
                units[ui]["rows"].append({})
            units[ui]["rows"][ri][field] = v.strip()
            continue

    # чекбоксы backorder_flag могут не прийти — нормализуем
    for u in units.values():
        for row in u["rows"]:
            row["quantity"] = int(row.get("quantity") or 0)
            row["backorder_flag"] = bool(row.get("backorder_flag") in ("on", "true", "1"))
            row["line_status"] = (row.get("line_status") or "search_ordered")
    # превращаем в упорядоченный список
    out = [units[i] for i in sorted(units.keys())]
    return out
@inventory_bp.get("/api/part_lookup")
@login_required
def api_part_lookup():
    from sqlalchemy import func
    from models import Part
    pn = (request.args.get("pn") or "").strip().upper()
    if not pn:
        return jsonify({"found": False})
    part = Part.query.filter(func.upper(Part.part_number) == pn).first()
    if not part:
        return jsonify({"found": False, "stock_hint": "—"})
    on_hand = int(getattr(part, "on_hand", 0) or getattr(part, "quantity", 0) or 0)
    stock_hint = "STOCK" if on_hand > 0 else "WAIT"
    wh = getattr(part, "location", None) or getattr(part, "wh", None) or ""
    return jsonify({"found": True, "name": getattr(part, "name", "") or "", "wh": wh, "stock_hint": stock_hint})

@inventory_bp.post("/work_orders/<int:wo_id>/units/<int:unit_id>/issue_instock")
@login_required
def wo_issue_instock_unit(wo_id, unit_id):
    if getattr(current_user, "role", "") not in ("admin", "superadmin"):
        flash("Access denied", "danger")
        return redirect(url_for("inventory.wo_detail", wo_id=wo_id))

    from datetime import datetime, timedelta
    from sqlalchemy import and_, func
    from urllib.parse import urlencode
    from flask import request, session
    from extensions import db

    from models import WorkOrder, Part, IssuedPartRecord, IssuedBatch, WorkOrderPart


    wo = WorkOrder.query.get_or_404(wo_id)
    unit = next((u for u in (wo.units or []) if u.id == unit_id), None)
    if not unit:
        flash("Unit not found", "danger")
        return redirect(url_for("inventory.wo_detail", wo_id=wo_id))

    # флаг автосмены статуса
    set_status = (request.form.get("set_status") or "").strip().lower()

    rows = compute_availability_unit(unit, wo.status)

    items = []
    # --- map PN -> INV# из WorkOrderPart.invoice_number ---
    pn_to_inv = {}
    for p in (wo.parts or []):
        pn_key = (getattr(p, "part_number", "") or "").strip().upper()
        inv = (getattr(p, "invoice_number", "") or "").strip()
        if pn_key and inv:
            pn_to_inv[pn_key] = inv[:32]

    for r in rows:
        if int(r.get("issue_now", 0)) > 0:
            pn = (r.get("part_number") or "").strip().upper()
            if not pn:
                continue

            part = Part.query.filter_by(part_number=pn).first()
            if not part:
                continue

            # ✅ фиксируем себестоимость склада в момент выдачи
            try:
                real_cost = float(part.unit_cost or 0.0)
            except Exception:
                real_cost = 0.0

            # берём INV# прямо из WorkOrderPart для этой unit + pn
            wop_inv = ""
            try:
                wop = next(
                    (p for p in (wo.parts or [])
                     if p.unit_id == unit_id and (p.part_number or "").strip().upper() == pn),
                    None
                )
                wop_inv = (getattr(wop, "invoice_number", "") or "").strip()
            except Exception:
                wop_inv = ""

            # INV# берём из WorkOrderPart.invoice_number по unit_id + part_number
            wop_inv = ""
            try:
                wop = next(
                    (p for p in (wo.parts or [])
                     if p.unit_id == unit_id and (p.part_number or "").strip().upper() == pn),
                    None
                )
                wop_inv = (getattr(wop, "invoice_number", "") or "").strip()
            except Exception:
                wop_inv = ""

            items.append({
                "part_id": part.id,
                "qty": int(r["issue_now"]),
                "unit_price": real_cost,
                "inv_ref": wop_inv,  # ✅ вот это критично
            })

    if not items:
        flash("Nothing available to issue for this unit.", "warning")
        return redirect(url_for("inventory.wo_detail", wo_id=wo_id))

    # --- создаём строки выдачи ---
    issue_date, maybe_records = _issue_records_bulk(
        issued_to=wo.technician_name,
        reference_job=wo.canonical_job,
        items=items
    )

    # поддержка обоих вариантов возврата (_issue_records_bulk может вернуть count)
    if isinstance(maybe_records, (list, tuple)) and maybe_records and hasattr(maybe_records[0], "id"):
        new_records = list(maybe_records)
    else:
        # fallback: собрать только что созданные строки по окну +/-2 минуты
        t0 = issue_date - timedelta(minutes=2)
        t1 = issue_date + timedelta(minutes=2)
        new_records = IssuedPartRecord.query.filter(
            and_(
                IssuedPartRecord.issued_to == wo.technician_name,
                IssuedPartRecord.reference_job == wo.canonical_job,
                IssuedPartRecord.invoice_number.is_(None),
                IssuedPartRecord.issue_date >= t0,
                IssuedPartRecord.issue_date <= t1,
            )
        ).all()

    if not new_records:
        # даже если не нашли "new_records" (крайне редко), мы всё равно можем завершить
        if set_status == "done":
            try:
                wo.status = "done"
                db.session.commit()
            except Exception:
                db.session.rollback()
        flash("Issued items saved, but could not collect records for invoice.", "warning")
        return redirect(url_for("inventory.wo_detail", wo_id=wo_id))

    # --- формируем инвойс/батч БЕЗ внешних хелперов ---
    try:
        inv_no = _reserve_invoice_number()

        batch = IssuedBatch(
            invoice_number=inv_no,
            issued_to=wo.technician_name,
            issued_by=getattr(current_user, "username", "system"),
            reference_job=wo.canonical_job,
            issue_date=issue_date,
            location=None,
        )
        db.session.add(batch)
        db.session.flush()  # чтобы появился batch.id

        for r in new_records:
            r.batch_id = batch.id
            r.invoice_number = inv_no

        db.session.commit()
    except Exception as e:
        db.session.rollback()
        if set_status == "done":
            try:
                wo.status = "done"
                db.session.commit()
            except Exception:
                db.session.rollback()
        flash(f"Error creating invoice batch: {e}", "danger")
        return redirect(url_for("inventory.wo_detail", wo_id=wo_id))

    # --- обновляем статус WO ---
    if set_status == "done":
        try:
            wo.status = "done"
            db.session.commit()
        except Exception:
            db.session.rollback()
    else:
        try:
            avail_all = compute_availability_multi(wo)
            still_wait = any(
                int(x.get("on_hand", 0)) < int(x.get("requested", 0))
                for x in avail_all
            )
            if not still_wait:
                wo.status = "done"
                db.session.commit()
        except Exception:
            db.session.rollback()

    # --- сохраняем URL инвойса и возвращаемся на страницу WO ---
    params = urlencode({
        "invoice_number": batch.invoice_number,
        "ref_job": (wo.canonical_job or "").strip(),
    })
    session["last_invoice_url"] = f"/invoice/pdf?{params}"

    return redirect(url_for("inventory.wo_detail", wo_id=wo_id))

@inventory_bp.post("/work_orders/<int:wo_id>/status")
@login_required
def wo_set_status(wo_id):
    if getattr(current_user, "role", "") not in ("admin", "superadmin"):
        flash("Access denied", "danger")
        return redirect(url_for("inventory.wo_detail", wo_id=wo_id))
    new_status = (request.form.get("status") or "").strip()
    if new_status not in ("search_ordered","ordered","done"):
        flash("Invalid status", "warning")
        return redirect(url_for("inventory.wo_detail", wo_id=wo_id))
    wo = WorkOrder.query.get_or_404(wo_id)
    wo.status = new_status
    db.session.commit()
    flash(f"Status set to {new_status}.", "success")
    return redirect(url_for("inventory.wo_detail", wo_id=wo_id))



@inventory_bp.post("/issue/batch")
@login_required
def issue_batch():
    # 🔐 только admin/superadmin
    if getattr(current_user, "role", "") not in ("admin", "superadmin"):
        return jsonify({"ok": False, "error": "Access denied"}), 403
    try:
        payload = request.get_json(force=True, silent=False) or {}
        issued_to = (payload.get("issued_to") or "").strip()
        reference_job = (payload.get("reference_job") or "").strip()
        items = payload.get("items") or []

        if not issued_to:
            return jsonify({"ok": False, "error": "issued_to is required"}), 400
        if not reference_job:
            return jsonify({"ok": False, "error": "reference_job is required"}), 400

        # 🧾 логируем состав выдачи (без персональных данных)
        try:
            safe_items = [
                {"part_id": int(i.get("part_id", 0)), "qty": int(i.get("qty", 0)),
                 "unit_price": None if i.get("unit_price") in (None, "") else float(i.get("unit_price"))}
                for i in items
            ]
            current_app.logger.info("ISSUE_BATCH by=%s job=%s items=%s",
                                    getattr(current_user, "username", "?"),
                                    reference_job, safe_items)
        except Exception:
            pass

        issue_date, created = _issue_records_bulk(issued_to, reference_job, items)
        today = issue_date.date().isoformat()
        params = urlencode({
            "start_date": today, "end_date": today,
            "recipient": issued_to, "reference_job": reference_job
        })
        return jsonify({
            "ok": True,
            "created": created,   # ← для тоста
            "redirect": f"/reports_grouped?{params}"
        }), 200

    except ValueError as ve:
        db.session.rollback()
        return jsonify({"ok": False, "error": str(ve)}), 400
    except SQLAlchemyError:
        db.session.rollback()
        return jsonify({"ok": False, "error": "DB error"}), 500
    except Exception as e:
        db.session.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500


@inventory_bp.post("/issue/line/<int:part_id>")
@login_required
def issue_line(part_id):
    """
    Выдать одну позицию (когда нужно выдать конкретный Part сейчас).
    Принимает JSON или form-data с полями:
      - qty (int) — обязательное, > 0
      - issued_to (str) — обязательное
      - reference_job (str) — обязательное
      - unit_price (float) — опционально (если уже учтены доставка/наценка)
    Возвращает:
      - JSON {"ok": True, "created": 1, "redirect": "..."} для JSON-запросов
      - redirect на /reports_grouped для форм
    """
    # 🔐 Разрешаем только admin / superadmin
    if getattr(current_user, "role", "") not in ("admin", "superadmin"):
        # Для JSON-запросов — JSON, для форм — flash + редирект
        if request.is_json:
            return jsonify({"ok": False, "error": "Access denied"}), 403
        flash("Access denied", "danger")
        return redirect(url_for("inventory.dashboard"))

    try:
        # Универсально читаем данные (JSON или форма)
        data = (request.get_json(silent=True) or {}) if request.is_json else (request.form or {})
        qty = int(data.get("qty") or 0)
        issued_to = (data.get("issued_to") or "").strip()
        reference_job = (data.get("reference_job") or "").strip()

        # unit_price — опционально
        unit_price_raw = data.get("unit_price")
        unit_price = None
        if unit_price_raw not in (None, ""):
            unit_price = float(unit_price_raw)

        # Валидация
        if qty <= 0:
            raise ValueError("qty must be > 0")
        if not issued_to or not reference_job:
            raise ValueError("issued_to and reference_job are required")

        # 🧾 Лог выдачи одной строки (без чувствительных данных)
        try:
            current_app.logger.info(
                "ISSUE_LINE by=%s job=%s part_id=%s qty=%s price=%s",
                getattr(current_user, "username", "?"),
                reference_job, part_id, qty, unit_price
            )
        except Exception:
            pass

        # Выдаём ровно одну позицию
        items = [{"part_id": int(part_id), "qty": qty, "unit_price": unit_price}]
        issue_date, _created = _issue_records_bulk(issued_to, reference_job, items)  # min(qty, on_hand) внутри

        # Готовим редирект на сгруппированный отчёт за сегодня
        today = issue_date.date().isoformat()
        params = urlencode({
            "start_date": today,
            "end_date": today,
            "recipient": issued_to,
            "reference_job": reference_job
        })
        target = f"/reports_grouped?{params}"

        if request.is_json:
            return jsonify({"ok": True, "created": 1, "redirect": target}), 200
        else:
            return redirect(target, code=303)

    except ValueError as ve:
        db.session.rollback()
        if request.is_json:
            return jsonify({"ok": False, "error": str(ve)}), 400
        flash(f"Issue failed: {ve}", "danger")
        return redirect(url_for("inventory.dashboard"))

    except Exception as e:
        db.session.rollback()
        if request.is_json:
            return jsonify({"ok": False, "error": "Internal error"}), 500
        current_app.logger.exception("ISSUE_LINE failed: %s", e)
        flash("Issue failed: internal error", "danger")
        return redirect(url_for("inventory.dashboard"))

# ----------------- Dashboard -----------------

@inventory_bp.route('/api/part/<part_number>', methods=['GET'])
@login_required
def get_part_by_number(part_number):
    """
    Вернёт информацию о детали по Part #.
    Ответ (200):
      {
        "id": int,
        "name": str,
        "location": str | null,
        "unit_cost": float,
        "quantity": int
      }
    Или (404): {"error": "Not found"}
    """
    try:
        pn = (part_number or "").strip().upper()
        if not pn:
            return jsonify({"error": "Part number is required"}), 400

        part = Part.query.filter_by(part_number=pn).first()
        if not part:
            return jsonify({"error": "Not found"}), 404

        # Нормализуем типы в ответе
        return jsonify({
            "id": int(part.id),
            "name": part.name or "",
            "location": part.location,
            "unit_cost": float(part.unit_cost or 0.0),
            "quantity": int(part.quantity or 0),
        }), 200

    except Exception as e:
        current_app.logger.exception("api/part failed for %r: %s", part_number, e)
        return jsonify({"error": "Internal error"}), 500




@inventory_bp.route('/api/part_lookup')
@login_required
def part_lookup():
    part_number = request.args.get('part_number', '').strip().upper()
    part = Part.query.filter_by(part_number=part_number).first()
    if part:
        return {
            'found': True,
            'id': part.id,
            'name': part.name,
            'quantity': part.quantity
        }
    return { 'found': False }

@inventory_bp.route('/')
@login_required
def dashboard():
    search_query = request.args.get('search', '').strip()
    parts = Part.query

    if search_query:
        parts = parts.filter(
            or_(
                Part.part_number.ilike(f"%{search_query}%"),
                Part.name.ilike(f"%{search_query}%")
            )
        )

    parts = parts.all()
    return render_template('index.html', parts=parts, search_query=search_query)

# @inventory_bp.route('/inventory_summary')
# @login_required
# def inventory_summary():
#     parts = Part.query.filter(Part.quantity > 0).all()
#
#     locations = defaultdict(lambda: {
#         'total_quantity': 0,
#         'total_value': 0.0,
#     })
#
#     for part in parts:
#         loc = part.location or 'Unknown'
#         locations[loc]['total_quantity'] += part.quantity
#         locations[loc]['total_value'] += part.quantity * part.unit_cost
#
#     grand_total_quantity = sum(data['total_quantity'] for data in locations.values())
#     grand_total_value = sum(data['total_value'] for data in locations.values())
#
#     return render_template('inventory_summary.html',
#                            locations=locations,
#                            grand_total_quantity=grand_total_quantity,
#                            grand_total_value=grand_total_value)

@inventory_bp.route('/dashboard/location_report', methods=['GET'], endpoint='location_report')
@login_required
def location_report():
    current_location_raw = (request.args.get('loc') or '').strip()
    norm = current_location_raw.lower()

    base_q = Part.query.filter(Part.quantity > 0)

    # ---- datalist всех локаций (уникальные, с красивыми названиями) ----
    loc_rows = (Part.query.with_entities(Part.location)
                .filter(Part.quantity > 0)
                .distinct().all())
    seen = set()
    all_locations = []
    for (loc,) in loc_rows:
        label = (loc or 'Unknown')
        key = (label.strip().lower() or 'unknown')
        if key not in seen:
            seen.add(key)
            all_locations.append(label)
    all_locations.sort()

    # ---- выборка по локации ----
    if norm:
        # Unknown-группа
        if norm in ('unknown', '—', '-', 'none', 'null', ''):
            parts = base_q.filter(or_(Part.location == None, Part.location == '')).all()
        else:
            # 1) точное совпадение без регистра
            exact = (base_q.filter(
                func.lower(func.coalesce(Part.location, '')) == norm
            ).all())

            if exact:
                parts = exact
            else:
                # 2) если пользователь явно использовал wildcard (*), делаем LIKE по шаблону
                if '*' in current_location_raw:
                    like_pat = norm.replace('*', '%')
                    parts = (base_q.filter(
                        func.lower(func.coalesce(Part.location, '')).like(like_pat)
                    ).all())
                else:
                    # 3) префиксный поиск (без регистра): norm%
                    #   (чтобы 'rel' не ловил всё подряд как contains)
                    parts = (base_q.filter(
                        func.lower(func.coalesce(Part.location, '')).like(f"{norm}%")
                    ).all())
    else:
        parts = base_q.all()

    # ---- группировка и итоги ----
    locations = defaultdict(lambda: {'parts': [], 'total_quantity': 0, 'total_value': 0.0})
    for part in parts:
        loc = (part.location or 'Unknown').strip() or 'Unknown'
        qty = int(part.quantity or 0)
        cost = float(part.unit_cost or 0.0)
        locations[loc]['parts'].append(part)
        locations[loc]['total_quantity'] += qty
        locations[loc]['total_value']    += qty * cost

    grand_total_quantity = sum(x['total_quantity'] for x in locations.values()) if locations else 0
    grand_total_value    = sum(x['total_value'] for x in locations.values()) if locations else 0.0

    return render_template(
        'location_report.html',
        locations=locations,
        all_locations=all_locations,
        current_location=current_location_raw,
        grand_total_quantity=grand_total_quantity,
        grand_total_value=grand_total_value
    )
# --- Печать детального отчёта по локациям (весь список товаров) ---
@inventory_bp.route('/dashboard/location_report/print')
@login_required
def print_location_report():
    parts = Part.query.filter(Part.quantity > 0).all()

    locations = defaultdict(lambda: {
        'parts': [],
        'total_quantity': 0,
        'total_value': 0.0,
    })

    for part in parts:
        loc = part.location or 'Unknown'
        locations[loc]['parts'].append(part)
        locations[loc]['total_quantity'] += part.quantity
        locations[loc]['total_value'] += part.quantity * part.unit_cost

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    elements = []

    styles = getSampleStyleSheet()
    title_style = styles['Heading2']

    for loc, data in locations.items():
        elements.append(Paragraph(f"Location: {loc}", title_style))
        elements.append(Paragraph(f"Total Quantity: {data['total_quantity']}", styles['Normal']))
        elements.append(Paragraph(f"Total Value: ${data['total_value']:.2f}", styles['Normal']))
        elements.append(Spacer(1, 12))

        table_data = [["Part Number", "Name", "Quantity", "Unit Cost", "Total Cost"]]
        for part in data['parts']:
            table_data.append([
                part.part_number,
                part.name,
                str(part.quantity),
                f"${part.unit_cost:.2f}",
                f"${part.quantity * part.unit_cost:.2f}"
            ])
        # Итог по локации
        table_data.append([
            "Location Total", "",
            str(data['total_quantity']), "", f"${data['total_value']:.2f}"
        ])

        table = Table(table_data, hAlign='LEFT')
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightblue),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.black),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 20))

    doc.build(elements)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name="location_report.pdf", mimetype='application/pdf')

# --- Печать сводного отчёта (итог по локациям без деталей) ---
@inventory_bp.route('/dashboard/inventory_summary/print')
@login_required
def print_inventory_summary():
    parts = Part.query.filter(Part.quantity > 0).all()

    locations = defaultdict(lambda: {
        'total_quantity': 0,
        'total_value': 0.0,
    })

    for part in parts:
        loc = part.location or 'Unknown'
        locations[loc]['total_quantity'] += part.quantity
        locations[loc]['total_value'] += part.quantity * part.unit_cost

    grand_total_quantity = sum(data['total_quantity'] for data in locations.values())
    grand_total_value = sum(data['total_value'] for data in locations.values())

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []

    styles = getSampleStyleSheet()
    elements.append(Paragraph("Inventory Summary Report by Location", styles['Heading2']))
    elements.append(Spacer(1, 12))

    table_data = [["Location", "Total Quantity", "Total Value"]]
    for loc, data in locations.items():
        table_data.append([
            loc,
            str(data['total_quantity']),
            f"${data['total_value']:.2f}"
        ])
    # Итог
    table_data.append([
        "Grand Total",
        str(grand_total_quantity),
        f"${grand_total_value:.2f}"
    ])

    table = Table(table_data, hAlign='LEFT')
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.black),
    ]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name="inventory_summary.pdf", mimetype='application/pdf')



@inventory_bp.route('/add', methods=['GET', 'POST'])
@login_required
def add_part():
    if request.method == 'POST':
        part_number = request.form['part_number'].strip().upper()
        name = request.form['name'].strip().upper()
        quantity = int(request.form['quantity'])
        unit_cost = float(request.form['unit_cost'])
        location = request.form['location'].strip().upper()

        existing = Part.query.filter_by(part_number=part_number).first()
        if existing:
            existing.quantity += quantity
            existing.unit_cost = unit_cost
            existing.name = name
            existing.location = location
        else:
            new_part = Part(
                name=name,
                part_number=part_number,
                quantity=quantity,
                unit_cost=unit_cost,
                location=location
            )
            db.session.add(new_part)

        db.session.commit()
        flash('Part saved successfully.', 'success')
        return redirect(url_for('inventory.dashboard'))

    return render_template('add_part.html')


# ----------------- Issue Part -----------------
@inventory_bp.route('/issue', methods=['GET', 'POST'])
@login_required
def issue_part():
    from datetime import datetime
    from flask import request, render_template, redirect, url_for, flash
    from flask_login import current_user
    from urllib.parse import urlencode

    from extensions import db
    from models import Part, IssuedPartRecord

    # ---------- helper: touch Work Order updated_by/updated_at ----------
    def _touch_work_order_from_ref(ref_job: str, ts: datetime, user_id: int) -> None:
        """
        Best-effort: if reference_job corresponds to a Work Order,
        set updated_by_id + updated_at so header shows correct user.
        Safe: wrapped with try/except, no crash if schema differs.
        """
        ref = (ref_job or "").strip()
        if not ref:
            return

        # 1) Try via JobIndex (your architecture)
        try:
            from models import JobIndex  # if exists
            ji = JobIndex.query.filter_by(job_number=ref).first()
            if ji and getattr(ji, "work_order_id", None):
                db.session.execute(
                    db.text("""
                        UPDATE work_orders
                        SET updated_at = :ts,
                            updated_by_id = :uid
                        WHERE id = :wo_id
                    """),
                    {"ts": ts, "uid": user_id, "wo_id": ji.work_order_id}
                )
                return
        except Exception:
            pass

        # 2) Fallback: try direct match on canonical_ref / job numbers (if those columns exist)
        # NOTE: if any of these columns don't exist, SQLite will error -> we catch.
        try:
            db.session.execute(
                db.text("""
                    UPDATE work_orders
                    SET updated_at = :ts,
                        updated_by_id = :uid
                    WHERE canonical_ref = :ref
                       OR job_num_a = :ref
                       OR job_num_b = :ref
                """),
                {"ts": ts, "uid": user_id, "ref": ref}
            )
        except Exception:
            # last fallback: do nothing
            return

    # этот список для отдельной страницы issue_part, можно оставить как есть
    parts = Part.query.filter(Part.quantity > 0).all()

    if request.method == 'POST':
        import json
        try:
            all_parts = json.loads(request.form.get('all_parts_json', '[]'))
        except json.JSONDecodeError:
            flash('Invalid part data.', 'danger')
            return redirect(url_for('.issue_part'))

        if not all_parts:
            flash('No parts to issue.', 'warning')
            return redirect(url_for('.issue_part'))

        new_records: list[IssuedPartRecord] = []

        # шапка батча — из первой строки
        first = all_parts[0]
        issued_to = (first.get('recipient') or '').strip()
        reference_job = (first.get('reference_job') or '').strip() or None
        issued_by = current_user.username
        issue_dt = datetime.utcnow()

        # batch-level location (опционально — по первой строке)
        batch_location = None

        # 1) валидируем сток и создаём строки выдачи
        for idx, item in enumerate(all_parts):
            part_id = item.get('part_id')
            part = Part.query.get(part_id) if part_id else None
            qty = int(item.get('quantity') or 0)

            # тип работы и INS-флаг (если не придут — будет False и логика как раньше)
            job_type = (item.get('job_type') or '').strip().upper()
            is_ins_flag = bool(
                item.get('is_ins')
                or item.get('is_insurance_supplied')
                or item.get('ins_flag')
            )
            is_insurance_case = (job_type == 'INSURANCE') and is_ins_flag

            if not part or qty <= 0:
                flash(f"Invalid part or quantity for item #{idx + 1}.", 'danger')
                return redirect(url_for('.issue_part'))

            # --- ОСОБЫЙ СЛУЧАЙ: страховая работа + INS-строка ---
            if is_insurance_case:
                # НИКАКИХ проверок склада и НИКАКОГО списания количества
                loc_snapshot = (getattr(part, "location", "") or "").strip()

                rec = IssuedPartRecord(
                    part_id=part.id,
                    quantity=qty,
                    issued_to=issued_to,
                    reference_job=reference_job,
                    issued_by=issued_by,
                    issue_date=issue_dt,
                    unit_cost_at_issue=0.0,          # INS supplied => 0 cost (как обычно делаем)
                    location="INS",                  # явный snapshot
                )
                db.session.add(rec)
                new_records.append(rec)

                if batch_location is None and loc_snapshot:
                    batch_location = loc_snapshot

                continue

            # --- Обычный случай: не страховая/не INS-строка ---
            if part.quantity < qty:
                flash(
                    f"Not enough stock for {item.get('part_number', 'UNKNOWN')}: "
                    f"need {qty}, have {part.quantity}.",
                    'danger'
                )
                return redirect(url_for('.issue_part'))

            # уменьшаем сток
            part.quantity -= qty

            # СНИМОК локации на момент выдачи
            loc_snapshot = (getattr(part, "location", "") or "").strip()

            rec = IssuedPartRecord(
                part_id=part.id,
                quantity=qty,
                issued_to=issued_to,
                reference_job=reference_job,
                issued_by=issued_by,
                issue_date=issue_dt,
                unit_cost_at_issue=float(getattr(part, "unit_cost", 0) or 0),
                location=loc_snapshot,  # snapshot
            )
            db.session.add(rec)
            new_records.append(rec)

            if batch_location is None and loc_snapshot:
                batch_location = loc_snapshot

        # 2) сразу создаём батч + номер инвойса
        try:
            _create_batch_for_records(
                records=new_records,
                issued_to=issued_to,
                issued_by=issued_by,
                reference_job=reference_job,
                issue_date=issue_dt,
                location=batch_location,
            )

            # ✅ ВАЖНО: после issue обновляем “кто обновил WO”
            if reference_job:
                _touch_work_order_from_ref(reference_job, issue_dt, current_user.id)

            db.session.commit()
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating invoice batch: {e}', 'danger')
            return redirect(url_for('.issue_part'))

        flash('All parts issued successfully.', 'success')

        # 3) редирект в grouped — инвойс уже существует
        today = issue_dt.date().isoformat()
        params = {'start_date': today, 'end_date': today}
        if issued_to:
            params['recipient'] = issued_to
        if reference_job:
            params['reference_job'] = reference_job

        return redirect('/reports_grouped?' + urlencode(params), code=303)

    return render_template('issue_part.html', parts=parts)


@inventory_bp.get("/issue_ui")
@login_required
def issue_ui():
    from flask_login import current_user
    from models import Part

    parts = Part.query.order_by(Part.part_number).limit(200).all()
    technician_name = getattr(current_user, "username", "TECH")
    canonical_ref = "TESTJOB123"  # подставишь свой JOB

    return render_template(
        "issue_ui.html",
        parts=parts,
        technician_name=technician_name,
        canonical_ref=canonical_ref
    )

@inventory_bp.route('/reports', methods=['GET', 'POST'])
@login_required
def reports():
    query = IssuedPartRecord.query.join(Part)
    start_date = request.form.get('start_date')
    end_date = request.form.get('end_date')
    recipient = request.form.get('recipient')
    reference_job = request.form.get('reference_job')

    if start_date:
        query = query.filter(IssuedPartRecord.issue_date >= start_date)
    if end_date:
        end_dt = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
        query = query.filter(IssuedPartRecord.issue_date <= end_dt)
    if recipient:
        query = query.filter(IssuedPartRecord.issued_to.ilike(f'%{recipient}%'))
    if reference_job:
        query = query.filter(IssuedPartRecord.reference_job.ilike(f'%{reference_job}%'))

    records = query.order_by(IssuedPartRecord.issue_date.desc()).all()

    # Группируем записи по ключу
    invoices_map = defaultdict(lambda: {
        'issued_to': '',
        'reference_job': '',
        'issued_by': '',
        'issue_date': None,
        'items': [],
        'total_value': 0.0,
    })

    grand_total = 0.0

    for r in records:
        key = (r.issued_to, r.reference_job or '', r.issued_by, r.issue_date.date())
        inv = invoices_map[key]
        inv['issued_to'] = r.issued_to
        inv['reference_job'] = r.reference_job
        inv['issued_by'] = r.issued_by
        inv['issue_date'] = r.issue_date
        inv['items'].append(r)
        line_total = r.quantity * r.unit_cost_at_issue
        inv['total_value'] += line_total
        grand_total += line_total

    invoices = list(invoices_map.values())

    return render_template(
        'reports_grouped.html',
        invoices=invoices,
        total=grand_total,
        start_date=start_date,
        end_date=end_date,
        recipient=recipient,
        reference_job=reference_job
    )
# inventory/routes.py
@inventory_bp.route('/reports_grouped', methods=['GET', 'POST'])
@login_required
def reports_grouped():
    from collections import defaultdict
    from datetime import datetime, time, timedelta, timezone
    from zoneinfo import ZoneInfo

    from sqlalchemy.orm import selectinload
    from sqlalchemy import func, or_, case

    from flask import render_template, request
    from flask_login import current_user

    from extensions import db
    from models import IssuedPartRecord, IssuedBatch, Part, utc_to_local

    # ✅ SAFE import: если модель/таблица ещё не существует — отчёт не падает
    try:
        from models import ReturnDestination
    except Exception:
        ReturnDestination = None

    DEFAULT_LIMIT = 50
    LA_TZ = ZoneInfo("America/Los_Angeles")

    def _parse_date_ymd(s: str | None):
        if not s:
            return None
        try:
            return datetime.strptime(s, "%Y-%m-%d")
        except Exception:
            return None

    # ✅ Vendor companies (id->name) for dropdown + stable display
    return_destinations = []
    dest_map = {}
    if ReturnDestination is not None:
        try:
            return_destinations = (
                db.session.query(ReturnDestination)
                .order_by(func.lower(ReturnDestination.name).asc())
                .all()
            )
            dest_map = {int(d.id): (d.name or "") for d in return_destinations}
        except Exception:
            # Никаких падений: просто без списка компаний
            return_destinations = []
            dest_map = {}

    # ---------- Параметры (GET/POST) ----------
    params         = request.values
    start_date_s   = (params.get('start_date') or '').strip()
    end_date_s     = (params.get('end_date') or '').strip()
    recipient_raw  = (params.get('recipient') or '').strip() or None
    reference_job  = (params.get('reference_job') or '').strip() or None
    invoice_s      = (params.get('invoice_number') or params.get('invoice') or params.get('invoice_no') or '').strip()
    invoice_search = invoice_s or None  # общий поиск (и invoice_number, и inv_ref)
    location       = (params.get('location') or '').strip() or None
    status         = (params.get('status') or '').strip().upper()

    # роль/текущий пользователь
    role_low = (getattr(current_user, "role", "") or "").strip().lower()
    me_user  = (getattr(current_user, "username", "") or "").strip()

    # ТЕХНИК: принудительно фильтруем только по себе
    recipient_effective = me_user if role_low == "technician" else recipient_raw

    # ---------- Даты: работаем по календарным дням ----------
    start_dt_raw = _parse_date_ymd(start_date_s)
    end_dt_raw   = _parse_date_ymd(end_date_s)
    start_day = start_dt_raw.date() if start_dt_raw else None
    end_day   = end_dt_raw.date()   if end_dt_raw   else None

    # ---------- Запрос с подзагрузкой связей ----------
    q = (
        db.session.query(IssuedPartRecord)
        .join(Part, IssuedPartRecord.part_id == Part.id)
        .outerjoin(IssuedBatch, IssuedBatch.id == IssuedPartRecord.batch_id)
        .options(
            selectinload(IssuedPartRecord.part),
            selectinload(IssuedPartRecord.batch),
        )
    )

    # ----- КЛЮЧЕВОЕ: правильная «отчётная дата» строки -----
    is_return = or_(
        IssuedPartRecord.quantity < 0,
        func.upper(func.coalesce(IssuedPartRecord.reference_job, '')).like('RETURN%')
    )
    date_expr = case(
        (is_return, IssuedBatch.issue_date),
        else_=func.coalesce(IssuedPartRecord.issue_date, IssuedBatch.issue_date)
    )

    # ✅ Default mode: если пользователь НЕ задал никаких фильтров — показываем последние 50 инвойсов
    has_any_filter = any([
        start_day, end_day,
        invoice_search,
        recipient_effective,
        reference_job,
        location,
        status
    ])
    default_mode = (not has_any_filter)

    # Фильтр по "кому выдано"
    if recipient_effective:
        if role_low == "technician":
            q = q.filter(
                or_(
                    func.trim(IssuedPartRecord.issued_to) == recipient_effective,
                    func.trim(IssuedBatch.issued_to)     == recipient_effective,
                )
            )
        else:
            like = f"%{recipient_effective}%"
            q = q.filter(
                or_(
                    IssuedPartRecord.issued_to.ilike(like),
                    IssuedBatch.issued_to.ilike(like),
                )
            )

    if reference_job:
        q = q.filter(IssuedPartRecord.reference_job.ilike(f'%{reference_job}%'))

    # ---------- Date filter (LA day boundaries -> UTC naive) ----------
    def _la_day_start_utc(d):
        return (
            datetime(d.year, d.month, d.day, 0, 0, 0, tzinfo=LA_TZ)
            .astimezone(timezone.utc)
            .replace(tzinfo=None)
        )

    def _la_next_day_start_utc(d):
        nd = d + timedelta(days=1)
        return (
            datetime(nd.year, nd.month, nd.day, 0, 0, 0, tzinfo=LA_TZ)
            .astimezone(timezone.utc)
            .replace(tzinfo=None)
        )

    if start_day:
        q = q.filter(date_expr >= _la_day_start_utc(start_day))
    if end_day:
        q = q.filter(date_expr < _la_next_day_start_utc(end_day))

    # ---------- Invoice#/INV# unified search ----------
    invoice_no = None
    if invoice_search:
        try:
            invoice_no = int(invoice_search)
        except ValueError:
            invoice_no = None

        like = f"%{invoice_search}%"
        if invoice_no is not None:
            q = q.filter(
                or_(
                    IssuedPartRecord.invoice_number == invoice_no,
                    func.coalesce(IssuedPartRecord.inv_ref, "").ilike(like),
                )
            )
        else:
            q = q.filter(IssuedPartRecord.inv_ref.ilike(like))

    if location:
        q = q.filter(IssuedPartRecord.location == location)

    # ---------- Фильтр по статусу строки ----------
    if status == "OPEN":
        q = q.filter(
            IssuedPartRecord.quantity > 0,
            func.coalesce(IssuedPartRecord.consumed_qty, 0) == 0
        )
    elif status == "PARTIAL":
        q = q.filter(
            IssuedPartRecord.quantity > 0,
            func.coalesce(IssuedPartRecord.consumed_qty, 0) > 0,
            func.coalesce(IssuedPartRecord.consumed_qty, 0) < IssuedPartRecord.quantity
        )
    elif status == "CONSUMED":
        q = q.filter(
            IssuedPartRecord.quantity > 0,
            func.coalesce(IssuedPartRecord.consumed_qty, 0) >= IssuedPartRecord.quantity
        )

    # ---------- DEFAULT: last 50 invoices (when no filters) ----------
    rows = None
    if default_mode:
        recent_inv = (
            q.with_entities(IssuedPartRecord.invoice_number)
             .filter(IssuedPartRecord.invoice_number.isnot(None))
             .order_by(date_expr.desc(), IssuedPartRecord.id.desc())
             .distinct()
             .limit(DEFAULT_LIMIT)
             .all()
        )
        recent_inv_numbers = [x[0] for x in recent_inv if x and x[0] is not None]

        if not recent_inv_numbers:
            rows = []
        else:
            q2 = (
                db.session.query(IssuedPartRecord)
                .join(Part, IssuedPartRecord.part_id == Part.id)
                .outerjoin(IssuedBatch, IssuedBatch.id == IssuedPartRecord.batch_id)
                .options(
                    selectinload(IssuedPartRecord.part),
                    selectinload(IssuedPartRecord.batch),
                )
                .filter(IssuedPartRecord.invoice_number.in_(recent_inv_numbers))
            )
            rows = q2.order_by(
                func.date(date_expr).desc(),
                IssuedPartRecord.id.desc()
            ).all()

            # ✅ не показываем дефолтные даты в форме
            start_date_s = ""
            end_date_s = ""

    if rows is None:
        rows = q.order_by(
            func.date(date_expr).desc(),
            IssuedPartRecord.id.desc()
        ).all()

    # ---------- Группировка: batch/legacy ----------
    grouped = defaultdict(list)
    for r in rows:
        if getattr(r, 'batch_id', None):
            key = ('BATCH', r.batch_id)
        else:
            is_ret = ((r.quantity or 0) < 0) or ((r.reference_job or '').upper().startswith('RETURN'))
            safe_dt = (
                r.batch.issue_date if (is_ret and getattr(r, "batch", None)) else
                (r.issue_date or (r.batch.issue_date if getattr(r, "batch", None) else None))
            )
            day = (safe_dt.date() if safe_dt else datetime.min.date())
            inv_num = getattr(r, 'invoice_number', None)
            key = ('LEGACY', r.issued_to, r.reference_job, r.issued_by, day, inv_num)
        grouped[key].append(r)

    invoices = []
    grand_total = 0.0

    def _is_return_records(items: list[IssuedPartRecord]) -> bool:
        if not items:
            return False
        if any((it.quantity or 0) < 0 for it in items):
            return True
        ref = (items[0].reference_job or '').strip().upper()
        return ref.startswith('RETURN')

    for gkey, items in grouped.items():
        items_sorted = sorted(items, key=lambda it: it.id)
        total_value = sum((it.quantity or 0) * (it.unit_cost_at_issue or 0.0) for it in items_sorted)
        grand_total += total_value

        if gkey[0] == 'BATCH':
            batch = items_sorted[0].batch
            batch_inv_number = batch.invoice_number or next(
                (it.invoice_number for it in reversed(items_sorted) if (it.invoice_number or 0) > 0),
                None
            )
            inv = {
                'id': f'B{batch.id}',
                'issued_to': batch.issued_to,
                'reference_job': batch.reference_job,
                'issued_by': batch.issued_by,
                'issue_date': batch.issue_date,
                'issue_date_local': utc_to_local(batch.issue_date),
                'invoice_number': batch_inv_number,
                'location': batch.location,
                'items': items_sorted,
                'total_value': total_value,
                'is_return': _is_return_records(items_sorted),
                '_sort_dt': max((it.issue_date for it in items_sorted if it.issue_date), default=batch.issue_date),
                '_sort_id': max((it.id for it in items_sorted), default=0),
            }
        else:
            _, issued_to, ref_job, issued_by, day, inv_num = gkey
            issue_dt = datetime.combine(day, time.min)
            first = items_sorted[0]
            inv_num_from_items = next(
                (it.invoice_number for it in reversed(items_sorted) if (it.invoice_number or 0) > 0),
                None
            )
            resolved_inv_num = inv_num or inv_num_from_items
            if not resolved_inv_num:
                batch_guess = (
                    db.session.query(IssuedBatch)
                    .filter(
                        func.trim(IssuedBatch.issued_to) == (issued_to or ''),
                        func.trim(IssuedBatch.reference_job) == (ref_job or ''),
                        func.trim(IssuedBatch.issued_by) == (issued_by or ''),
                        func.date(IssuedBatch.issue_date) == day
                    )
                    .order_by(IssuedBatch.id.desc())
                    .first()
                )
                if batch_guess and (batch_guess.invoice_number or 0) > 0:
                    resolved_inv_num = batch_guess.invoice_number

            inv = {
                'id': f'K{issued_to}|{issued_by}|{ref_job or ""}|{day.isoformat()}|{resolved_inv_num or ""}',
                'issued_to': issued_to,
                'reference_job': ref_job,
                'issued_by': issued_by,
                'issue_date': issue_dt,
                'invoice_number': resolved_inv_num,
                'issue_date_local': utc_to_local(issue_dt),
                'location': first.location,
                'items': items_sorted,
                'total_value': total_value,
                'is_return': _is_return_records(items_sorted),
                '_sort_dt': max((it.issue_date for it in items_sorted if it.issue_date), default=issue_dt),
                '_sort_id': max((it.id for it in items_sorted), default=0),
            }

        invoices.append(inv)

    # ---------- Сортировка карточек ----------
    invoices.sort(
        key=lambda g: (
            (g.get('_sort_dt') or datetime.min),
            (g.get('invoice_number') or 0),
            (g.get('_sort_id') or 0),
        ),
        reverse=True
    )

    return render_template(
        'reports_grouped.html',
        invoices=invoices,
        total=grand_total,
        start_date=start_date_s,
        end_date=end_date_s,
        recipient=(recipient_effective or ''),
        reference_job=reference_job or '',
        invoice=invoice_s or '',
        location=location or '',
        status=status or '',

        # ✅ For dropdown + stable display in template
        return_destinations=return_destinations,
        dest_map=dest_map,
    )


@inventory_bp.route("/invoice/printdirect")
@login_required
def invoice_printdirect():
    """
    Больше не рендерим промежуточную HTML-страницу и не вызываем window.print().
    Просто отправляем пользователя в обычный PDF viewer (/invoice/pdf),
    передавая полученные query-параметры как есть.
    """
    from flask import request, redirect, url_for

    inv_s         = (request.args.get("invoice_number") or "").strip()
    issued_to_in  = (request.args.get("issued_to") or "").strip()
    reference_job = (request.args.get("reference_job") or "").strip() or None
    issued_by     = (request.args.get("issued_by") or "").strip()
    issue_date_s  = (request.args.get("issue_date") or "").strip()

    return redirect(url_for(
        'inventory.view_invoice_pdf',
        invoice_number=inv_s,
        issued_to=issued_to_in,
        reference_job=reference_job,
        issued_by=issued_by,
        issue_date=issue_date_s
    ), code=302)



@inventory_bp.route("/invoice/pdf")
@login_required
def view_invoice_pdf():
    """
    Печать инвойса.

    Техник/tech:
      - видит ТОЛЬКО свои строки (issued_to == его имя).
        Чужие строки тихо вырезаются.
        Если после фильтрации ничего не осталось — редиректим его назад в grouped.
    Admin/superadmin:
      - полный доступ.

    Legacy-кейс:
      Если у группы ещё НЕТ invoice_number и это старые строки без batch,
      аккуратно создаём batch / резервируем номер, коммитим,
      и печатаем уже с присвоенным номером.

    Параметр ?print=1 пока просто игнорируется на уровне ответа (PDF всё равно inline).
    В будущем можно будет использовать его внутри generate_invoice_pdf,
    если ты сделаешь HTML+window.print().
    """

    from extensions import db
    from models import IssuedPartRecord, IssuedBatch
    from datetime import datetime, time as _time
    from sqlalchemy import func, or_
    from flask import request, make_response, flash, redirect, url_for
    from flask_login import current_user

    # ---- роль текущего пользователя ----
    role_raw = (getattr(current_user, "role", "") or "").strip().lower()
    is_admin_like = role_raw in ("admin", "superadmin")
    is_technician = role_raw in ("technician", "tech")

    # нормализованное имя техника (для сравнения с issued_to в строках)
    my_name_norm = (getattr(current_user, "username", "") or "").strip().lower()

    # ---- вспомогалки --------------------------------------------------------
    def _next_invoice_number():
        mb = (
            db.session.query(func.coalesce(func.max(IssuedBatch.invoice_number), 0))
            .scalar()
            or 0
        )
        ml = (
            db.session.query(func.coalesce(func.max(IssuedPartRecord.invoice_number), 0))
            .scalar()
            or 0
        )
        return max(int(mb), int(ml)) + 1

    def _ensure_invoice_number_for_records(
        records,
        issued_to,
        issued_by,
        reference_job,
        issue_date,
        location,
    ):
        """
        Гарантирует, что у набора records появится invoice_number:
        1) Пытаемся через твой нормальный хелпер _create_batch_for_records,
           чтобы всё было красиво.
        2) Если не вышло — делаем fallback: создаём IssuedBatch вручную,
           присваиваем номер всем строкам.
        Возвращает invoice_number (int) или None.
        """

        # Если в списке уже есть номер хоть у одной строки — просто вернём его.
        if any(getattr(r, "invoice_number", None) for r in records):
            return getattr(records[0], "invoice_number", None)

        # Попробуем через твой штатный хелпер.
        try:
            batch = _create_batch_for_records(
                records=records,
                issued_to=issued_to,
                issued_by=issued_by,
                reference_job=reference_job,
                issue_date=issue_date,
                location=location,
            )
            return getattr(batch, "invoice_number", None)
        except Exception:
            db.session.rollback()

        # Fallback: вручную резервируем новый номер и создаём IssuedBatch
        for _ in range(5):
            inv_no_try = _next_invoice_number()
            try:
                with db.session.begin_nested():
                    batch = IssuedBatch(
                        invoice_number=inv_no_try,
                        issued_to=issued_to,
                        issued_by=issued_by or "system",
                        reference_job=reference_job,
                        issue_date=issue_date,
                        location=(location or None),
                    )
                    db.session.add(batch)
                    db.session.flush()

                    for r in records:
                        r.batch_id = batch.id
                        r.invoice_number = inv_no_try

                    db.session.flush()

                return inv_no_try
            except Exception:
                db.session.rollback()
                continue

        raise RuntimeError("Failed to reserve invoice number")

    def _parse_dt_flex(s: str | None):
        if not s:
            return None
        for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%m/%d/%Y"):
            try:
                return datetime.strptime(s, fmt)
            except Exception:
                pass
        return None

    # ---- входные параметры (query string) -----------------------------------
    inv_s         = (request.args.get("invoice_number") or "").strip()
    issued_to_in  = (request.args.get("issued_to") or "").strip()
    reference_job = (request.args.get("reference_job") or "").strip() or None
    issued_by     = (request.args.get("issued_by") or "").strip()
    issue_date_s  = (request.args.get("issue_date") or "").strip()
    # хотим автопечать? пока не используем, но future-proof оставим прочитанным
    want_auto     = (request.args.get("print") or "").strip() == "1"

    inv_no = int(inv_s) if inv_s.isdigit() else None

    # ---- загружаем строки recs и собираем hdr -------------------------------
    recs = []
    hdr  = None

    if inv_no is not None:
        # 1) Найдём batch_ids с таким invoice_number
        batch_ids = [
            bid
            for (bid,) in db.session.query(IssuedBatch.id)
            .filter(IssuedBatch.invoice_number == inv_no)
            .all()
        ]

        # 2) Берём IssuedPartRecord с этим номером или batch_id
        q = IssuedPartRecord.query
        if batch_ids:
            recs = (
                q.filter(
                    or_(
                        IssuedPartRecord.invoice_number == inv_no,
                        IssuedPartRecord.batch_id.in_(batch_ids),
                    )
                )
                .order_by(IssuedPartRecord.id.asc())
                .all()
            )
        else:
            recs = (
                q.filter_by(invoice_number=inv_no)
                .order_by(IssuedPartRecord.id.asc())
                .all()
            )

        if not recs:
            flash(f"Invoice #{inv_no} not found.", "warning")
            return redirect(url_for("inventory.reports_grouped"))

        # 3) Построим hdr на основе батча или первой строки
        if batch_ids:
            b = db.session.get(IssuedBatch, batch_ids[0])
            hdr = {
                "issued_to": b.issued_to,
                "reference_job": b.reference_job,
                "issued_by": b.issued_by,
                "issue_date": b.issue_date,
                "invoice_number": inv_no,
                "location": getattr(b, "location", None),
            }
        else:
            first = recs[0]
            hdr = {
                "issued_to": first.issued_to,
                "reference_job": first.reference_job,
                "issued_by": first.issued_by,
                "issue_date": first.issue_date,
                "invoice_number": inv_no,
                "location": first.location,
            }

        # 4) Добавим legacy-строки того же дня без номера (старые "висячие" строки)
        day = hdr["issue_date"].date() if hdr.get("issue_date") else None
        if day:
            extra = (
                IssuedPartRecord.query.filter(
                    func.trim(IssuedPartRecord.issued_to)
                    == (hdr["issued_to"] or ""),
                    func.trim(IssuedPartRecord.issued_by)
                    == (hdr["issued_by"] or ""),
                    func.trim(IssuedPartRecord.reference_job)
                    == (hdr["reference_job"] or ""),
                    func.date(IssuedPartRecord.issue_date) == day,
                    or_(
                        IssuedPartRecord.invoice_number.is_(None),
                        IssuedPartRecord.invoice_number == 0,
                    ),
                    IssuedPartRecord.batch_id.is_(None),
                )
                .order_by(IssuedPartRecord.id.asc())
                .all()
            )

            if extra:
                have_ids = {r.id for r in recs}
                for r in extra:
                    if r.id not in have_ids:
                        recs.append(r)

        # 5) итог сортируем по id
        recs.sort(key=lambda r: r.id)

    else:
        # legacy режим: инвойс без номера, ищем по issued_to / issued_by / reference_job / дате
        if issued_to_in and issued_by and issue_date_s:
            dt = _parse_dt_flex(issue_date_s) or datetime.utcnow()

            start = datetime.combine(dt.date(), _time.min)
            end   = datetime.combine(dt.date(), _time.max)

            recs = (
                IssuedPartRecord.query.filter(
                    IssuedPartRecord.issued_to == issued_to_in,
                    IssuedPartRecord.issued_by == issued_by,
                    IssuedPartRecord.reference_job == reference_job,
                    IssuedPartRecord.issue_date.between(start, end),
                )
                .order_by(IssuedPartRecord.id.asc())
                .all()
            )

    # Ничего не нашли → назад
    if not recs:
        flash("Invoice lines not found.", "warning")
        return redirect(url_for("inventory.reports_grouped"))

    # ---- Ограничение доступа техника ---------------------------------------
    # Техник видит ТОЛЬКО строки, где issued_to == его имя.
    if is_technician and (not is_admin_like):
        allowed_name = my_name_norm

        safe_recs = []
        for r in recs:
            rec_issued_to_norm = (getattr(r, "issued_to", "") or "").strip().lower()
            if rec_issued_to_norm == allowed_name:
                safe_recs.append(r)

        recs = safe_recs

        # После фильтрации пусто? Возвращаем техника в его отчёт.
        if not recs:
            flash("Access denied for this invoice.", "warning")
            return redirect(url_for("inventory.reports_grouped"))

    # ---- Присвоение invoice_number для legacy инвойсов ---------------------
    if inv_no is None and all(getattr(r, "invoice_number", None) is None for r in recs):
        base = recs[0]
        try:
            new_no = _ensure_invoice_number_for_records(
                records=recs,
                issued_to=getattr(base, "issued_to", issued_to_in),
                issued_by=getattr(base, "issued_by", issued_by),
                reference_job=getattr(base, "reference_job", reference_job),
                issue_date=getattr(base, "issue_date", _parse_dt_flex(issue_date_s) or datetime.utcnow()),
                location=getattr(base, "location", None),
            )
            db.session.commit()
            inv_no = new_no
        except Exception:
            db.session.rollback()
            inv_no = None

    # ---- Генерация PDF (как раньше, без авто_print аргумента) ---------------
    pdf_bytes = generate_invoice_pdf(
        recs,
        invoice_number=inv_no
    )

    # Ответ
    resp = make_response(pdf_bytes)

    fname_base = inv_no or getattr(recs[0], "id", "NO_NUM")
    if isinstance(fname_base, int):
        fname = f"INVOICE_{int(fname_base):06d}.pdf"
    else:
        fname = "INVOICE.pdf"

    # PDF inline (браузерный viewer)
    resp.headers["Content-Type"] = "application/pdf"
    resp.headers["Content-Disposition"] = f'inline; filename="{fname}"'

    return resp


@inventory_bp.route('/reports/update_record/<int:record_id>', methods=['POST'])
@login_required
def update_report_record(record_id):
    if current_user.role not in ['superadmin', 'admin']:
        flash("Access denied", "danger")
        return redirect(url_for('inventory.reports'))

    r = IssuedPartRecord.query.get_or_404(record_id)

    # обновляем разрешённые поля
    if 'issued_to' in request.form:
        v = (request.form.get('issued_to') or '').strip()
        if v: r.issued_to = v

    if 'reference_job' in request.form:
        v = (request.form.get('reference_job') or '').strip() or None
        r.reference_job = v

    if 'unit_cost' in request.form:
        raw = (request.form.get('unit_cost') or '').strip()
        try:
            r.unit_cost_at_issue = float(raw)
        except Exception:
            pass

    if current_user.role == 'superadmin' and 'issue_date' in request.form:
        s = (request.form.get('issue_date') or '').strip()
        if s:
            try:
                # сохраняем ТУ ЖЕ дату с полночью, время не трогаем, чтобы не прыгал порядок
                from datetime import datetime
                d = datetime.strptime(s, '%Y-%m-%d').date()
                r.issue_date = datetime.combine(d, r.issue_date.time())
            except Exception:
                pass

    db.session.commit()

    # КЛЮЧЕВОЕ: после изменения Reference Job/Issued To
    # нужно перерисовать страницы, чтобы кнопки карточки получили НОВЫЕ record_ids[]
    flash("Line saved.", "success")
    return redirect(url_for('inventory.reports'))

# Отмена отдельной позиции
@inventory_bp.route('/reports/cancel/<int:record_id>', methods=['POST'])
@login_required
def cancel_issued_record(record_id):
    if current_user.role not in ['superadmin', 'admin']:
        flash("Access denied", "danger")
        return redirect(url_for('inventory.reports'))

    record = IssuedPartRecord.query.get_or_404(record_id)
    part = Part.query.get(record.part_id)
    if part:
        part.quantity += record.quantity

    db.session.delete(record)
    db.session.commit()

    flash(f"Issued record #{record.id} canceled and stock restored.", "success")
    return redirect(url_for('inventory.reports'))

@inventory_bp.route('/reports/update_invoice', methods=['POST'])
@login_required
def update_invoice():
    """
    Save / Return / Delete from the grouped report.

    Additions in this version:
    - `apply_scope`: 'all' (default) or 'selected'.
    - Invoice number is assigned only for scope='all' and only if ALL rows in the group have no number yet.
    - When scope='all' and there's no number, records are resolved by the group's legacy keys.
    - **Hard cap for Return Selected**: you cannot return more than has been issued for that line.
      The backend calculates already-returned quantity and trims the requested amount accordingly.
    """
    from extensions import db
    from models import IssuedPartRecord, IssuedBatch, Part
    from datetime import datetime, time as _time
    from sqlalchemy import func

    # ---------- helpers ----------
    def _is_return_row(r):
        """A row is a 'return' when its quantity is negative."""
        return (getattr(r, 'quantity', 0) or 0) < 0

    def _next_invoice_number():
        """Safe next invoice number from max(IssuedBatch, IssuedPartRecord)."""
        mb = db.session.query(func.coalesce(func.max(IssuedBatch.invoice_number), 0)).scalar() or 0
        ml = db.session.query(func.coalesce(func.max(IssuedPartRecord.invoice_number), 0)).scalar() or 0
        return max(int(mb), int(ml)) + 1

    def _ensure_invoice_number_for_records(
        records,
        issued_to,
        issued_by,
        reference_job,
        issue_date,
        location,
        force_new=False
    ):
        """
        Safe version: when force_new=True — always creates a new batch with a unique invoice_number.
        Used for repeated returns / extra issues.
        """
        from extensions import db
        from models import IssuedBatch

        if not records:
            return None

        # If not forced and at least one record already has a number — do nothing
        if not force_new and any(getattr(r, "invoice_number", None) for r in records):
            return None

        # Primary path — your helper
        try:
            batch = _create_batch_for_records(
                records=records,
                issued_to=issued_to,
                issued_by=issued_by,
                reference_job=reference_job,
                issue_date=issue_date,
                location=location,
            )
            return batch
        except Exception:
            db.session.rollback()

        # Fallback — reserve manually
        for _ in range(5):
            inv_no = _next_invoice_number()
            try:
                with db.session.begin_nested():
                    batch = IssuedBatch(
                        invoice_number=inv_no,
                        issued_to=issued_to,
                        issued_by=issued_by or "system",
                        reference_job=reference_job,
                        issue_date=issue_date,
                        location=(location or None),
                    )
                    db.session.add(batch)
                    db.session.flush()

                    for r in records:
                        r.batch_id = batch.id
                        r.invoice_number = inv_no
                    db.session.flush()
                return batch
            except Exception:
                db.session.rollback()
        raise RuntimeError("Failed to reserve invoice number")

    def _already_returned_qty_for(r) -> int:
        """
        How many units have already been returned against the same *issued* row r.
        """
        if r is None or (r.quantity or 0) <= 0 or not getattr(r, "part_id", None):
            return 0

        part_id = r.part_id
        issued_to = (r.issued_to or '')
        ref_orig = (r.reference_job or '').strip()

        base = (db.session.query(func.coalesce(func.sum(IssuedPartRecord.quantity), 0))
                .filter(IssuedPartRecord.part_id == part_id,
                        IssuedPartRecord.issued_to == issued_to,
                        IssuedPartRecord.quantity < 0))
        if ref_orig:
            base = base.filter(IssuedPartRecord.reference_job == f"RETURN {ref_orig}")
        else:
            base = base.filter(IssuedPartRecord.reference_job.ilike("RETURN%"))

        total_neg = base.scalar() or 0
        return abs(int(total_neg))

    def _available_to_return_for(r) -> int:
        """Remaining quantity that can still be returned for this issued row."""
        issued_qty = int(r.quantity or 0)
        already_ret = _already_returned_qty_for(r)
        return max(0, issued_qty - already_ret)

    # ---------- access ----------
    if current_user.role not in ['superadmin', 'admin']:
        flash("Access denied", "danger")
        return redirect(url_for('inventory.reports_grouped'))

    # ---------- flags / scope ----------
    do_return        = (request.form.get('do_return') or '') == '1'
    do_delete_ret    = (request.form.get('delete_return') or '') == '1'
    delete_invoice_s = (request.form.get('delete_invoice') or '').strip()
    delete_invoice_no = int(delete_invoice_s) if delete_invoice_s.isdigit() else None
    apply_scope = (request.form.get('apply_scope') or 'all').lower()  # 'all' | 'selected'

    # ---------- header values ----------
    location  = (request.form.get('location') or '').strip() or None
    issued_by = (request.form.get('issued_by') or getattr(current_user, 'username', '') or '').strip()

    s_date = (request.form.get('issue_date') or request.form.get('issue_date_old') or '').strip()
    issue_date = _parse_dt_flex(s_date) if s_date else datetime.utcnow()

    invoice_s = (request.form.get('invoice_number') or '').strip()
    form_invoice_no = int(invoice_s) if invoice_s.isdigit() else None

    # Group keys for scope='all' when invoice number is missing
    g_issued_to = (request.form.get('group_issued_to') or '').strip()
    g_ref_job   = (request.form.get('group_reference_job') or '').strip() or None
    g_issued_by = (request.form.get('group_issued_by') or '').strip()
    g_issue_s   = (request.form.get('group_issue_date') or '').strip()

    # ---------- collect selected ids ----------
    raw_ids = request.form.getlist('record_ids[]') or request.form.getlist('record_ids')
    sel_ids = [int(x) for x in raw_ids if str(x).strip().isdigit()]

    # ---------- load records ----------
    recs = []
    if apply_scope == 'selected' and sel_ids:
        recs = IssuedPartRecord.query.filter(IssuedPartRecord.id.in_(sel_ids)).all()

    if not recs and form_invoice_no is not None:
        recs = IssuedPartRecord.query.filter_by(invoice_number=form_invoice_no).all()

    if not recs and apply_scope == 'all':
        # When invoice has no number yet — resolve by legacy keys within the day
        if g_issued_to and g_issued_by and g_issue_s:
            dt = _parse_dt_flex(g_issue_s) or issue_date
            start = datetime.combine(dt.date(), _time.min)
            end   = datetime.combine(dt.date(), _time.max)
            recs = (IssuedPartRecord.query
                    .filter(IssuedPartRecord.issued_to == g_issued_to,
                            IssuedPartRecord.issued_by == g_issued_by,
                            IssuedPartRecord.reference_job == g_ref_job,
                            IssuedPartRecord.issue_date.between(start, end))
                    .order_by(IssuedPartRecord.id.asc()).all())

    if not recs:
        flash("Invoice lines not found.", "warning")
        return redirect(url_for('inventory.reports_grouped'))

    # ---------- DELETE INVOICE ----------
    if delete_invoice_no is not None:
        if current_user.role != 'superadmin':
            flash("Only superadmin may delete invoice.", "danger")
            return redirect(url_for('inventory.reports_grouped'))
        touched = set()
        try:
            for r in list(recs):
                if r.part:
                    # Revert stock impact (return issued qty to stock; for negative rows this subtracts)
                    r.part.quantity = int(r.part.quantity or 0) + int(r.quantity or 0)
                if r.batch_id:
                    touched.add(r.batch_id)
                db.session.delete(r)
            # Remove empty batches
            for bid in touched:
                still = db.session.query(IssuedPartRecord.id).filter_by(batch_id=bid).first()
                if not still:
                    b = db.session.get(IssuedBatch, bid)
                    if b:
                        db.session.delete(b)
            db.session.commit()
            flash(f"Invoice #{delete_invoice_no:06d} deleted.", "success")
        except Exception as e:
            db.session.rollback()
            flash(f"Failed to delete invoice: {e}", "danger")
        return redirect(url_for('inventory.reports_grouped'))

    # ---------- DELETE RETURN ----------
    if do_delete_ret:
        if current_user.role != 'superadmin':
            flash("Only superadmin may delete return invoices.", "danger")
            return redirect(url_for('inventory.reports_grouped'))
        touched = set()
        try:
            for r in recs:
                if not _is_return_row(r):
                    continue
                if r.part and (r.quantity or 0) < 0:
                    # Deleting a return (negative) reduces stock back
                    r.part.quantity = int(r.part.quantity or 0) - abs(int(r.quantity or 0))
                if r.batch_id:
                    touched.add(r.batch_id)
                db.session.delete(r)
            for bid in list(touched):
                still = db.session.query(IssuedPartRecord.id).filter_by(batch_id=bid).first()
                if not still:
                    b = db.session.get(IssuedBatch, bid)
                    if b:
                        db.session.delete(b)
            db.session.commit()
            flash("Return lines deleted.", "success")
        except Exception as e:
            db.session.rollback()
            flash(f"Failed to delete return invoice: {e}", "danger")
        return redirect(url_for('inventory.reports_grouped'))

    # ---------- RETURN SELECTED (with hard cap) ----------
    if do_return:
        try:
            created = 0
            trimmed_any = False

            for r in recs:
                # Only process against a positive "issued" row
                if (r.quantity or 0) <= 0:
                    continue

                available = _available_to_return_for(r)
                if available <= 0:
                    continue

                raw = (request.form.get(f"qty_{r.id}") or "1").strip()
                try:
                    qty_req = int(raw)
                except Exception:
                    qty_req = 1

                if qty_req < 0:
                    qty_req = 0

                if qty_req > available:
                    qty_req = available
                    trimmed_any = True

                if qty_req <= 0:
                    continue

                # Create a negative "return" row
                ret = IssuedPartRecord(
                    part_id=r.part_id,
                    issued_to=r.issued_to,
                    issued_by=issued_by or r.issued_by,
                    quantity=-qty_req,
                    unit_cost_at_issue=(
                        r.unit_cost_at_issue
                        if r.unit_cost_at_issue is not None
                        else (r.part.unit_cost if r.part else 0.0)
                    ),
                    reference_job=(
                        f"RETURN {r.reference_job}"
                        if (r.reference_job and not str(r.reference_job).upper().startswith("RETURN"))
                        else (r.reference_job or "RETURN")
                    ),
                    issue_date=issue_date,
                    location=(location or r.location),
                    invoice_number=None,
                    batch_id=None
                )
                db.session.add(ret)

                if r.part:
                    r.part.quantity = int(r.part.quantity or 0) + qty_req

                created += 1

            # assign a fresh invoice number for the newly created return rows
            if created:
                new_returns = (
                    db.session.query(IssuedPartRecord)
                    .filter(IssuedPartRecord.invoice_number.is_(None))
                    .filter(IssuedPartRecord.quantity < 0)
                    .order_by(IssuedPartRecord.id.desc())
                    .limit(created)
                    .all()
                )
                if new_returns:
                    return_ref = getattr(new_returns[0], "reference_job", None) or "RETURN"
                    _ensure_invoice_number_for_records(
                        records=new_returns,
                        issued_to=getattr(new_returns[0], "issued_to", None),
                        issued_by=issued_by,
                        reference_job=return_ref,
                        issue_date=datetime.utcnow(),
                        location=getattr(new_returns[0], "location", None) or location,
                        force_new=True
                    )

            db.session.commit()

            if created:
                note = " (some requested quantities were trimmed)" if trimmed_any else ""
                flash(f"Created {created} return line(s){note}.", "success")
            else:
                flash("No lines were eligible for return.", "warning")

        except Exception as e:
            db.session.rollback()
            flash(f"Failed to create return: {e}", "danger")

        return redirect(url_for('inventory.reports_grouped'))

    # ---------- SAVE (edit + assign invoice number for scope='all') ----------

    # ---------- SAVE (edit + assign invoice number for scope='all') ----------
    try:
        role = (getattr(current_user, "role", "") or "").strip().lower()
        can_edit_refjob = role in ("superadmin", "admin")
        is_super = role == "superadmin"

        # 1) SUPERADMIN: full edit (qty/ucost/issued_to/refjob)
        if is_super:
            for r in recs:
                # Qty edit
                new_qty_s = request.form.get(f"edit_qty_{r.id}")
                if new_qty_s not in (None, ""):
                    try:
                        new_qty = int(new_qty_s)
                        old_qty = int(r.quantity or 0)
                        if r.part:
                            # Put back the difference: old - new
                            r.part.quantity = int(r.part.quantity or 0) + (old_qty - new_qty)
                        r.quantity = new_qty
                    except Exception:
                        pass

                # Unit cost edit
                new_ucost_s = request.form.get(f"edit_ucost_{r.id}")
                if new_ucost_s not in (None, ""):
                    try:
                        r.unit_cost_at_issue = float(new_ucost_s)
                    except Exception:
                        pass

                # Issued To edit
                new_it = request.form.get(f"edit_issued_to_{r.id}")
                if new_it is not None:
                    r.issued_to = new_it.strip()

                # Reference Job edit
                new_rj = request.form.get(f"edit_refjob_{r.id}")
                if new_rj is not None:
                    r.reference_job = (new_rj.strip() or None)

        # 2) ADMIN: only Reference Job edit
        elif can_edit_refjob:
            changed_ref = None
            for r in recs:
                new_rj = request.form.get(f"edit_refjob_{r.id}")
                if new_rj is None:
                    continue
                new_val = (new_rj.strip() or None)
                if new_val != r.reference_job:
                    r.reference_job = new_val
                    changed_ref = new_val

            # If we changed ref job — also update batch header(s) so the card "Ref:" updates
            if changed_ref is not None:
                batch_ids = {getattr(r, "batch_id", None) for r in recs if getattr(r, "batch_id", None)}
                if batch_ids:
                    for bid in batch_ids:
                        b = db.session.get(IssuedBatch, bid)
                        if b:
                            b.reference_job = changed_ref

        # 3) Common header fields (keep as-is)
        for r in recs:
            r.issued_by = issued_by or r.issued_by
            r.issue_date = issue_date or r.issue_date
            if location:
                r.location = location

        # 4) Assign invoice number only when saving the whole card and only if none of the rows has a number
        if apply_scope == 'all' and all(getattr(r, "invoice_number", None) is None for r in recs):
            base = recs[0]
            _ensure_invoice_number_for_records(
                records=recs,
                issued_to=getattr(base, "issued_to", ""),
                issued_by=getattr(base, "issued_by", ""),
                reference_job=getattr(base, "reference_job", None),
                issue_date=getattr(base, "issue_date", issue_date),
                location=getattr(base, "location", location),
            )

        db.session.commit()
        flash("Invoice saved.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Failed to save invoice: {e}", "danger")

    return redirect(url_for('inventory.reports_grouped'))

@inventory_bp.route('/reports/cancel_invoice', methods=['POST'])
@login_required
def cancel_invoice():
    if current_user.role not in ['superadmin', 'admin']:
        flash("Access denied", "danger")
        return redirect(url_for('inventory.reports'))

    ids = request.form.getlist('record_ids[]') or request.form.getlist('record_ids')
    records = []
    if ids:
        try:
            ids = [int(x) for x in ids]
        except ValueError:
            ids = []
        if ids:
            records = IssuedPartRecord.query.filter(IssuedPartRecord.id.in_(ids)).all()

    if not records:
        issued_to  = (request.form.get('issued_to') or request.args.get('issued_to') or '').strip()
        reference  = (request.form.get('reference_job') or request.args.get('reference_job') or '').strip() or None
        issued_by  = (request.form.get('issued_by') or request.args.get('issued_by') or '').strip()
        s          = (request.form.get('issue_date') or request.args.get('issue_date') or '').strip()
        issue_date = _parse_dt_flex(s)

        if not (issued_to and issued_by and issue_date):
            flash("Invoice identifiers are missing.", "danger")
            return redirect(url_for('inventory.reports'))

        start = datetime.combine(issue_date.date(), time.min)
        end   = datetime.combine(issue_date.date(), time.max)
        records = IssuedPartRecord.query.filter(
            IssuedPartRecord.issued_to == issued_to,
            IssuedPartRecord.reference_job == reference,
            IssuedPartRecord.issued_by == issued_by,
            IssuedPartRecord.issue_date.between(start, end)
        ).all()

    if not records:
        flash("Invoice not found.", "warning")
        return redirect(url_for('inventory.reports'))

    for r in records:
        part = Part.query.get(r.part_id)
        if part:
            part.quantity = int(part.quantity or 0) + int(r.quantity or 0)
        db.session.delete(r)
    db.session.commit()

    flash("Invoice canceled and stock restored.", "success")
    return redirect(url_for('inventory.reports'))

# ----------------- Download Invoice -----------------
@inventory_bp.route("/import", methods=["GET", "POST"])
@login_required
def import_parts():
    """
    Загрузка PDF → парсинг → превью.
    Безопасно хранит путь к файлу в meta['attachment_path'].
    """
    import os
    from datetime import datetime
    from flask import current_app, request, redirect, url_for, render_template, flash
    from werkzeug.utils import secure_filename

    if request.method == "GET":
        return render_template("receiving_import_upload.html")

    # ---------- POST ----------
    file = request.files.get("file")
    if not file or not file.filename.strip():
        flash("Select a PDF file.", "warning")
        return redirect(url_for("inventory.import_parts"))

    src_name = secure_filename(file.filename)               # исходное имя файла
    ext = os.path.splitext(src_name)[1].lower()
    if ext != ".pdf":
        flash("Only PDF is supported.", "warning")
        return redirect(url_for("inventory.import_parts"))

    # Куда сохраняем
    upload_dir = os.path.join(current_app.instance_path, "uploads")
    os.makedirs(upload_dir, exist_ok=True)
    pdf_path = os.path.join(upload_dir, f"{datetime.utcnow():%Y%m%d_%H%M%S}_{src_name}")

    # Сохраняем файл
    file.save(pdf_path)

    # Парсим PDF → DataFrame → нормализация строк
    try_ocr = bool(request.form.get("try_ocr"))
    df = dataframe_from_pdf(pdf_path, try_ocr=try_ocr)      # твой хелпер
    rows = _norm_cols(df) if not df.empty else []
    if not rows:
        flash("Nothing parsed from PDF. Try OCR option.", "warning")
        return redirect(url_for("inventory.import_parts"))

    # --- хелпер разбора даты + валидация ---
    def _parse_invoice_date(raw: str | None):
        raw = (raw or "").strip()
        if not raw:
            return None
        for fmt in ("%Y-%m-%d", "%m/%d/%Y"):
            try:
                return datetime.strptime(raw, fmt).date()
            except Exception:
                continue
        return None

    # Метаданные заголовка (можно править на превью)
    supplier = (request.form.get("supplier") or rows[0].get("supplier") or "").strip()
    invoice  = (request.form.get("invoice_number") or "").strip()
    date_s   = (request.form.get("invoice_date") or "").strip()
    notes    = (request.form.get("notes") or f"Imported from {src_name}").strip()

    # ---- НЕ ДАЁМ использовать дату из будущего ----
    today = datetime.utcnow().date()
    inv_date_val = _parse_invoice_date(date_s)
    if inv_date_val and inv_date_val > today:
        # удаляем сохранённый pdf, чтобы не копить мусор
        try:
            os.remove(pdf_path)
        except OSError:
            pass

        flash("Invoice date cannot be in the future. Please correct it and try again.", "danger")
        return redirect(url_for("inventory.import_parts"))

    # Рендер превью, ПЕРЕДАЁМ путь скрытым полем
    return render_template(
        "receiving_import_preview.html",
        rows=rows,
        meta=dict(
            supplier=supplier,
            invoice=invoice,
            date=date_s,          # покажем как ввёл, можно исправить на превью
            notes=notes,
            currency="USD",
            attachment_path=pdf_path,        # ← тут путь для hidden
        ),
        src_file=src_name
    )


@inventory_bp.route('/invoice/<int:record_id>')
@login_required
def invoice(record_id):
    record = IssuedPartRecord.query.get(record_id)
    if not record:
        return "Record not found", 404

    pdf_data = generate_invoice_pdf(record)
    return send_file(BytesIO(pdf_data), as_attachment=True, download_name='invoice.pdf')


ALLOWED_EXTS = {".xlsx", ".xls", ".csv", ".pdf"}

def allowed_file(filename: str) -> bool:
    return os.path.splitext(filename)[1].lower() in ALLOWED_EXTS


@inventory_bp.route('/reports/download')
@login_required
def download_report_pdf():
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from datetime import datetime
    from io import BytesIO

    start = request.args.get('start_date')
    end = request.args.get('end_date')
    recipient = request.args.get('recipient')
    reference_job = request.args.get('reference_job')

    query = IssuedPartRecord.query.join(Part)
    if start:
        query = query.filter(IssuedPartRecord.issue_date >= start)
    if end:
        end_dt = datetime.strptime(end, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
        query = query.filter(IssuedPartRecord.issue_date <= end_dt)
    if recipient:
        query = query.filter(IssuedPartRecord.issued_to.ilike(f"%{recipient}%"))
    if reference_job:
        query = query.filter(IssuedPartRecord.reference_job.ilike(f"%{reference_job}%"))

    records = query.order_by(IssuedPartRecord.issue_date.desc()).all()

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=20)
    elements = []
    styles = getSampleStyleSheet()

    title = Paragraph("Issued Parts Report", styles["Heading1"])
    elements.append(title)
    elements.append(Spacer(1, 12))

    filter_text = ""
    if start and end:
        filter_text = f"Filters: from {datetime.strptime(start, '%Y-%m-%d').strftime('%m/%d/%Y')} to {datetime.strptime(end, '%Y-%m-%d').strftime('%m/%d/%Y')}"
    elif start:
        filter_text = f"Filters: from {datetime.strptime(start, '%Y-%m-%d').strftime('%m/%d/%Y')}"
    elif end:
        filter_text = f"Filters: up to {datetime.strptime(end, '%Y-%m-%d').strftime('%m/%d/%Y')}"

    if filter_text:
        elements.append(Paragraph(filter_text, styles["Normal"]))
        elements.append(Spacer(1, 12))

    data = [["Date", "Part #", "Name", "Qty", "Unit Cost", "Total", "Issued To", "Job Ref."]]

    total_sum = 0
    for r in records:
        total = r.quantity * r.part.unit_cost
        total_sum += total
        row = [
            r.issue_date.strftime('%m/%d/%Y'),
            Paragraph(r.part.part_number, styles['Normal']),
            Paragraph(r.part.name, styles['Normal']),
            str(r.quantity),
            f"${r.unit_cost_at_issue:.2f}",
            f"${total:.2f}",
            Paragraph(r.issued_to, styles['Normal']),
            Paragraph(r.reference_job or '—', styles['Normal'])
        ]
        data.append(row)

    # Итоговая строка
    data.append([
        "", "", "", "", "TOTAL:",
        f"${total_sum:.2f}", "", ""
    ])

    col_widths = [65, 130, 150, 35, 65, 65, 90, 110]

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#d3d3d3")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (3, 1), (5, -2), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),

        # Здесь рисуем сетку по всей таблице
        ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),

        # Вертикальная линия слева от колонки "Unit Cost"
        ('LINEBEFORE', (5, 0), (5, -2), 0.25, colors.grey),

        # Линия над итоговой строкой
        ('LINEABOVE', (4, -1), (5, -1), 0.25, colors.grey),
        # Итоговая строка - светлый фон, отступы, выравнивание
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor("#f0f0f0")),  # светло-серый фон
        ('TOPPADDING', (0, -1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, -1), (-1, -1), 8),

        # Итоговые шрифты
        ('FONTNAME', (4, -1), (4, -1), 'Helvetica-Bold'),
        ('FONTNAME', (5, -1), (5, -1), 'Helvetica-Bold'),

        # Выравнивание итога по правому краю
        ('ALIGN', (5, -1), (5, -1), 'RIGHT'),


    ]))

    elements.append(table)
    doc.build(elements)

    buffer.seek(0)
    return send_file(buffer,
                     as_attachment=True,
                     download_name=f"report_{start or 'all'}_{end or 'all'}.pdf",
                     mimetype='application/pdf')


@inventory_bp.route("/users", methods=["GET", "POST"])
@login_required
def users():
    # ---- Создание нового пользователя (разрешено только superadmin)
    if request.method == "POST":
        if current_user.role != ROLE_SUPERADMIN:
            flash("Access denied", "danger")
            return redirect(url_for("inventory.users"))

        username = (request.form.get("username") or "").strip()
        password = request.form.get("password") or ""
        role     = (request.form.get("role") or "").strip().lower()

        if not username or not password:
            flash("Username and password are required.", "warning")
            return redirect(url_for("inventory.users"))

        # case-insensitive проверка уникальности
        exists = (
            db.session.query(User.id)
            .filter(func.lower(User.username) == username.lower())
            .first()
        )
        if exists:
            flash("User already exists.", "warning")
            return redirect(url_for("inventory.users"))

        # модель сама нормализует роль (валидатор), дефолт — technician
        u = User(username=username, role=role or ROLE_TECHNICIAN)
        u.password = password
        db.session.add(u)
        db.session.commit()
        flash("User created.", "success")
        return redirect(url_for("inventory.users"))

    # ---- Список пользователей
    users = User.query.order_by(User.username.asc()).all()

    # Опции ролей для селекта
    role_options = [
        ("technician", "Technician"),
        ("user",       "User"),
        ("viewer",     "Viewer"),
        ("admin",      "Admin"),
        ("superadmin", "Superadmin"),
    ]

    return render_template("users.html", users=users, role_options=role_options)
@inventory_bp.route('/users/add', methods=['GET', 'POST'])
@login_required
def add_user():
    if current_user.role not in ['superadmin', 'admin']:
        flash("Access denied", "danger")
        return redirect(url_for('inventory.dashboard'))

    if request.method == 'POST':
        username = request.form['username'].strip()
        role = request.form['role']
        password = request.form['password']

        # Админ может создавать только user
        if current_user.role == 'admin' and role != 'user':
            flash("Admins can only create users with role 'user'.", "danger")
            return redirect(url_for('inventory.add_user'))

        if User.query.filter_by(username=username).first():
            flash("Username already exists", "danger")
            return redirect(url_for('inventory.add_user'))

        new_user = User(
            username=username,
            role=role,
            password_hash=generate_password_hash(password)
        )
        db.session.add(new_user)
        db.session.commit()
        flash("User added successfully", "success")
        return redirect(url_for('inventory.users'))

    allowed_roles = ['user'] if current_user.role == 'admin' else ['user', 'admin', 'superadmin']
    return render_template('add_user.html', allowed_roles=allowed_roles)


@inventory_bp.route('/users/edit/<int:user_id>', methods=['GET', 'POST'])
@login_required
def edit_user(user_id):
    from sqlalchemy import func
    user = User.query.get_or_404(user_id)

    me_role = (current_user.role or '').lower()
    target_role = (user.role or '').lower()

    # Права: superadmin — всех; admin — только user/viewer/technician и себя
    if me_role == 'admin':
        if not (user.id == current_user.id or target_role in {'user', 'viewer', 'technician'}):
            flash("Admins can only edit users with role 'user/viewer/technician' or themselves.", "danger")
            return redirect(url_for('inventory.dashboard'))
    elif me_role != 'superadmin':
        flash("Access denied", "danger")
        return redirect(url_for('inventory.dashboard'))

    # Список ролей для селекта:
    # - супер-админ видит все роли
    # - админ может назначать только user/viewer/technician; себе — оставлять admin
    role_options = []
    if me_role == 'superadmin':
        role_options = [
            ('technician', 'Technician'),
            ('user', 'User'),
            ('viewer', 'Viewer'),
            ('admin', 'Admin'),
            ('superadmin', 'Superadmin'),
        ]
    else:  # admin
        role_options = [
            ('technician', 'Technician'),
            ('user', 'User'),
            ('viewer', 'Viewer'),
        ]
        # если админ редактирует себя — оставляем возможность иметь admin (но не менять на супер-админа)
        if user.id == current_user.id:
            role_options.append(('admin', 'Admin'))

    if request.method == 'POST':
        new_username = (request.form.get('username') or '').strip()
        new_role_raw = (request.form.get('role') or '').strip().lower()

        # username можно менять обоим (как было), без усложнения логики
        if new_username:
            # защита от дублей по нижнему регистру
            exists = (
                db.session.query(User.id)
                .filter(func.lower(User.username) == new_username.lower(), User.id != user.id)
                .first()
            )
            if exists:
                flash("Username already exists.", "danger")
                return redirect(url_for('inventory.edit_user', user_id=user.id))
            user.username = new_username

        # роль — только в пределах разрешённых опций
        allowed_values = {val for val, _ in role_options}
        if new_role_raw in allowed_values:
            user.role = new_role_raw  # у вас в модели стоит нормализация/дефолт
        else:
            # ничего не меняем, чтобы не сломать существующее значение
            pass

        db.session.commit()
        flash("User updated successfully", "success")
        return redirect(url_for('inventory.users'))

    return render_template('edit_user.html', user=user, role_options=role_options)

@inventory_bp.route('/users/change_password/<int:user_id>', methods=['GET', 'POST'])
@login_required
def change_password(user_id):
    from flask import request, flash, redirect, url_for, render_template, jsonify
    from flask_login import current_user
    from werkzeug.security import generate_password_hash
    from extensions import db
    from models import User

    user = User.query.get_or_404(user_id)

    role = (current_user.role or '').lower()
    is_self = current_user.id == user_id

    # superadmin → любой
    # admin → только user и себя
    # остальные → только себя
    if role == 'admin':
        if (user.role or '').lower() != 'user' and not is_self:
            # not allowed
            if request.headers.get('X-Requested-With') == 'fetch':
                return jsonify({"ok": False, "error": "Access denied"}), 403
            flash("Admins can only change passwords for users with role 'user' or themselves.", "danger")
            return redirect(url_for('inventory.dashboard'))

    elif role != 'superadmin' and not is_self:
        if request.headers.get('X-Requested-With') == 'fetch':
            return jsonify({"ok": False, "error": "Access denied"}), 403
        flash("Access denied", "danger")
        return redirect(url_for('inventory.dashboard'))

    if request.method == 'POST':
        # security check for self-change (non-superadmin must enter current password)
        if is_self and role != 'superadmin':
            current_password = (request.form.get('current_password') or '').strip()
            if not current_password or not user.check_password(current_password):
                if request.headers.get('X-Requested-With') == 'fetch':
                    return jsonify({"ok": False, "error": "Current password is incorrect."}), 400
                flash("Current password is incorrect.", "danger")
                return redirect(url_for('inventory.change_password', user_id=user_id))

        new_password = (request.form.get('password') or '').strip()
        confirm_password = (request.form.get('confirm_password') or '').strip()

        if not new_password:
            if request.headers.get('X-Requested-With') == 'fetch':
                return jsonify({"ok": False, "error": "New password cannot be empty."}), 400
            flash("New password cannot be empty.", "danger")
            return redirect(url_for('inventory.change_password', user_id=user_id))

        if new_password != confirm_password:
            if request.headers.get('X-Requested-With') == 'fetch':
                return jsonify({"ok": False, "error": "Passwords do not match."}), 400
            flash("Passwords do not match.", "danger")
            return redirect(url_for('inventory.change_password', user_id=user_id))

        user.password_hash = generate_password_hash(new_password)
        db.session.commit()

        # если это fetch-запрос из модалки -> вернём json (без redirect)
        if request.headers.get('X-Requested-With') == 'fetch':
            return jsonify({"ok": True})

        # обычный сценарий (юзер сам зашёл на страницу смены пароля)
        flash("Password changed successfully", "success")
        if role == 'superadmin':
            return redirect(url_for('inventory.users'))
        else:
            return redirect(url_for('inventory.dashboard'))

    # GET (обычный режим, не через модалку)
    need_current = is_self and role != 'superadmin'
    return render_template('change_password.html', user=user, need_current=need_current)

@inventory_bp.route('/users/delete/<int:user_id>', methods=['POST'])
@login_required
def delete_user(user_id):
    user = User.query.get_or_404(user_id)

    # Только superadmin может удалять
    if current_user.role != 'superadmin':
        flash("Access denied", "danger")
        return redirect(url_for('inventory.dashboard'))

    if user.id == current_user.id:
        flash("You cannot delete yourself!", "danger")
        return redirect(url_for('inventory.users'))

    db.session.delete(user)
    db.session.commit()
    flash("User deleted successfully", "success")
    return redirect(url_for('inventory.users'))


@inventory_bp.route('/clear_issued_records')
@login_required
def clear_issued_records():
    if current_user.role != 'superadmin':
        flash('Only superadmin can clear records.', 'danger')
        return redirect(url_for('inventory.dashboard'))

    from models import IssuedPartRecord
    from extensions import db

    IssuedPartRecord.query.delete()
    db.session.commit()
    flash('All issued records cleared.', 'success')
    return redirect(url_for('inventory.dashboard'))

@inventory_bp.route('/update_part/<int:part_id>', methods=['POST'])
@login_required
def update_part_field(part_id):
    if current_user.role != 'superadmin':
        flash("Access denied", "danger")
        return redirect(url_for('inventory.dashboard'))

    part = Part.query.get_or_404(part_id)

    quantity = request.form.get('quantity')
    unit_cost = request.form.get('unit_cost')

    if quantity is not None:
        try:
            part.quantity = int(quantity)
        except ValueError:
            flash("Invalid quantity value", "danger")
            return redirect(url_for('inventory.dashboard'))

    if unit_cost is not None:
        try:
            part.unit_cost = float(unit_cost)
        except ValueError:
            flash("Invalid unit cost value", "danger")
            return redirect(url_for('inventory.dashboard'))

    db.session.commit()
    flash("Part updated successfully.", "success")
    return redirect(url_for('inventory.dashboard'))

@inventory_bp.route('/delete/<int:part_id>', methods=['POST'])
@login_required
def delete_part(part_id):
    if current_user.role != 'superadmin':
        flash('Only Superadmin can delete parts.', 'danger')
        return redirect(url_for('inventory.dashboard'))

    part = Part.query.get_or_404(part_id)

    # 1) Проверяем, использовалась ли деталь в выдачах
    issued_count = IssuedPartRecord.query.filter_by(part_id=part.id).count()
    if issued_count:
        flash(
            f'Cannot delete part {part.part_number}: '
            f'it is referenced in {issued_count} issued record(s). '
            f'Set quantity to 0 or mark it as inactive instead.',
            'danger'
        )
        # Вернёмся на дашборд с поиском по этой детали
        return redirect(url_for('inventory.dashboard', search=part.part_number))

    # 2) Пытаемся удалить (на случай других связей – подстрахуемся try/except)
    try:
        db.session.delete(part)
        db.session.commit()
        flash(f'Part {part.part_number} deleted.', 'success')
    except IntegrityError as e:
        db.session.rollback()
        flash(
            'Cannot delete this part because it is still referenced in other '
            'receiving / stock / return records. '
            'Please clear those references first.',
            'danger'
        )

    return redirect(url_for('inventory.dashboard'))

@inventory_bp.route('/edit/<int:part_id>', methods=['GET', 'POST'])
@login_required
def edit_part(part_id):
    if current_user.role != 'superadmin':
        flash("Access denied", "danger")
        return redirect(url_for('inventory.dashboard'))

    part = Part.query.get_or_404(part_id)

    if request.method == 'POST':
        part.name = request.form['name'].strip().upper()
        part.part_number = request.form['part_number'].strip().upper()
        try:
            part.quantity = int(request.form['quantity'])
            part.unit_cost = float(request.form['unit_cost'])
        except ValueError:
            flash("Invalid quantity or unit cost", "danger")
            return redirect(url_for('inventory.edit_part', part_id=part_id))

        part.location = request.form['location'].strip().upper()

        db.session.commit()
        flash("Part updated successfully.", "success")
        return redirect(url_for('inventory.dashboard'))

    return render_template('edit_part.html', part=part)


@inventory_bp.route('/reports/delete/<int:record_id>', methods=['POST'])
@login_required
def delete_report_record(record_id):
    if current_user.role != 'superadmin':
        flash('Access denied', 'danger')
        return redirect(url_for('inventory.reports'))

    record = IssuedPartRecord.query.get_or_404(record_id)
    db.session.delete(record)
    db.session.commit()
    flash('Issued record deleted successfully.', 'success')
    return redirect(url_for('inventory.reports'))

@inventory_bp.route('/clear_parts')
@login_required
def clear_parts():
    if current_user.role != 'superadmin':
        flash('Only superadmin can clear parts.', 'danger')
        return redirect(url_for('inventory.dashboard'))

    from models import Part
    from extensions import db

    Part.query.delete()
    db.session.commit()
    flash('All parts cleared.', 'success')
    return redirect(url_for('inventory.dashboard'))

@inventory_bp.route('/compare_cart')
def compare_cart():
    try:
        flash("🔍 Collecting Marcone cart...", "info")
        items = get_marcone_items()

        flash("📦 Comparing with inventory...", "info")
        result = check_cart_items(items)

        filepath = os.path.join(UPLOAD_DIR, "marcone_inventory_report.docx")
        export_to_docx(result, filename=filepath)

        flash("✅ Marcone report generated! Click below to download.", "success")
        return redirect(url_for("inventory.dashboard"))

    except Exception as e:
        flash(f"❌ Error (Marcone): {str(e)}", "danger")
        return redirect(url_for("inventory.dashboard"))

# @inventory_bp.route('/download_marcone_report')
# def download_marcone_report():
#     filepath = os.path.join(UPLOAD_DIR, "marcone_inventory_report.docx")
#     if os.path.exists(filepath):
#         return send_file(filepath, as_attachment=True)
#     flash("❌ Marcone report not found!", "danger")
#     return redirect(url_for("inventory.dashboard"))


@inventory_bp.route('/compare_reliable')
def compare_reliable():
    try:
        flash("🔍 Collecting Reliable cart...", "info")
        items = get_reliable_items()

        flash("📦 Comparing with inventory...", "info")
        result = check_cart_items(items)

        filepath = os.path.join(UPLOAD_DIR, "reliable_inventory_report.docx")
        export_to_docx(result, filename=filepath)

        flash("✅ Reliable report generated! Click below to download.", "success")
        return redirect(url_for("inventory.dashboard"))

    except Exception as e:
        flash(f"❌ Error (Reliable): {str(e)}", "danger")
        return redirect(url_for("inventory.dashboard"))

# @inventory_bp.route('/download_reliable_report')
# def download_reliable_report():
#     filepath = os.path.join(UPLOAD_DIR, "reliable_inventory_report.docx")
#     if os.path.exists(filepath):
#         return send_file(filepath, as_attachment=True)
#     flash("❌ Reliable report not found!", "danger")
#     return redirect(url_for("inventory.dashboard"))

@inventory_bp.route('/download_marcone_report')
@login_required
def download_marcone_report():
    from compare_cart.run_compare import get_marcone_items, check_cart_items, export_to_docx

    filepath = os.path.join(UPLOAD_DIR, "marcone_inventory_report.docx")

    # 1. Получаем данные из корзины Marcone
    items = get_marcone_items()

    # 2. Сравниваем с базой данных
    result = check_cart_items(items)

    # 3. Генерируем новый отчет
    export_to_docx(result, filename=filepath)

    # 4. Отдаем файл пользователю
    if os.path.exists(filepath):
        @after_this_request
        def remove_file(response):
            try:
                os.remove(filepath)
            except Exception as e:
                print(f"Failed to delete file: {e}")
            return response

        return send_file(filepath, as_attachment=True)

    flash("❌ Marcone report generation failed!", "danger")
    return redirect(url_for("inventory.dashboard"))


@inventory_bp.route('/download_reliable_report')
@login_required
def download_reliable_report():
    from compare_cart.run_compare_reliable import get_reliable_items, check_cart_items, export_to_docx

    filepath = os.path.join(UPLOAD_DIR, "reliable_inventory_report.docx")

    items = get_reliable_items()
    result = check_cart_items(items)
    export_to_docx(result, filename=filepath)

    if os.path.exists(filepath):
        @after_this_request
        def remove_file(response):
            try:
                os.remove(filepath)
            except Exception as e:
                print(f"Failed to delete file: {e}")
            return response

        return send_file(filepath, as_attachment=True)

    flash("❌ Reliable report generation failed!", "danger")
    return redirect(url_for("inventory.dashboard"))

def parse_preview_rows_relaxed(form):
    """
    Собирает строки превью из request.form:
    [{"part_number":..., "part_name":..., "quantity":..., "unit_cost_base":..., "unit_cost":..., "location":...}]
    """
    buckets = defaultdict(dict)

    # 1) units[0][rows][i][field]
    for key, val in form.items():
        m = _rows_re.match(key)
        if m:
            unit_idx, row_idx, field = m.groups()
            buckets[int(row_idx)][field] = (val or "").strip()

    # 2) rows[i][field] fallback
    for key, val in form.items():
        m = _rows_flat_re.match(key)
        if m:
            row_idx, field = m.groups()
            buckets[int(row_idx)][field] = (val or "").strip()

    out = []

    for i in sorted(buckets.keys()):
        r = buckets[i]

        # qty
        try:
            qty = int((r.get("quantity") or "0").replace(",", ""))
        except Exception:
            qty = 0

        # base cost (editable)
        try:
            base = float((r.get("unit_cost_base") or "0").replace("$", "").replace(",", ""))
        except Exception:
            base = 0.0

        # adjusted cost (readonly)
        try:
            adj = float((r.get("unit_cost") or "").replace("$", "").replace(",", ""))
        except Exception:
            adj = 0.0

        # если adj пустой, а base есть → берём base
        if adj <= 0 and base > 0:
            adj = base

        pn   = (r.get("part_number") or "").strip()
        name = (r.get("part_name") or "").strip()

        # выкидываем только реально пустые строки
        if not pn and not name and qty <= 0 and base <= 0 and adj <= 0:
            continue

        loc = (r.get("location") or "").strip().upper() or "MAIN"

        out.append({
            "part_number":     pn,
            "part_name":       name,
            "quantity":        qty,
            "unit_cost_base":  base,
            "unit_cost":       adj,
            "location":        loc,
            "supplier":        (r.get("supplier") or "").strip(),
        })

    return out

from datetime import datetime, date
import os, logging
from sqlalchemy.exc import IntegrityError

@inventory_bp.route("/import-parts", methods=["GET", "POST"], endpoint="import_parts_upload")
def import_parts_upload():
    import os
    import re
    import logging
    from datetime import datetime, date
    from sqlalchemy.exc import IntegrityError

    # ===== helpers ============================================================
    def _supplier_to_default_location(supplier_hint: str | None) -> str:
        s = (supplier_hint or "").lower()
        for key, loc in SUPPLIER_LOC_DEFAULTS.items():
            if key in s:
                return loc
        return "MAIN"

    def _pick_field(model, candidates):
        for f in candidates:
            if hasattr(model, f):
                return f
        return None

    def _coerce_norm_df(df):
        import pandas as pd
        if df is None:
            return None
        if isinstance(df, pd.DataFrame):
            return df
        try:
            return pd.DataFrame(df)
        except Exception:
            return None

    def _infer_invoice_number(norm_df, source_file: str, supplier_hint: str | None) -> str:
        # unchanged logic, но с правильным импортом pdfminer
        if norm_df is not None:
            for col in ("invoice_no", "invoice", "order_no", "ref", "reference"):
                if col in norm_df.columns:
                    vals = [str(x).strip() for x in norm_df[col].dropna().astype(str).tolist()]
                    vals = [v for v in vals if v and v.lower() not in ("nan", "none", "null")]
                    if vals:
                        from collections import Counter
                        c = Counter(vals)
                        best = c.most_common(1)[0][0]
                        m = re.search(r"\b\d{6,}\b", best.replace(" ", ""))
                        if m:
                            return m.group(0)
                        return best

        base = os.path.basename(source_file or "")
        m2 = re.findall(r"\d{6,}", base)
        if m2:
            return max(m2, key=len)

        if str(source_file).lower().endswith(".pdf"):
            text = ""
            try:
                import fitz
                with fitz.open(source_file) as d:
                    for i in range(min(3, d.page_count)):
                        text += d[i].get_text() + "\n"
            except Exception:
                try:
                    # <<< ВАЖНО: правильный модуль pdfminer >>>
                    from pdfminer.high_level import extract_text
                    text = extract_text(source_file) or ""
                except Exception:
                    text = ""
            if text:
                mm = re.search(r"Invoice[^0-9]*([0-9]{6,})", text, flags=re.I)
                if mm:
                    return mm.group(1)
        return ""

    def _ensure_norm_columns(df, default_loc: str, saved_path: str):
        import pandas as pd
        import os

        # если совсем нечего – вернём пустой df с нужными колонками
        if df is None:
            cols = [
                "part_number", "part_name", "qty", "quantity",
                "unit_cost", "unit_cost_base",
                "location", "row_key", "source_file", "supplier",
                "order_no", "invoice_no", "date",
            ]
            return pd.DataFrame(columns=cols)

        df = df.copy()

        # гарантируем наличие всех колонок
        need_cols = [
            "part_number", "part_name", "qty", "quantity",
            "unit_cost", "unit_cost_base",
            "location", "row_key", "source_file", "supplier",
            "order_no", "invoice_no", "date",
        ]
        for c in need_cols:
            if c not in df.columns:
                df[c] = None

        # ----- qty / quantity -----
        qty = pd.to_numeric(df["qty"], errors="coerce")
        quantity = pd.to_numeric(df["quantity"], errors="coerce")

        df.loc[qty.isna() & quantity.notna(), "qty"] = quantity
        df.loc[quantity.isna() & qty.notna(), "quantity"] = qty

        df["qty"] = pd.to_numeric(df["qty"], errors="coerce").fillna(0).astype(int)
        df.loc[df["qty"] < 0, "qty"] = 0
        df["quantity"] = df["qty"].astype(int)

        # ----- unit_cost / unit_cost_base -----
        df["unit_cost_base"] = pd.to_numeric(df["unit_cost_base"], errors="coerce")
        df["unit_cost"]      = pd.to_numeric(df["unit_cost"], errors="coerce")

        # где base пустая, подставляем unit_cost
        mask_no_base = df["unit_cost_base"].isna()
        df.loc[mask_no_base, "unit_cost_base"] = df.loc[mask_no_base, "unit_cost"]

        # ----- source_file -----
        sf_empty = df["source_file"].isna() | (df["source_file"].astype(str).str.strip() == "")
        df.loc[sf_empty, "source_file"] = saved_path

        # ----- location -----
        loc_empty = df["location"].isna() | (df["location"].astype(str).str.strip() == "")
        df.loc[loc_empty, "location"] = (default_loc or "MAIN")
        df["location"] = (
            df["location"]
            .astype(str)
            .str.strip()
            .str.upper()
        )

        # ----- нормализация строковых полей (БЕЗ row_key!) -----
        for col in ("part_number", "part_name", "supplier",
                    "order_no", "invoice_no", "source_file"):
            df[col] = (
                df[col]
                .astype(str)
                .replace({"None": ""})
                .fillna("")
                .str.strip()
            )

        # ===== ВСЕГДА пересчитываем row_key по новой схеме =====
        # чтобы не использовать старые ключи вида "137292700/REL"
        df = df.reset_index(drop=False).rename(columns={"index": "__row_i"})
        file_id = os.path.basename(str(saved_path or ""))

        def _mk_key(row):
            pn = str(row.get("part_number", "")).strip().upper()
            loc = str(row.get("location", "")).strip().upper()
            try:
                qty_local = int(pd.to_numeric(row.get("qty", 0), errors="coerce") or 0)
            except Exception:
                qty_local = 0
            cost_base = row.get("unit_cost_base", None)
            if pd.isna(cost_base):
                cost_base = "NA"
            else:
                try:
                    cost_base = float(cost_base)
                except Exception:
                    cost_base = "NA"
            i = int(row.get("__row_i", 0) or 0)
            return f"{file_id}|{i}|{pn}|{loc}|{qty_local}|{cost_base}"

        df["row_key"] = df.apply(_mk_key, axis=1)

        if "__row_i" in df.columns:
            df = df.drop(columns=["__row_i"])

        # ----- выбрасываем полностью пустые строки -----
        drop_mask = (
            (df["part_number"].astype(str).str.strip() == "") &
            (df["part_name"].astype(str).str.strip() == "") &
            (df["qty"] == 0) &
            (df["unit_cost_base"].fillna(0) == 0)
        )
        if drop_mask.any():
            df = df[~drop_mask].copy()

        return df

    def _detect_supplier_from_content(saved_path: str, df_hint):
        """
        Пытается определить поставщика:
        1) по имени файла
        2) по колонкам / значениям df_hint
        3) по тексту PDF:
           - сначала fitz (если установлен)
           - потом pdfminer (если установлен)
           - если оба недоступны, используем pdfplumber как fallback
        """
        try:
            base = (saved_path or "").lower()
            if "reliable" in base:
                return "Reliable Parts"
            if "marcone" in base or "marcon" in base:
                return "Marcone"

            blob = ""
            # --- 1) df_hint: заголовки + первые строки
            if df_hint is not None:
                try:
                    blob = " ".join(df_hint.columns.astype(str))
                except Exception:
                    pass
                try:
                    blob += " " + " ".join(
                        [str(x) for x in df_hint.head(30).astype(str).values.ravel()]
                    )
                except Exception:
                    pass

            blob_l = blob.lower()
            if "reliable" in blob_l:
                return "Reliable Parts"
            if "marcone" in blob_l or "marcon" in blob_l:
                return "Marcone"

            # --- 2) PDF текст
            text = ""
            # 2a. fitz (PyMuPDF), если есть
            try:
                import fitz  # type: ignore
                with fitz.open(saved_path) as d:
                    for i in range(min(3, d.page_count)):
                        text += d[i].get_text() + "\n"
            except Exception:
                # 2b. pdfminer, если есть
                try:
                    from pdfminer.high_level import extract_text  # type: ignore
                    text = extract_text(saved_path) or ""
                except Exception:
                    text = ""

            # 2c. fallback через pdfplumber, если текст всё ещё пустой
            if not text:
                try:
                    import pdfplumber  # type: ignore
                    with pdfplumber.open(saved_path) as pdf:
                        for page in pdf.pages[:3]:
                            t = page.extract_text() or ""
                            if t:
                                text += t + "\n"
                except Exception:
                    pass

            tl = (text or "").lower()
            if "reliable parts" in tl or "reliable parts inc" in tl or "reliableparts" in tl:
                return "Reliable Parts"
            if "marcone" in tl or "marcone supply" in tl:
                return "Marcone"
        except Exception:
            pass
        return None

    def _merge_locations_stable(old_loc: str | None, new_loc: str | None) -> str:
        out = []
        for raw in (old_loc, new_loc):
            if not raw:
                continue
            for token in str(raw).upper().split("/"):
                t = token.strip()
                if t and t not in out:
                    out.append(t)
        return "/".join(out)

    def _parse_invoice_date_from_form(field_name: str = "invoice_date") -> date:
        raw = (request.form.get(field_name) or "").strip()
        if not raw:
            return date.today()
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(raw, fmt).date()
            except ValueError:
                continue
        return date.today()

    # -------- extra: применить ОДИН РАЗ --------------------------------------
    def _apply_extra_to_df_once(df, extra_expenses_float, eps=1e-6):
        import pandas as pd

        df = df.copy()
        df["quantity"] = pd.to_numeric(df.get("quantity"), errors="coerce").fillna(0).astype(int)
        df["unit_cost_base"] = pd.to_numeric(df.get("unit_cost_base"), errors="coerce").fillna(0.0)
        if "unit_cost" not in df.columns:
            df["unit_cost"] = df["unit_cost_base"]
        df["unit_cost"] = pd.to_numeric(df.get("unit_cost"), errors="coerce")

        subtotal_base = float((df["quantity"] * df["unit_cost_base"]).sum() or 0.0)
        current_total = float((df["quantity"] * df["unit_cost"]).sum() or 0.0)
        extra = float(extra_expenses_float or 0.0)
        target_total = subtotal_base + extra

        if abs(current_total - target_total) <= float(eps or 1e-6):
            return df, subtotal_base, target_total

        return _distribute_extra_and_adjust_costs(df, extra)

    def _distribute_extra_and_adjust_costs(df, extra_expenses_float):
        import pandas as pd

        df = df.copy()
        df["quantity"] = pd.to_numeric(df["quantity"], errors="coerce").fillna(0).astype(int)
        df["unit_cost_base"] = pd.to_numeric(df["unit_cost_base"], errors="coerce")

        df["line_base_total"] = df["quantity"] * df["unit_cost_base"]
        subtotal_base = float(df["line_base_total"].sum() or 0.0)

        if subtotal_base > 0 and extra_expenses_float:
            def _adj_cost(row):
                qty = row["quantity"]
                base_cost = row["unit_cost_base"]
                if qty <= 0:
                    return base_cost
                share = (row["line_base_total"] / subtotal_base) if subtotal_base > 0 else 0.0
                extra_for_row = extra_expenses_float * share
                per_item_fee = extra_for_row / qty
                return base_cost + per_item_fee

            df["unit_cost"] = df.apply(_adj_cost, axis=1)
        else:
            df["unit_cost"] = df["unit_cost_base"]

        df["unit_cost"] = df["unit_cost"].astype(float)
        grand_total = subtotal_base + float(extra_expenses_float or 0.0)
        return df, subtotal_base, grand_total

    # ==========================================================================
    enabled = int(current_app.config.get("WCCR_IMPORT_ENABLED", 0))
    dry     = int(current_app.config.get("WCCR_IMPORT_DRY_RUN", 1))

    if request.method == "POST":
        keys = list(request.form.keys())
        flash(
            f"DEBUG: method=POST, keys={keys[:12]}{' ...' if len(keys) > 12 else ''}, "
            f"enabled={enabled}, dry={dry}",
            "info"
        )

    # ===== A) POST из превью (Save / Apply) ===================================
    if (
        request.method == "POST" and
        (
            ("save" in request.form) or
            ("apply" in request.form) or
            any(k.startswith("units[") or k.startswith("rows[") for k in request.form.keys())
        )
    ):
        saved_path = (request.form.get("saved_path") or "").strip()
        extra_expenses_raw = (request.form.get("extra_expenses") or "").strip()
        try:
            extra_expenses_val = float(extra_expenses_raw)
        except Exception:
            extra_expenses_val = 0.0

        invoice_date_raw = (request.form.get("invoice_date") or "").strip()
        invoice_date_val = _parse_invoice_date_from_form("invoice_date")

        # НОВОЕ: читаем supplier из visible-поля или скрытого supplier_hint
        supplier_from_form = (request.form.get("supplier") or
                              request.form.get("supplier_hint") or "").strip()

        if not saved_path:
            flash("Saved path is empty. Upload the file again.", "warning")

        rows = parse_preview_rows_relaxed(request.form)
        if not rows:
            flash("Нет данных в форме (пустая таблица).", "warning")
            return render_template(
                "import_preview.html",
                rows=[],
                saved_path=saved_path,
                supplier_hint=supplier_from_form,
                extra_expenses=extra_expenses_val,
                subtotal_base=0.0,
                grand_total=extra_expenses_val,
                invoice_date=invoice_date_raw,
            )

        norm = rows_to_norm_df(rows, saved_path)
        norm = _coerce_norm_df(norm)

        # если юзер не ввёл supplier — пробуем угадать; если всё равно пусто, ругаемся
        supplier_hint = (
            supplier_from_form or
            _detect_supplier_from_content(saved_path, norm) or ""
        )
        if not supplier_hint:
            flash("Supplier is required. Please fill Supplier field.", "danger")

        default_loc = _supplier_to_default_location(supplier_hint)

        if norm is None or norm.empty:
            flash("Нет данных для применения импорта (пустой набор строк).", "warning")
            return render_template(
                "import_preview.html",
                rows=[],
                saved_path=saved_path,
                supplier_hint=supplier_hint,
                extra_expenses=extra_expenses_val,
                subtotal_base=0.0,
                grand_total=extra_expenses_val,
                invoice_date=invoice_date_raw,
            )

        norm = _ensure_norm_columns(norm, default_loc, saved_path)

        import pandas as pd
        norm["quantity"] = pd.to_numeric(norm["quantity"], errors="coerce").fillna(0).astype(int)
        norm["unit_cost"] = pd.to_numeric(norm["unit_cost"], errors="coerce").fillna(0.0)
        norm["unit_cost_base"] = pd.to_numeric(norm["unit_cost_base"], errors="coerce").fillna(0.0)

        subtotal_base = float((norm["quantity"] * norm["unit_cost_base"]).sum() or 0.0)
        grand_total = float((norm["quantity"] * norm["unit_cost"]).sum() or 0.0)

        # extra не распределяем здесь – пока 0
        extra_expenses_val = 0.0

        flash(
            f"Supplier hint: {supplier_hint or 'None'}, default location: {default_loc}",
            "info"
        )

        # <<< НОВОЕ: пробуем взять Invoice # из формы, а если пусто — авто-парсим >>>
        invoice_guess = (request.form.get("invoice_number") or "").strip()
        if not invoice_guess:
            invoice_guess = _infer_invoice_number(norm, saved_path, supplier_hint) or ""


        # --- ВАЛИДАЦИЯ: дата не может быть в будущем ---
        today = date.today()
        if invoice_date_val and invoice_date_val > today:
            flash("Invoice date cannot be in the future. Please correct Invoice Date.", "danger")
            # Возвращаемся на превью, ничего не создаём в базе
            return render_template(
                "import_preview.html",
                rows=norm.to_dict(orient="records"),
                saved_path=saved_path,
                supplier_hint=supplier_hint,
                extra_expenses=extra_expenses_val,
                subtotal_base=subtotal_base,
                grand_total=grand_total,
                default_loc=default_loc or "MAIN",
                invoice_date=invoice_date_raw,   # показываем то, что он ввёл (даже если будущее)
            )


        # ----- SAVE (остаться на превью) -----
        if "save" in request.form:
            return render_template(
                "import_preview.html",
                rows=norm.to_dict(orient="records"),
                saved_path=saved_path,
                supplier_hint=supplier_hint,
                extra_expenses=extra_expenses_val,
                subtotal_base=subtotal_base,
                grand_total=grand_total,
                default_loc=default_loc or "MAIN",
                invoice_date=invoice_date_raw,
                invoice_number=invoice_guess,
            )


        # ----- APPLY (создать batch) -----
        if "apply" in request.form:
            if dry or not enabled:
                flash("Импорт в режиме предпросмотра (DRY) или отключён конфигом.", "info")
                return render_template(
                    "import_preview.html",
                    rows=norm.to_dict(orient="records"),
                    saved_path=saved_path,
                    supplier_hint=supplier_hint,
                    extra_expenses=extra_expenses_val,
                    subtotal_base=subtotal_base,
                    grand_total=grand_total,
                    default_loc=default_loc or "MAIN",
                    invoice_date=invoice_date_raw,
                    invoice_number=invoice_guess,
                )


            session = db.session

            try:
                from models import ReceivingBatch as BatchModel
            except Exception:
                try:
                    from models import Receiving as BatchModel
                except Exception:
                    BatchModel = None

            try:
                from models import ReceivingItem as ItemModel
            except Exception:
                try:
                    from models import ReceivingLine as ItemModel
                except Exception:
                    ItemModel = None

            batch = None
            batch_id_field = None
            if BatchModel is not None:
                B_SUP   = _pick_field(BatchModel, ["supplier_name","supplier","vendor","provider"])
                B_INV   = _pick_field(BatchModel, ["invoice_number","invoice_no","invoice","number"])
                B_DATE  = _pick_field(BatchModel, ["date","invoice_date","received_date"])
                B_CURR  = _pick_field(BatchModel, ["currency"])
                B_NOTES = _pick_field(BatchModel, ["notes","comment","description"])
                B_STAT  = _pick_field(BatchModel, ["status","state"])
                B_C_AT  = _pick_field(BatchModel, ["created_at","created"])
                B_C_BY  = _pick_field(BatchModel, ["created_by"])
                B_P_AT  = _pick_field(BatchModel, ["posted_at"])
                B_P_BY  = _pick_field(BatchModel, ["posted_by"])
                B_ATTP  = _pick_field(BatchModel, ["attachment_path","attachment"])
                B_EXTRA = _pick_field(BatchModel, ["extra_expenses","shipping_fee","freight","expenses"])

                batch_id_field = _pick_field(BatchModel, ["id","batch_id"])

                batch_kwargs = {}
                if B_SUP:   batch_kwargs[B_SUP]   = (supplier_hint or "Unknown")
                if B_INV:   batch_kwargs[B_INV]   = invoice_guess
                if B_DATE:  batch_kwargs[B_DATE]  = invoice_date_val
                if B_CURR:  batch_kwargs[B_CURR]  = "USD"
                note_extra = f" (extra applied: {extra_expenses_val:.2f})" if extra_expenses_val else ""
                if B_NOTES: batch_kwargs[B_NOTES] = f"Imported from {os.path.basename(saved_path)}{note_extra}"
                if B_STAT:  batch_kwargs[B_STAT]  = "new"
                if B_C_AT:  batch_kwargs[B_C_AT]  = datetime.utcnow()
                if B_C_BY:  batch_kwargs[B_C_BY]  = 0
                if B_P_AT:  batch_kwargs[B_P_AT]  = None
                if B_P_BY:  batch_kwargs[B_P_BY]  = None
                if B_ATTP:  batch_kwargs[B_ATTP]  = ""
                if B_EXTRA: batch_kwargs[B_EXTRA] = 0.0

                try:
                    batch = BatchModel(**batch_kwargs)
                    session.add(batch)
                    session.flush()
                except IntegrityError as e:
                    session.rollback()
                    batch = None
                    logging.exception("DB error while creating receiving batch: %s", e)
                    flash("DB error while creating receiving batch (см. лог). Продолжаю без батча.", "danger")
                except Exception as e:
                    session.rollback()
                    batch = None
                    logging.exception("Failed to create ReceivingBatch: %s", e)
                    flash("Не удалось создать ReceivingBatch. Продолжаю без батча.", "warning")

            def duplicate_exists(rk: str) -> bool:
                return has_key(rk)

            def _merge_locations_stable_local(old_loc: str | None, new_loc: str | None) -> str:
                return _merge_locations_stable(old_loc, new_loc)

            def make_movement(m: dict) -> None:
                PartModel = Part

                PN_FIELDS   = ["part_number","number","sku","code","partnum","pn"]
                NAME_FIELDS = ["name","part_name","descr","description","title"]
                QTY_FIELDS  = ["quantity","qty","on_hand","stock","count"]
                LOC_FIELDS  = ["location","bin","shelf","place","loc"]
                COST_FIELDS = ["unit_cost","cost","price","unitprice","last_cost"]
                SUP_FIELDS  = ["supplier","vendor","provider"]

                pn_field   = _pick_field(PartModel, PN_FIELDS)
                name_field = _pick_field(PartModel, NAME_FIELDS)
                qty_field  = _pick_field(PartModel, QTY_FIELDS)
                loc_field  = _pick_field(PartModel, LOC_FIELDS)
                cost_field = _pick_field(PartModel, COST_FIELDS)
                sup_field  = _pick_field(PartModel, SUP_FIELDS)

                if pn_field is None or qty_field is None:
                    raise RuntimeError("Не найдено поле PART # или QTY в модели Part.")

                incoming_pn   = m["part_number"]
                incoming_qty  = int(m.get("qty") or m.get("quantity") or 0)
                incoming_cost = m.get("unit_cost")
                incoming_name = m.get("part_name") or ""
                incoming_loc  = (m.get("location") or "").strip().upper()
                incoming_sup  = supplier_hint or (m.get("supplier") or "")

                part = PartModel.query.filter(
                    getattr(PartModel, pn_field) == incoming_pn
                ).first()

                if not part:
                    kwargs = {
                        pn_field: incoming_pn,
                        qty_field: 0,
                    }
                    if name_field and incoming_name:
                        kwargs[name_field] = incoming_name
                    if loc_field:
                        kwargs[loc_field] = incoming_loc
                    if cost_field and (incoming_cost is not None):
                        kwargs[cost_field] = float(incoming_cost)
                    if sup_field and incoming_sup:
                        kwargs[sup_field] = incoming_sup

                    part = PartModel(**kwargs)
                    session.add(part)
                    try:
                        session.flush()
                    except IntegrityError:
                        session.rollback()
                        part = PartModel.query.filter(
                            getattr(PartModel, pn_field) == incoming_pn
                        ).first()
                        if not part:
                            raise

                on_hand_before = int(getattr(part, qty_field) or 0)
                old_loc_before = getattr(part, loc_field) if loc_field else None

                setattr(part, qty_field, on_hand_before + incoming_qty)

                if cost_field and (incoming_cost is not None):
                    setattr(part, cost_field, float(incoming_cost))

                if name_field and incoming_name and not getattr(part, name_field):
                    setattr(part, name_field, incoming_name)

                if loc_field:
                    incoming_loc_up = incoming_loc
                    if incoming_loc_up:
                        if on_hand_before > 0:
                            merged = _merge_locations_stable_local(old_loc_before, incoming_loc_up)
                            setattr(part, loc_field, merged)
                        else:
                            setattr(part, loc_field, incoming_loc_up)

                if ItemModel is not None and batch is not None and batch_id_field is not None:
                    I_BATCH = _pick_field(ItemModel, ["goods_receipt_id","batch_id","receiving_id"])
                    I_PART  = _pick_field(ItemModel, ["part_id","item_part_id"])
                    I_PN    = _pick_field(ItemModel, ["part_number","part","sku","code"])
                    I_PNAME = _pick_field(ItemModel, ["part_name","name","descr","description","title"])
                    I_QTY   = _pick_field(ItemModel, ["qty","quantity"])
                    I_COST  = _pick_field(ItemModel, ["unit_cost","cost","price","unitprice"])
                    I_LOC   = _pick_field(ItemModel, ["location","bin","shelf","place","loc"])

                    try:
                        item_kwargs = {}
                        if I_BATCH:
                            item_kwargs[I_BATCH] = getattr(batch, batch_id_field)
                        if I_PART:
                            item_kwargs[I_PART] = getattr(part, "id", None)
                        else:
                            if I_PN:
                                item_kwargs[I_PN] = incoming_pn
                            if I_PNAME:
                                item_kwargs[I_PNAME] = incoming_name
                        if I_QTY:
                            item_kwargs[I_QTY] = incoming_qty
                        if I_COST and (incoming_cost is not None):
                            item_kwargs[I_COST] = float(incoming_cost)
                        if I_LOC:
                            item_kwargs[I_LOC] = incoming_loc

                        if item_kwargs.get(I_QTY, 0) > 0:
                            session.add(ItemModel(**item_kwargs))
                    except IntegrityError as e:
                        session.rollback()
                        logging.exception("DB error while creating ReceivingItem: %s", e)
                        flash("DB error while creating ReceivingItem (см. лог).", "danger")
                    except Exception as e:
                        session.rollback()
                        logging.exception("Failed to create ReceivingItem: %s", e)
                        flash("Не удалось создать строку приёмки (см. лог).", "warning")

                meta = {
                    "file": m.get("source_file"),
                    "supplier": supplier_hint,
                    "invoice": (invoice_guess or "").strip(),
                }
                try:
                    if batch is not None and batch_id_field:
                        meta["batch_id"] = getattr(batch, batch_id_field)
                except Exception:
                    pass
                add_key(m["row_key"], meta)

            built, errors = build_receive_movements(
                norm,
                duplicate_exists_func=duplicate_exists,
                make_movement_func=make_movement
            )

            try:
                session.commit()
            except Exception as e:
                session.rollback()
                logging.exception("Final commit failed: %s", e)
                flash("Ошибка при окончательной записи в базу. Изменения откатились.", "danger")
                return redirect(url_for("inventory.import_parts_upload"))

            bid = None
            if batch is not None and batch_id_field:
                try:
                    bid = getattr(batch, batch_id_field or "id", None)
                    fresh_batch = BatchModel.query.get(bid)
                    if fresh_batch is not None:
                        B_STAT  = _pick_field(BatchModel, ["status","state"])
                        B_P_AT  = _pick_field(BatchModel, ["posted_at"])
                        B_P_BY  = _pick_field(BatchModel, ["posted_by"])
                        B_EXTRA = _pick_field(BatchModel, ["extra_expenses","shipping_fee","freight","expenses"])

                        if B_STAT:
                            setattr(fresh_batch, B_STAT, "posted")
                        if B_P_AT:
                            setattr(fresh_batch, B_P_AT, datetime.utcnow())
                        if B_P_BY:
                            try:
                                uid = getattr(current_user, "id", 0)
                            except Exception:
                                uid = 0
                            setattr(fresh_batch, B_P_BY, uid)

                        if B_EXTRA and getattr(fresh_batch, B_EXTRA, 0) not in (0, 0.0, None):
                            setattr(fresh_batch, B_EXTRA, 0.0)

                        session.add(fresh_batch)
                        session.commit()
                except Exception as e:
                    session.rollback()
                    logging.exception("Failed to finalize batch status: %s", e)
                    flash("Приход сохранён, но статус партии не удалось установить 'posted'.", "warning")

            for e in errors:
                flash(e, "danger")

            if bid is not None:
                flash(
                    f"Stock received and posted. Batch #{bid}. Создано строк: {len(built)}",
                    "success"
                )
                return redirect(url_for("inventory.receiving_detail", batch_id=bid))
            else:
                flash(
                    f"Создано приходов: {len(built)} (без записи ReceivingBatch)",
                    "warning"
                )
                return redirect(url_for("inventory.import_parts_upload"))

    # ===== B) первая загрузка файла → превью ==================================
    if request.method == "POST":
        f = request.files.get("file")
        if not f or not f.filename.strip():
            flash("Выберите файл (.pdf, .xlsx, .xls или .csv)", "warning")
            return redirect(request.url)

        filename   = secure_filename(f.filename)
        upload_dir = current_app.config.get(
            "UPLOAD_FOLDER",
            os.path.join(current_app.instance_path, "uploads")
        )
        os.makedirs(upload_dir, exist_ok=True)
        path = os.path.join(upload_dir, filename)
        f.save(path)

        ext = os.path.splitext(path)[1].lower()
        df  = dataframe_from_pdf(path, try_ocr=False) if ext == ".pdf" else load_table(path)

        supplier_hint = _detect_supplier_from_content(path, df)
        default_loc   = _supplier_to_default_location(supplier_hint)
        flash(
            f"Supplier hint: {supplier_hint or 'None'}, default location: {default_loc}",
            "info"
        )

        df = drop_vendor_noise_rows(df)
        df = fix_pn_and_description_in_df(df)

        if df is None or df.empty:
            flash("В этом файле не удалось распознать таблицы (возможно отсканированный PDF).", "danger")
            return render_template(
                "import_preview.html",
                rows=[],
                saved_path=path,
                supplier_hint=supplier_hint or "",
                extra_expenses=0.0,
                subtotal_base=0.0,
                grand_total=0.0,
            )

        norm, issues = normalize_table(
            df,
            supplier_hint=supplier_hint,
            source_file=path,
            default_location=default_loc,
        )
        for msg in issues:
            flash(msg, "warning")

        norm = _ensure_norm_columns(norm, default_loc, path)
        norm, subtotal_base, grand_total = _distribute_extra_and_adjust_costs(norm, 0.0)

        rows = norm.to_dict(orient="records")
        rows = fix_norm_records(rows, default_loc)

        # НОВОЕ: авто-догадка Invoice #
        invoice_guess = _infer_invoice_number(norm, path, supplier_hint) or ""

        return render_template(
            "import_preview.html",
            rows=rows,
            saved_path=path,
            supplier_hint=supplier_hint or "",
            extra_expenses=0.0,
            subtotal_base=subtotal_base,
            grand_total=grand_total,
            invoice_number=invoice_guess,
        )


    # ===== C) GET → показать форму загрузки ===================================
    return render_template("import_parts.html")

@inventory_bp.get("/orders/", endpoint="list_orders")
def list_orders():
    from flask import request, render_template
    from extensions import db
    # from models.order_item import OrderItem

    q = (request.args.get("q") or "").strip()
    base = OrderItem.query.order_by(OrderItem.date_ordered.desc(), OrderItem.id.desc())
    if q:
        like = f"%{q}%"
        base = base.filter(
            db.or_(
                OrderItem.technician.ilike(like),
                OrderItem.order_number.ilike(like),
            )
        )
    items = base.limit(500).all()
    return render_template("orders_list.html", items=items, q=q)

# ========= ORDERS: отметить как получено (и пополнить склад) =========
@inventory_bp.post("/orders/<int:item_id>/received", endpoint="mark_received")
def mark_received(item_id):
    from flask import redirect, url_for, flash
    from flask_login import current_user
    from datetime import datetime
    from models import db
    from models.order_item import OrderItem
    from models import Part  # твоя существующая модель склада

    # Помощник: подобрать имена полей из твоей модели Part
    PN_FIELDS   = ["part_number","number","sku","code","partnum","pn"]
    QTY_FIELDS  = ["quantity","qty","on_hand","stock","count"]
    LOC_FIELDS  = ["location","bin","shelf","place","loc"]
    NAME_FIELDS = ["name","part_name","descr","description","title"]
    COST_FIELDS = ["unit_cost","cost","price","unitprice","last_cost"]
    SUP_FIELDS  = ["supplier","vendor","provider"]
    def pick_field(model, candidates):
        for f in candidates:
            if hasattr(model, f): return f
        return None

    item = OrderItem.query.get_or_404(item_id)
    if item.status == "received":
        flash("Already received.", "info")
        return redirect(url_for("inventory.list_orders"))

    pn_field   = pick_field(Part, PN_FIELDS)
    qty_field  = pick_field(Part, QTY_FIELDS)
    loc_field  = pick_field(Part, LOC_FIELDS)
    name_field = pick_field(Part, NAME_FIELDS)
    cost_field = pick_field(Part, COST_FIELDS)
    sup_field  = pick_field(Part, SUP_FIELDS)
    if not pn_field or not qty_field:
        flash("Part model missing PN or QTY field.", "danger")
        return redirect(url_for("inventory.list_orders"))

    # Найти/создать позицию на складе (по PN + локация, если есть)
    filters = {pn_field: item.part_number}
    if loc_field and item.location:
        filters[loc_field] = item.location
    part = Part.query.filter_by(**filters).first()
    if not part:
        kwargs = dict(filters)
        kwargs[qty_field] = 0
        if name_field and item.part_name: kwargs[name_field] = item.part_name
        if cost_field and item.unit_cost is not None: kwargs[cost_field] = float(item.unit_cost)
        if sup_field and item.supplier: kwargs[sup_field] = item.supplier
        part = Part(**kwargs)
        db.session.add(part)
        db.session.flush()

    # + остаток
    setattr(part, qty_field, (getattr(part, qty_field) or 0) + int(item.qty_ordered or 0))

    # статус заказа
    item.status = "received"
    item.date_received = datetime.utcnow()
    who = getattr(current_user, "email", "system")
    item.notes = (item.notes or "")
    if who not in (item.notes or ""):
        item.notes = (item.notes + f"\nReceived by {who} at {item.date_received.isoformat()}").strip()

    db.session.commit()
    flash("Marked as received and stock updated.", "success")
    return redirect(url_for("inventory.list_orders"))

@inventory_bp.route("/import-settings", methods=["GET", "POST"], endpoint="import_settings")
@login_required
def import_settings():
    # простая защита: только админ/суперадмин
    if getattr(current_user, "role", None) not in (ROLE_ADMIN, ROLE_SUPERADMIN):
        abort(403)

    if request.method == "POST":
        ocr_on = 1 if request.form.get("pdf_ocr_enabled") == "on" else 0
        set_setting("pdf_ocr_enabled", bool(ocr_on))
        flash("Настройки сохранены.", "success")
        return redirect(url_for("inventory.import_settings"))

    ocr_enabled = bool(get_setting("pdf_ocr_enabled", False))
    return render_template("import_settings.html", ocr_enabled=ocr_enabled)
@inventory_bp.route('/reports/create_return_invoice', methods=['POST'])
@login_required
def create_return_invoice():
    if current_user.role not in ['superadmin', 'admin']:
        flash("Access denied", "danger")
        return redirect(url_for('inventory.reports'))

    # идентификаторы исходной группы (на всякий случай)
    issued_to_old     = (request.form.get('issued_to_old') or request.form.get('issued_to') or '').strip()
    reference_job_old = (request.form.get('reference_job_old') or request.form.get('reference_job') or '').strip() or None
    s                 = (request.form.get('issue_date_old') or request.form.get('issue_date') or '').strip()
    issued_by         = (request.form.get('issued_by') or '').strip()
    issue_date_old    = _parse_dt_flex(s)

    # выбранные строки для возврата
    line_ids = request.form.getlist('line_ids') or request.form.getlist('line_ids[]')
    if not line_ids:
        flash("No lines selected.", "warning")
        return redirect(url_for('inventory.reports'))

    try:
        line_ids = [int(x) for x in line_ids]
    except ValueError:
        flash("Invalid selection.", "danger")
        return redirect(url_for('inventory.reports'))

    # под каждую строку возьмём qty_X
    rows = IssuedPartRecord.query.filter(IssuedPartRecord.id.in_(line_ids)).all()
    if not rows:
        flash("Selected lines not found.", "warning")
        return redirect(url_for('inventory.reports'))

    created = 0
    for r in rows:
        qty_raw = request.form.get(f"qty_{r.id}")
        try:
            qty = int(qty_raw or 0)
        except Exception:
            qty = 0
        qty = max(0, min(qty, int(r.quantity or 0)))  # нельзя вернуть больше, чем выдали
        if qty <= 0:
            continue

        # создаём возвратную запись (отрицательное количество)
        ret = IssuedPartRecord(
            part_id=r.part_id,
            quantity=-qty,
            issued_to=r.issued_to,
            reference_job=f"RETURN {r.reference_job or ''}".strip(),
            issued_by=current_user.username,
            issue_date=datetime.utcnow(),
            unit_cost_at_issue=r.unit_cost_at_issue,
        )
        # при возврате на склад добавим кол-во обратно
        part = Part.query.get(r.part_id)
        if part:
            part.quantity = int(part.quantity or 0) + qty

        db.session.add(ret)
        created += 1

    if created == 0:
        flash("Nothing to return.", "warning")
        return redirect(url_for('inventory.reports'))

    db.session.commit()
    flash(f"Return invoice created ({created} lines).", "success")
    return redirect(url_for('inventory.reports'))

@inventory_bp.route('/reports/delete_return_invoice', methods=['POST'])
@login_required
def delete_return_invoice():
    if current_user.role != 'superadmin':
        flash("Access denied", "danger")
        return redirect(url_for('inventory.reports_grouped'))

    raw_ids = request.form.getlist('record_ids[]') or request.form.getlist('record_ids')
    try:
        ids = [int(x) for x in raw_ids if str(x).strip()]
    except ValueError:
        ids = []

    if not ids:
        flash("Nothing selected to delete.", "warning")
        return redirect(url_for('inventory.reports_grouped'))

    rows = IssuedPartRecord.query.filter(IssuedPartRecord.id.in_(ids)).all()
    if not rows:
        flash("Records not found.", "warning")
        return redirect(url_for('inventory.reports_grouped'))

    touched_batch_ids = set()
    try:
        for r in rows:
            # удаляем только возвратные строки
            if (r.quantity or 0) >= 0 and not ((r.reference_job or '').upper().startswith('RETURN')):
                continue

            # откатываем склад: удаляем возврат → уменьшаем склад
            if r.part and (r.quantity or 0) < 0:
                r.part.quantity = (r.part.quantity or 0) - abs(int(r.quantity or 0))

            if r.batch_id:
                touched_batch_ids.add(r.batch_id)

            db.session.delete(r)

        # чистим пустые батчи
        if touched_batch_ids:
            for bid in list(touched_batch_ids):
                still_has = db.session.query(IssuedPartRecord.id).filter_by(batch_id=bid).first()
                if not still_has:
                    b = db.session.get(IssuedBatch, bid)
                    if b:
                        db.session.delete(b)

        db.session.commit()
        flash("Return invoice deleted.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Failed to delete return invoice: {e}", "danger")

    return redirect(url_for('inventory.reports_grouped'))

@inventory_bp.route('/reports/return_selected', methods=['POST'])
@login_required
def return_selected():
    if current_user.role not in ['superadmin', 'admin']:
        flash("Access denied", "danger")
        return redirect(url_for('inventory.reports_grouped'))

    # ✅ ADD (safe)
    from sqlalchemy import func
    from models import ReturnDestination
    from datetime import datetime, timezone
    from zoneinfo import ZoneInfo

    LA_TZ = ZoneInfo("America/Los_Angeles")

    now_la = datetime.now(LA_TZ)  # aware LA time
    now_dt = now_la.astimezone(timezone.utc).replace(tzinfo=None)  # naive UTC for DB

    raw_ids = request.form.getlist('record_ids[]') or request.form.getlist('record_ids')
    try:
        selected_ids = [int(x) for x in raw_ids if str(x).strip()]
    except ValueError:
        selected_ids = []

    if not selected_ids:
        flash("No rows selected.", "warning")
        return redirect(url_for('inventory.reports_grouped'))

    rows = IssuedPartRecord.query.filter(IssuedPartRecord.id.in_(selected_ids)).all()
    if not rows:
        flash("Selected records not found.", "warning")
        return redirect(url_for('inventory.reports_grouped'))


    issued_by = getattr(current_user, 'username', '') or 'system'
    first = rows[0]
    batch_issued_to = first.issued_to
    batch_reference = (first.reference_job or '').strip() or 'STOCK'
    return_reference = f"RETURN {batch_reference.upper()}"
    batch_location   = first.location

    created = []
    for src in rows:
        # то, что пользователь попросил вернуть
        try:
            want = int(request.form.get(f"qty_{src.id}", "1") or "1")
        except ValueError:
            want = 1
        if want <= 0:
            continue

        # Сколько уже возвращено по этой детали/получателю
        already = _already_returned_qty_for_source(src)

        # Максимум к возврату сейчас = выдано по строке − уже возвращено (не ниже 0)
        issued_qty = int(src.quantity or 0)
        max_can_return = max(0, issued_qty - already)

        qty = min(want, max_can_return)
        if qty == 0:
            continue

        src_inv = (src.invoice_number if src.invoice_number is not None else None)

        # --- return destination meta (per source row) ---
        r_to = (request.form.get(f"return_to_{src.id}") or "").strip().upper()

        # NOTE: can be ID (digits) or free text name (when DB empty)
        r_dest_raw = (request.form.get(f"return_dest_{src.id}") or "").strip()

        # ✅ IMPORTANT: define dest_id always (avoid UnboundLocalError)
        dest_id = None

        # required only if qty > 0 (то есть реально возвращаем)
        if qty > 0:
            if r_to not in ("STOCK", "VENDOR"):
                flash("Please select Return To (STOCK or VENDOR) for all returned lines.", "danger")
                db.session.rollback()
                return redirect(url_for('inventory.reports_grouped'))

            if r_to == "VENDOR":
                if not r_dest_raw:
                    flash("Please select Vendor Company for all VENDOR return lines.", "danger")
                    db.session.rollback()
                    return redirect(url_for('inventory.reports_grouped'))

                # 1) dropdown: digits => ReturnDestination.id
                if r_dest_raw.isdigit():
                    dest_id = int(r_dest_raw)
                else:
                    # 2) free text => find/create ReturnDestination by name
                    name = r_dest_raw.strip()
                    existing = (
                        db.session.query(ReturnDestination)
                        .filter(func.lower(ReturnDestination.name) == name.lower())
                        .first()
                    )
                    if not existing:
                        existing = ReturnDestination(name=name, is_active=True)
                        db.session.add(existing)
                        db.session.flush()  # get id without commit
                    dest_id = existing.id

        ret = IssuedPartRecord(
            part_id=src.part_id,
            quantity=-qty,
            issued_to=src.issued_to,
            issued_by=issued_by,
            reference_job=return_reference,
            issue_date=now_dt,
            unit_cost_at_issue=src.unit_cost_at_issue,
            location=src.location,

            # ✅ связь возврата с исходным invoice (для корректного already-returned)
            inv_ref=(str(src_inv) if src_inv is not None else None),
            return_to=r_to,
            return_destination_id=dest_id,
        )

        db.session.add(ret)
        created.append(ret)

        # Возвращаем на склад
        if src.part:
            src.part.quantity = (src.part.quantity or 0) + qty

    if not created:
        flash("Nothing to return.", "warning")
        db.session.rollback()
        return redirect(url_for('inventory.reports_grouped'))

    try:
        batch = _create_batch_for_records(
            created,
            issued_to=batch_issued_to,
            issued_by=issued_by,
            reference_job=return_reference,
            issue_date=now_dt,
            location=batch_location,
        )
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f"Failed to create return invoice: {e}", "danger")
        # ✅ don't reference batch if it failed before assignment
        return redirect(url_for('inventory.reports_grouped'))

    flash(f"Return invoice #{batch.invoice_number} created ({len(created)} lines).", "success")
    return redirect(url_for('inventory.reports_grouped', invoice_number=batch.invoice_number))

def _already_returned_qty_for_source(src) -> int:
    """
    Сколько уже возвращено по той же детали этому же получателю.

    Новый безопасный алгоритм:
    1) Если у исходной строки есть invoice_number — сначала считаем RETURN только
       внутри этого invoice (через inv_ref == invoice_number).
    2) Если (1) дал 0 — fallback на старую логику (по part_id + issued_to + RETURN%),
       чтобы старые возвраты без inv_ref продолжали учитываться и ничего не сломалось.
    """

    # --- 1) точный подсчет по исходному invoice (если он известен) ---
    src_inv = getattr(src, "invoice_number", None)
    if src_inv is not None:
        inv_key = str(int(src_inv))  # нормализуем в "1552"
        total_neg_strict = (
            db.session.query(func.coalesce(func.sum(IssuedPartRecord.quantity), 0))
            .filter(
                IssuedPartRecord.part_id == src.part_id,
                IssuedPartRecord.issued_to == src.issued_to,
                IssuedPartRecord.reference_job.ilike("RETURN%"),
                IssuedPartRecord.quantity < 0,
                IssuedPartRecord.inv_ref == inv_key,
            )
            .scalar()
            or 0
        )

        strict_abs = abs(int(total_neg_strict))
        if strict_abs > 0:
            return strict_abs

        # если строго 0 — идем в fallback (для старых возвратов без inv_ref)

    # --- 2) fallback (старое поведение) ---
    total_neg_fallback = (
        db.session.query(func.coalesce(func.sum(IssuedPartRecord.quantity), 0))
        .filter(
            IssuedPartRecord.part_id == src.part_id,
            IssuedPartRecord.issued_to == src.issued_to,
            IssuedPartRecord.reference_job.ilike("RETURN%"),
            IssuedPartRecord.quantity < 0,
        )
        .scalar()
        or 0
    )
    return abs(int(total_neg_fallback))


@inventory_bp.get("/receiving/by-invoice/<path:inv>", endpoint="receiving_by_invoice")
@login_required
def receiving_by_invoice(inv):
    from flask import redirect, url_for, flash, request
    from sqlalchemy import func
    from extensions import db
    from models import GoodsReceipt  # или ReceivingBatch (как у тебя модель называется)

    inv = (inv or "").strip()
    if not inv:
        return redirect(url_for("inventory.receiving_list"))

    # ищем самый свежий batch по invoice_number (строкой)
    batch = (
        db.session.query(GoodsReceipt)
        .filter(func.trim(GoodsReceipt.invoice_number) == inv)
        .order_by(
            GoodsReceipt.invoice_date.desc().nullslast(),
            GoodsReceipt.id.desc()
        )
        .first()
    )

    if not batch:
        flash(f"Receiving invoice not found: {inv}", "warning")
        # удобно вернуть в список receiving уже с поиском
        return redirect(url_for("inventory.receiving_list", q=inv))

    # ⚠️ имя endpoint может отличаться: receiving_detail / receiving_view / receiving_edit
    return redirect(url_for("inventory.receiving_detail", batch_id=batch.id))


@inventory_bp.get("/receiving", endpoint="receiving_list")
@login_required
def receiving_list():
    from sqlalchemy import func  # or_ можно удалить, если больше не нужен

    current_app.logger.debug(
        "### DEBUG receiving_list USING MODEL %s FROM %s TABLENAME=%s",
        ReceivingBatch, __file__, getattr(ReceivingBatch, "__tablename__", "?")
    )

    q = ReceivingBatch.query
    DEFAULT_LIMIT = 200

    # ОДНА строка поиска
    global_q = (request.args.get("q") or "").strip()

    d1     = (request.args.get("date_from") or "").strip()
    d2     = (request.args.get("date_to") or "").strip()
    status = (request.args.get("status") or "").strip()

    # --- статус ---
    if status in ("draft", "posted"):
        q = q.filter(ReceivingBatch.status == status)

    # --- даты ---
    def _parse_date(s):
        try:
            return datetime.strptime(s, "%Y-%m-%d").date()
        except Exception:
            return None

    d1p, d2p = _parse_date(d1), _parse_date(d2)
    if d1p:
        q = q.filter(or_(ReceivingBatch.invoice_date.is_(None),
                         ReceivingBatch.invoice_date >= d1p))
    if d2p:
        q = q.filter(or_(ReceivingBatch.invoice_date.is_(None),
                         ReceivingBatch.invoice_date <= d2p))

    # Забираем батчи из БД (без текстового фильтра)
    qry = q.order_by(
        ReceivingBatch.invoice_date.desc().nullslast(),
        ReceivingBatch.id.desc()
    )

    # IMPORTANT: limit only when NO filters (иначе пропадают “старые” инвойсы)
    if not global_q and not d1p and not d2p and not status:
        qry = qry.limit(DEFAULT_LIMIT)

    batches = qry.all()

    # --- текстовый поиск по Supplier / Invoice / Part # / Part Name ---
    if global_q:
        needle = global_q.lower()
        filtered = []

        for b in batches:
            # supplier / invoice
            sup = (getattr(b, "supplier_name", "") or "").lower()
            inv = (getattr(b, "invoice_number", "") or "").lower()

            match_batch = (needle in sup) or (needle in inv)

            # если по заголовку не нашли — ищем в строках
            if not match_batch:
                lines = getattr(b, "items", None) or getattr(b, "lines", None) or []
                for ln in lines:
                    pn = (
                        getattr(ln, "part_number", None)
                        or getattr(ln, "pn", None)
                        or ""
                    )
                    name = (
                        getattr(ln, "part_name", None)
                        or (
                            getattr(getattr(ln, "part", None), "name", None)
                            if getattr(ln, "part", None) is not None
                            else None
                        )
                        or ""
                    )
                    if needle in str(pn).lower() or needle in str(name).lower():
                        match_batch = True
                        break

            if match_batch:
                filtered.append(b)

        batches = filtered

    # --- Totals ---
    totals = {}
    for b in batches:
        try:
            lines = getattr(b, "items", []) or getattr(b, "lines", []) or []
            total = 0.0
            for ln in lines:
                qty  = getattr(ln, "qty", None) or getattr(ln, "quantity", 0) or 0
                cost = getattr(ln, "unit_cost", None) or getattr(ln, "price", None) or 0.0
                try:
                    total += float(cost) * int(qty)
                except Exception:
                    pass
            totals[b.id] = total
        except Exception:
            totals[b.id] = 0.0

    # --- права админа ---
    can_admin = bool(
        getattr(current_user, "is_superadmin", False) or
        getattr(current_user, "is_super_admin", False) or
        getattr(current_user, "is_admin", False)
    )

    return render_template(
        "receiving_list.html",
        batches=batches,
        totals=totals,
        can_admin=can_admin,
        limit=DEFAULT_LIMIT,
        filters={
            "q": global_q,
            "date_from": d1,
            "date_to": d2,
            "status": status,
        }
    )

# --- Receiving: toggle posted/draft (superadmin only) ------------------------
# --- helper: запрещаем unpost, если уже было списание в IssuedPartRecord -----

def _batch_consumed_forbid_unpost(batch) -> bool:
    """
    Возвращает True если из ЭТОГО прихода уже что-то ушло техникам.
    В этом случае unpost НЕЛЬЗЯ делать (иначе сломаешь остатки).

    Логика:
    - соберём все part_number из строк прихода (batch.items / batch.lines)
    - найдём Part.id для этих part_number
    - проверим IssuedPartRecord по этим Part.id
    - если есть хотя бы одна выдача -> True
    """
    from sqlalchemy import func
    from extensions import db
    from models import Part, IssuedPartRecord

    # получаем строки партии
    lines = (
        getattr(batch, "lines", None)
        or getattr(batch, "items", None)
        or []
    )

    # соберём PN из строк
    part_numbers_upper = []
    for it in lines:
        pn = (getattr(it, "part_number", "") or "").strip()
        if pn:
            part_numbers_upper.append(pn.upper())

    # если нет нормальных PN -> считаем что расход не доказан => unpost разрешаем
    if not part_numbers_upper:
        return False

    # найдём ID деталей по PN
    part_ids = [
        row[0]
        for row in (
            db.session.query(Part.id)
            .filter(func.upper(func.trim(Part.part_number)).in_(part_numbers_upper))
            .all()
        )
    ]

    if not part_ids:
        # не нашли Part вообще -> считаем что нечего сверять => не блокируем
        return False

    # ищем хоть одну выдачу по этим Part.id
    used = (
        db.session.query(IssuedPartRecord.id)
        .filter(IssuedPartRecord.part_id.in_(part_ids))
        .limit(1)
        .first()
    )

    # True => УЖЕ ЕСТЬ РАСХОД -> блокируем unpost
    return used is not None


# --- Receiving: post/unpost toggle -------------------------------------------

@inventory_bp.post("/receiving/<int:batch_id>/toggle", endpoint="receiving_toggle")
@login_required
def receiving_toggle(batch_id: int):
    from flask import flash, redirect, url_for, current_app
    from flask_login import current_user
    from extensions import db
    from models import ReceivingBatch
    from services.receiving import post_receiving_batch, unpost_receiving_batch

    log = current_app.logger

    # --- кто может жать Post / Unpost через toggle ---
    role_low = (getattr(current_user, "role", "") or "").strip().lower()
    is_adminish = (
        role_low in ("admin", "superadmin")
        or getattr(current_user, "is_admin", False)
        or getattr(current_user, "is_superadmin", False)
        or getattr(current_user, "is_super_admin", False)
    )
    if not is_adminish:
        flash("Access denied. Admin only.", "danger")
        return redirect(url_for("inventory.receiving_detail", batch_id=batch_id))

    batch = db.session.get(ReceivingBatch, batch_id)
    if not batch:
        flash("Batch not found.", "warning")
        return redirect(url_for("inventory.receiving_list"))

    status_now = (getattr(batch, "status", "") or "").strip().lower()
    was_posted_at = getattr(batch, "posted_at", None)

    # =========================================================
    # CASE 1: сейчас не POSTED -> хотим POST (приход на склад)
    # =========================================================
    if status_now != "posted":
        # защита от двойного прихода:
        # если batch уже когда-то был проведён (posted_at не пустой),
        # то склад повторно не трогаем, просто вернём статус "posted".
        if was_posted_at:
            log.warning(
                "[RECEIVING_TOGGLE] Batch %s already had stock applied earlier "
                "(posted_at=%s). Setting status back to POSTED without re-applying stock.",
                batch.id, was_posted_at
            )
            batch.status = "posted"
            db.session.commit()
            flash("Batch marked as POSTED (stock was already applied earlier).", "info")
            return redirect(url_for("inventory.receiving_detail", batch_id=batch.id))

        # нормальный путь: первая проводка -> плюсуем остаток через сервис
        try:
            post_receiving_batch(
                batch.id,
                getattr(current_user, "id", None)
            )
            flash(
                f"Batch #{batch.id} posted and stock updated.",
                "success"
            )
        except Exception as e:
            db.session.rollback()
            log.exception(
                "[RECEIVING_TOGGLE] post failed for batch %s",
                batch.id
            )
            flash(f"Failed to post: {e}", "danger")

        return redirect(url_for("inventory.receiving_detail", batch_id=batch.id))

    # =========================================================
    # CASE 2: сейчас POSTED -> хотим UNPOST (снять со склада)
    # =========================================================
    # helper: супер-админ?
    def is_superadmin(u) -> bool:
        r = (getattr(u, "role", "") or "").strip().lower()
        return (
            r == "superadmin"
            or getattr(u, "is_superadmin", False)
            or getattr(u, "is_super_admin", False)
        )

    # Если партия уже "потреблена" (из неё что-то выдали технику после постинга),
    # то мы не даём обычному админу снимать склад, чтобы не уйти в минус.
    # superadmin может сделать форс-чистку через отдельную кнопку /receiving/<id>/unpost.
    if _batch_consumed_forbid_unpost(batch):
        log.debug(
            "[RECEIVING_TOGGLE] Blocked UNPOST for batch %s: already consumed.",
            batch.id,
        )

        if not is_superadmin(current_user):
            flash(
                "Cannot unpost: items from this batch were already issued to technicians.",
                "danger"
            )
            return redirect(url_for("inventory.receiving_detail", batch_id=batch.id))

        # супер-админу не делаем авто-rollback тут,
        # чтобы не словить неожиданный минус: он должен нажать свою отдельную Unpost кнопку,
        # которая идёт на /receiving/<id>/unpost (там только superadmin).
        flash(
            "This batch was already consumed. Use SUPERADMIN Unpost button instead.",
            "warning"
        )
        return redirect(url_for("inventory.receiving_detail", batch_id=batch.id))

    # нормально: партия posted и не была использована -> можно безопасно снять со склада
    try:
        unpost_receiving_batch(
            batch.id,
            getattr(current_user, "id", None)
        )
        flash(
            f"Batch #{batch.id} reverted to draft and stock rolled back.",
            "warning"
        )
    except Exception as e:
        db.session.rollback()
        log.exception(
            "[RECEIVING_TOGGLE] unpost failed for batch %s",
            batch.id
        )
        flash(f"Failed to unpost: {e}", "danger")

    return redirect(url_for("inventory.receiving_detail", batch_id=batch.id))

@inventory_bp.get("/receiving/new", endpoint="receiving_new")
@login_required
def receiving_new():
    from flask import flash, redirect, url_for
    from flask_login import current_user

    role = (getattr(current_user, "role", "") or "").lower()

    # Разрешаем admin и superadmin заходить на страницу создания НОВОГО батча
    if role not in ("admin", "superadmin"):
        flash("Access denied. only admin or superadmin can create Receiving batch.", "danger")
        return redirect(url_for("inventory.receiving_list"))

    # НИЧЕГО не флэшить про "only superadmin can edit Receiving"
    # потому что это не редактирование существующего, это НОВЫЙ (batch=None)

    return render_template(
        "receiving_edit.html",
        batch=None,
        today=datetime.utcnow().date()
    )


@inventory_bp.post("/receiving/<int:batch_id>/unpost", endpoint="receiving_unpost")
@login_required
def receiving_unpost(batch_id: int):
    if not _is_superadmin():
        flash("Access denied: only superadmin can unpost.", "danger")
        return redirect(url_for("inventory.receiving_detail", batch_id=batch_id))

    try:
        unpost_receiving_batch(batch_id, getattr(current_user, "id", None))
        flash("Batch unposted & stock rolled back.", "success")
    except Exception as e:
        current_app.logger.exception("Unpost failed")
        flash(f"Unpost failed: {e}", "danger")
    return redirect(url_for("inventory.receiving_detail", batch_id=batch_id))
@inventory_bp.post("/receiving/save", endpoint="receiving_save")
@login_required
def receiving_save():
    """
    Создаёт или обновляет ReceivingBatch (GoodsReceipt) + строки.
    Добавлено:
      - extra_expenses (shipping/etc) лежит в batch.extra_expenses
      - unit_cost строк пересчитывается с учётом распределения extra_expenses
        пропорционально (qty * base_cost).
    Если action == "post", вызываем post_receiving_batch() чтобы
    списать в склад и пометить batch как posted.
    """
    from flask import request, redirect, url_for, flash
    from datetime import datetime
    from flask_login import current_user

    from extensions import db
    from models import ReceivingBatch, ReceivingItem, Part
    from sqlalchemy import func

    today = datetime.utcnow().date()

    def _parse_invoice_date(raw: str | None):
        raw = (raw or "").strip()
        if not raw:
            return None
        for fmt in ("%Y-%m-%d", "%m/%d/%Y"):
            try:
                return datetime.strptime(raw, fmt).date()
            except Exception:
                continue
        return None

    f = request.form
    batch_id = (f.get("batch_id") or "").strip()
    action = (f.get("action") or "").strip().lower()
    force_existing = (f.get("force_existing") or "").strip() == "1"

    # читаем extra_expenses как float
    try:
        extra_expenses_total = float(f.get("extra_expenses") or 0.0)
    except Exception:
        extra_expenses_total = 0.0
    if extra_expenses_total < 0:
        extra_expenses_total = 0.0

    # ---------------- ROLE HELPERS ----------------
    role_low = (getattr(current_user, "role", "") or "").strip().lower()

    def is_super(u) -> bool:
        rl = (getattr(u, "role", "") or "").strip().lower()
        if rl == "superadmin":
            return True
        if getattr(u, "is_superadmin", False):
            return True
        if getattr(u, "is_super_admin", False):
            return True
        return False

    def is_adminish(u) -> bool:
        rl = (getattr(u, "role", "") or "").strip().lower()
        if rl in ("admin", "superadmin"):
            return True
        if getattr(u, "is_admin", False):
            return True
        if getattr(u, "is_superadmin", False):
            return True
        if getattr(u, "is_super_admin", False):
            return True
        return False

    user_is_super = is_super(current_user)
    user_is_adminish = is_adminish(current_user)

    # figure out: new batch or editing existing
    is_new_batch = (not batch_id) and (not force_existing)

    # ------------- ACCESS CONTROL -------------
    # 1) новый батч может создать admin / superadmin
    # 2) существующий батч:
    #    - draft: admin или superadmin может редактировать
    #    - posted:
    #         обычные админы редактировать не могут;
    #         superadmin может править ТОЛЬКО шапку (supplier, invoice, date, notes, currency)
    #         без изменения строк.
    if is_new_batch:
        if not user_is_adminish:
            flash("Access denied. Only admin or superadmin can create receiving.", "danger")
            return redirect(url_for("inventory.receiving_list"))

        batch = ReceivingBatch(created_by=getattr(current_user, "id", None))
        batch.status = "draft"
        batch.created_at = datetime.utcnow()

        # header fields
        batch.supplier_name = (f.get("supplier_name") or "").strip() or "UNKNOWN"
        batch.invoice_number = (f.get("invoice_number") or "").strip() or None

        # --- Валидация даты: не позволяем будущее ---
        inv_date_val = _parse_invoice_date(f.get("invoice_date"))
        if inv_date_val and inv_date_val > today:
            flash("Invoice date cannot be in the future.", "danger")
            return redirect(url_for("inventory.receiving_new"))
        batch.invoice_date = inv_date_val

        batch.currency = (f.get("currency") or "USD").strip()[:8] or "USD"
        batch.notes    = (f.get("notes") or "").strip() or None

    else:
        # existing batch
        try:
            batch = ReceivingBatch.query.get(int(batch_id))
        except Exception:
            batch = None

        if not batch:
            flash("Receiving batch not found.", "danger")
            return redirect(url_for("inventory.receiving_list"))

        status_low = (getattr(batch, "status", "") or "").strip().lower()

        # режим "только шапка" для superadmin, когда партия уже posted
        header_only_mode = False
        if status_low == "posted":
            if not user_is_super:
                flash("Batch is posted. Unpost first, then edit.", "danger")
                return redirect(url_for("inventory.receiving_detail", batch_id=batch.id))
            header_only_mode = True

        # update header fields from form
        batch.supplier_name  = (f.get("supplier_name") or "").strip() or (batch.supplier_name or "UNKNOWN")
        batch.invoice_number = (f.get("invoice_number") or "").strip() or None

        # --- Валидация даты: не позволяем будущее ---
        inv_date_val = _parse_invoice_date(f.get("invoice_date"))
        if inv_date_val and inv_date_val > today:
            flash("Invoice date cannot be in the future.", "danger")
            return redirect(url_for("inventory.receiving_detail", batch_id=batch.id))
        batch.invoice_date = inv_date_val

        batch.currency = ((f.get("currency") or "USD").strip()[:8] or "USD")
        batch.notes    = (f.get("notes") or "").strip() or None

        # Если superadmin редактирует уже posted партию — меняем только шапку,
        # строки НЕ трогаем.
        if header_only_mode:
            db.session.add(batch)
            db.session.commit()
            flash("Batch header updated.", "success")
            return redirect(url_for("inventory.receiving_detail", batch_id=batch.id))

        # draft-режим: можно править всё, перестраиваем строки
        batch.items.clear()

    # ------------- ПАРСИМ СТРОКИ И СЧИТАЕМ РАСПРЕДЕЛЕНИЕ ----------------
    # сначала соберём все строки в память, чтобы:
    # 1) посчитать subtotal = сумма qty * base_cost
    # 2) потом рассчитать для каждой строки adjusted_cost
    tmp_rows = []
    idx = 0
    while True:
        basekey = f"rows[{idx}]"
        if not any(k.startswith(basekey) for k in f.keys()):
            break

        pn_val = (f.get(f"{basekey}[part_number]") or "").strip().upper()
        if pn_val:
            try:
                qty_val = int(f.get(f"{basekey}[quantity]") or 0)
            except Exception:
                qty_val = 0
            try:
                base_cost_val = float(f.get(f"{basekey}[unit_cost]") or 0.0)
            except Exception:
                base_cost_val = 0.0

            loc_val = (f.get(f"{basekey}[location]") or "").strip()[:64]
            pn_name = (f.get(f"{basekey}[part_name]") or "").strip()

            tmp_rows.append({
                "part_number": pn_val,
                "part_name": pn_name,
                "qty": qty_val,
                "base_cost": base_cost_val,
                "location": loc_val,
            })

        idx += 1

    # считаем subtotal
    subtotal = 0.0
    for r in tmp_rows:
        subtotal += (r["qty"] or 0) * (r["base_cost"] or 0.0)

    # теперь считаем adjusted cost для каждой строки
    for r in tmp_rows:
        qty = r["qty"] or 0
        base_cost = r["base_cost"] or 0.0
        adj_cost = base_cost
        if subtotal > 0 and qty > 0:
            line_total = qty * base_cost
            share = line_total / subtotal
            extra_for_line = extra_expenses_total * share
            per_item_fee = extra_for_line / qty
            adj_cost = base_cost + per_item_fee
        r["adj_cost"] = adj_cost

    # пишем extra_expenses в сам батч
    batch.extra_expenses = float(extra_expenses_total or 0.0)

    # создаём и добавляем ReceivingItem строки (уже с adj_cost)
    for r in tmp_rows:
        if not r["part_number"]:
            continue
        item = ReceivingItem(
            part_number = r["part_number"],
            part_name   = r["part_name"] or "",
            quantity    = int(r["qty"] or 0),
            unit_cost   = float(r["adj_cost"] or 0.0),  # ВАЖНО: уже с fee
            location    = r["location"] or "",
        )
        batch.items.append(item)

    # sync part names into Part table (non-destructive)
    if batch.items:
        for it in batch.items:
            pn_sync = (it.part_number or "").strip().upper()
            new_name = (it.part_name or "").strip()
            if not pn_sync or not new_name:
                continue
            part_obj = Part.query.filter(func.upper(Part.part_number) == pn_sync).first()
            if part_obj:
                old_name = (getattr(part_obj, "name", "") or "").strip()
                if old_name != new_name and new_name:
                    part_obj.name = new_name

    # сохранить draft (с extra_expenses и с adj_cost в строках)
    db.session.add(batch)
    db.session.commit()

    # если нажали "Receive & Post Stock"
    if action == "post":
        from services.receiving import post_receiving_batch
        post_receiving_batch(batch.id, getattr(current_user, "id", None))
        flash("Batch posted & stock updated.", "success")
    else:
        flash("Draft saved.", "success")

    return redirect(url_for("inventory.receiving_detail", batch_id=batch.id))

@inventory_bp.get("/receiving/<int:batch_id>", endpoint="receiving_detail")
@login_required
def receiving_detail(batch_id):
    from models import ReceivingBatch, Part, IssuedPartRecord
    from sqlalchemy import func
    from flask import current_app

    log = current_app.logger

    batch = db.session.get(ReceivingBatch, batch_id)
    if not batch:
        flash("Batch not found.", "warning")
        return redirect(url_for("inventory.receiving_list"))

    # собрать строки батча
    lines = (
        getattr(batch, "lines", None)
        or getattr(batch, "items", None)
        or []
    )

    # ---------- helpers to compute money safely ----------
    def _to_int(x):
        try:
            return int(x)
        except Exception:
            try:
                return int(float(str(x).replace(",", ".")))
            except Exception:
                return 0

    def _to_float(x):
        try:
            return float(x)
        except Exception:
            try:
                return float(str(x).replace(",", ".").replace("$", ""))
            except Exception:
                return 0.0

    batch_total = 0.0
    for it in lines:
        qty = _to_int(getattr(it, "quantity", None) or getattr(it, "qty", 0))
        cost = _to_float(
            getattr(it, "unit_cost", None) or getattr(it, "unit_cost_at_issue", 0.0)
        )
        batch_total += qty * cost

    # ---------- detect "consumed" ----------
    def _batch_has_been_consumed(_batch):
        """
        consumed = True если мы видим, что хотя бы одна из деталей из этой партии
        уже фигурирует в IssuedPartRecord (то есть была выдана технику).

        Логика:
        1. Собираем все part_id напрямую из строк партии (если там есть поле part_id).
        2. Собираем все part_number из строк партии, находим им Part.id.
        3. Берём объединение этих Part.id.
        4. Проверяем IssuedPartRecord по этим Part.id.
        """

        status_low_local = (getattr(_batch, "status", "") or "").strip().lower()
        if status_low_local != "posted":
            # не posted => не считаем как "использовано", ещё черновик
            log.warning(f"[RECV_DETAIL consume-check] batch {batch_id}: not posted -> consumed=False")
            print(f"[RECV_DETAIL consume-check] batch {batch_id}: not posted -> consumed=False")
            return False

        # (1) part_ids напрямую из строк, если есть
        direct_part_ids = []
        for row in (
            getattr(_batch, "lines", None)
            or getattr(_batch, "items", None)
            or []
        ):
            if hasattr(row, "part_id") and getattr(row, "part_id") is not None:
                try:
                    direct_part_ids.append(int(row.part_id))
                except Exception:
                    pass

        # (2) part_numbers -> Part.id
        raw_pns = []
        for row in (
            getattr(_batch, "lines", None)
            or getattr(_batch, "items", None)
            or []
        ):
            pn_val = getattr(row, "part_number", None)
            if pn_val:
                pn_norm = str(pn_val).strip().upper()
                if pn_norm:
                    raw_pns.append(pn_norm)

        # делаем уникальные
        raw_pns = list({p for p in raw_pns})
        direct_part_ids = list({pid for pid in direct_part_ids})

        log.warning(f"[RECV_DETAIL consume-check] batch {batch_id}: raw_pns={raw_pns}, direct_part_ids(pre)={direct_part_ids}")
        print(f"[RECV_DETAIL consume-check] batch {batch_id}: raw_pns={raw_pns}, direct_part_ids(pre)={direct_part_ids}")

        # (2b) достаём ID по PN
        part_ids_from_pn = []
        if raw_pns:
            rows_parts = (
                db.session.query(Part.id, Part.part_number)
                .filter(func.upper(func.trim(Part.part_number)).in_(raw_pns))
                .all()
            )
            part_ids_from_pn = [row[0] for row in rows_parts]

            log.warning(f"[RECV_DETAIL consume-check] batch {batch_id}: part_ids_from_pn={part_ids_from_pn}, parts_matched={[p for _, p in rows_parts]}")
            print(f"[RECV_DETAIL consume-check] batch {batch_id}: part_ids_from_pn={part_ids_from_pn}")

        # (3) объединяем
        all_part_ids = set(direct_part_ids) | set(part_ids_from_pn)
        all_part_ids = list(all_part_ids)

        log.warning(f"[RECV_DETAIL consume-check] batch {batch_id}: all_part_ids(final)={all_part_ids}")
        print(f"[RECV_DETAIL consume-check] batch {batch_id}: all_part_ids(final)={all_part_ids}")

        if not all_part_ids:
            log.warning(f"[RECV_DETAIL consume-check] batch {batch_id}: no part ids -> consumed=False")
            print(f"[RECV_DETAIL consume-check] batch {batch_id}: no part ids -> consumed=False")
            return False

        # (4) проверяем, есть ли выдачи по этим Part.id
        issued_rows = (
            db.session.query(
                IssuedPartRecord.id,
                IssuedPartRecord.part_id,
                IssuedPartRecord.quantity,
                IssuedPartRecord.issue_date
            )
            .filter(IssuedPartRecord.part_id.in_(all_part_ids))
            .limit(5)
            .all()
        )

        log.warning(f"[RECV_DETAIL consume-check] batch {batch_id}: issued_rows={issued_rows}")
        print(f"[RECV_DETAIL consume-check] batch {batch_id}: issued_rows={issued_rows}")

        return len(issued_rows) > 0

    consumed = _batch_has_been_consumed(batch)

    # ---------- role / permissions ----------
    role_low = (getattr(current_user, "role", "") or "").lower()
    is_super = (
        role_low in ["admin", "superadmin"]
        or getattr(current_user, "is_admin", False)
        or getattr(current_user, "is_superadmin", False)
        or getattr(current_user, "is_super_admin", False)
    )

    status_low = (getattr(batch, "status", "") or "").strip().lower()
    is_posted = (status_low == "posted")

    # can_edit: супер + не consumed
    can_edit = bool(is_super and (not consumed))

    # can_unpost: супер + posted + не consumed
    can_unpost = bool(is_super and is_posted and (not consumed))

    log.warning(
        f"[RECV_DETAIL FLAGS] batch {batch_id} "
        f"status={batch.status} posted?={is_posted} consumed?={consumed} "
        f"is_super={is_super} => can_edit={can_edit} can_unpost={can_unpost}"
    )
    print(
        f"[RECV_DETAIL FLAGS] batch {batch_id} "
        f"status={batch.status} posted?={is_posted} consumed?={consumed} "
        f"is_super={is_super} => can_edit={can_edit} can_unpost={can_unpost}"
    )

    return render_template(
        "receiving_detail.html",
        batch=batch,
        lines=lines,
        batch_total=batch_total,

        readonly=(not can_edit),
        can_edit=can_edit,
        can_unpost=can_unpost,
        consumed=consumed,
    )

@inventory_bp.post("/receiving/<int:batch_id>/post", endpoint="receiving_post")
@login_required
def receiving_post(batch_id):
    from extensions import db
    from models import ReceivingBatch

    batch = db.session.get(ReceivingBatch, batch_id)
    if not batch:
        flash("Batch not found", "danger")
        return redirect(url_for("inventory.receiving_list"))

    status_now = (batch.status or "").strip().lower()

    # КРИТИЧЕСКИЙ ФИЛЬТР:
    # если партию уже когда-то фактически оприходовали (твой "призрак"),
    # то повторно склад трогать нельзя.
    #
    # Простая защита: если status уже 'posted' -> не зовём post_receiving_batch
    # (это уже было), НО также если status == 'draft', НО batch.posted_at уже не пустой,
    # значит её уже проводили и потом кто-то откатил только поле status.
    #
    already_posted_once = False
    if status_now == "posted":
        already_posted_once = True
    else:
        # иногда баг откатывает только поле status, но оставляет posted_at
        if getattr(batch, "posted_at", None):
            already_posted_once = True

    if already_posted_once:
        flash("Batch already applied to stock. Status updated to POSTED.", "info")
        # просто выставим статус в БД, без повторного прихода
        batch.status = "posted"
        db.session.commit()
        return redirect(url_for("inventory.receiving_detail", batch_id=batch_id))

    # нормальный путь: реально ещё не приходили -> делаем приход
    post_receiving_batch(batch_id, getattr(current_user, "id", None))
    flash("Batch posted", "success")
    return redirect(url_for("inventory.receiving_detail", batch_id=batch_id))

# __________________________________________________________________________________________________________
def create_receiving_from_rows(
    *,
    supplier_name: str,
    invoice_number: str | None,
    invoice_date: date | None,
    currency: str | None,
    notes: str | None,
    rows: list[dict],
    created_by=None,
    auto_post: bool = False
) -> ReceivingBatch:

    """
    Делает так:
      1. создаёт ReceivingBatch как draft + строки
      2. commit (чтоб был id)
      3. если auto_post == True:
            - вызывает post_receiving_batch(batch.id)
              (это плюсанёт склад и поставит status='posted')
            - перечитывает batch из базы (уже posted)
            - expunge() эту свежую версию, чтобы статус 'posted' не затёрся обратно
         и возвращает именно ЭТУ версию
      4. если auto_post == False:
            - просто возвращает draft, тоже expunge, чтоб не затирал потом
    """

    # 1. СОЗДАТЬ batch как draft
    batch = ReceivingBatch(
        supplier_name=(supplier_name or "").strip(),
        invoice_number=(invoice_number or "").strip() or None,
        invoice_date=invoice_date,
        currency=(currency or "USD").strip()[:8],
        notes=(notes or "").strip() or None,
        status="draft",
        created_at=datetime.utcnow(),
        created_by=created_by,
    )
    db.session.add(batch)
    db.session.flush()  # batch.id теперь есть

    # 1.5. СОЗДАТЬ строки
    line_no = 1
    for r in rows:
        pn = (r.get("part_number") or r.get("pn") or "").strip()
        if not pn:
            continue

        qty = int((r.get("quantity") or r.get("qty") or 0) or 0)
        if qty <= 0:
            continue

        line_kwargs = dict(
            line_no=line_no,
            part_number=pn,
            part_name=(r.get("part_name") or r.get("description") or r.get("descr") or "").strip() or None,
            quantity=qty,
            unit_cost=float((r.get("unit_cost") or r.get("price") or 0) or 0),
            location=(r.get("location") or r.get("supplier") or "").strip() or None,
        )

        # ВАЖНО: у тебя ReceivingItem сейчас создаётся с batch_id=gr.id
        # если у твоей модели строк реально поле называется goods_receipt_id,
        # то замени здесь на goods_receipt_id=batch.id
        line_kwargs["batch_id"] = batch.id

        line = ReceivingItem(**line_kwargs)
        db.session.add(line)

        line_no += 1

    # 2. commit draft+строки, теперь batch и items в базе
    db.session.commit()

    # 3. если НЕ нужно автопостить → просто вернуть draft (но вынести из сессии)
    if not auto_post:
        fresh_draft = db.session.get(ReceivingBatch, batch.id)
        try:
            from flask import current_app
            current_app.logger.debug(
                "### DEBUG create_receiving_from_rows draft return id=%s status=%s",
                fresh_draft.id,
                fresh_draft.status,
            )
        except Exception:
            pass

        db.session.expunge(fresh_draft)
        return fresh_draft

    # 4. auto_post == True → проводим на склад и ставим posted
    post_receiving_batch(batch.id, current_user_id=created_by)

    # перечитываем уже ПОСТНУТУЮ версию
    posted_batch = db.session.get(ReceivingBatch, batch.id)

    try:
        from flask import current_app
        current_app.logger.debug(
            "### DEBUG create_receiving_from_rows AFTER POST id=%s status=%s",
            posted_batch.id,
            posted_batch.status,
        )
    except Exception:
        pass

    # КРИТИЧЕСКИЙ МОМЕНТ:
    # вынимаем объект из сессии, чтобы никакой следующий commit в этом же request
    # не смог случайно залить обратно старый статус 'draft'
    db.session.expunge(posted_batch)

    try:
        from flask import current_app
        current_app.logger.debug(
            "### DEBUG CFR (create_receiving_from_rows) RETURNING batch_id=%s status=%s auto_post=%s file=%s",
            gr.id,
            getattr(gr, "status", None),
            auto_post,
            __file__,
        )
    except Exception:
        pass

    return posted_batch


def _norm_cols(df):
    """
    Нормализация колонок DF к именам: part_number, part_name, quantity, unit_cost, supplier, location.
    + Фикс: разлепить 'PN DESCRIPTION' -> PN и DESCRIPTION по первому пробелу.
    + Отсев: 'Shipment marking/Shipmentmarking' и итоговые строки (Order total/Subtotal/...).
    """
    import re
    cols = {c.lower().strip(): c for c in df.columns}

    def pick(*names):
        for n in names:
            for key, orig in cols.items():
                if key == n.lower():
                    return orig
        # «умные» совпадения
        for key, orig in cols.items():
            if any(k in key for k in names):
                return orig
        return None

    c_pn  = pick("PART #","PART#","PART","PN","part_number")
    c_nm  = pick("DESCR.","DESCRIPTION","DESC","NAME","part_name")
    c_qty = pick("QTY","QUANTITY","qty ordered","qty shipped","quantity")
    c_uc  = pick("UNIT COST","UNIT PRICE","COST","PRICE","unit_cost")
    c_sup = pick("SUPPLIER","VENDOR","FROM","supplier")
    c_loc = pick("LOCATION","BIN","SHELF","location")

    # --- Регэкспы и фильтры ---
    marker_rx = re.compile(r'^\s*shipment\s*mark(?:ing)?\s*[:\-]?', re.I)
    totals_rx = re.compile(r'^\s*(order\s*total|orderinetot|ordertot|subtotal|total)\b', re.I)
    # допустимый токен PN: буквы/цифры и часто используемые символы ('-','/','.', '_')
    pn_token_rx = re.compile(r'^[A-Z0-9][A-Z0-9\-\/\._]*$', re.I)

    def split_pn_desc(text: str):
        """Разделить 'PN rest...' -> (PN, 'rest...') если первый токен похож на PN."""
        s = (text or "").strip()
        if not s:
            return "", ""
        parts = s.split()
        if len(parts) <= 1:
            return s, ""
        first, rest = parts[0], " ".join(parts[1:])
        if pn_token_rx.match(first):
            return first, rest
        return "", s  # первый токен не похож на PN — считаем всё описанием

    out = []
    for _, row in df.iterrows():
        raw_pn = str(row.get(c_pn, "")).strip() if c_pn else ""
        raw_nm = str(row.get(c_nm, "")).strip() if c_nm else ""

        # 1) Отсев мусора: Shipment marking, Order total и т.п.
        if (raw_pn and marker_rx.match(raw_pn)) or (raw_nm and marker_rx.match(raw_nm)):
            continue
        if totals_rx.match(raw_pn) or totals_rx.match(raw_nm):
            continue

        # 2) Приведение PN/NAME
        pn, nm = raw_pn, raw_nm

        # a) случай: PN склеен с описанием в самом PN (есть пробел)
        if pn and " " in pn:
            p_guess, rest = split_pn_desc(pn)
            if p_guess:
                pn = p_guess
                # если описания не было — берём остаток; если было — приписываем справа
                nm = (nm or "").strip()
                nm = (rest if not nm else f"{nm} {rest}").strip()

        # b) случай: PN пуст, а в DESCRIPTION лежит "PN rest..."
        if not pn and nm:
            p_guess, rest = split_pn_desc(nm)
            if p_guess:
                pn = p_guess
                nm = rest

        # c) иногда обе колонки одинаковые (дубликат) — очистим описание
        if nm and pn and nm.strip().upper() == pn.strip().upper():
            nm = ""

        # 3) Числа
        qty_raw = str(row.get(c_qty, "")).replace(",", "") if c_qty else ""
        uc_raw  = str(row.get(c_uc, "")).replace("$", "").replace(",", "") if c_uc else ""
        try:
            qty = int(float(qty_raw or "0"))
        except Exception:
            qty = 0
        try:
            uc = float(uc_raw or "0")
        except Exception:
            uc = 0.0

        # 4) Пустые строки выкидываем
        if not pn and not nm and qty == 0 and uc == 0.0:
            continue

        out.append({
            "part_number": pn,
            "part_name": nm,
            "quantity": qty,
            "unit_cost": uc,
            "supplier": (str(row.get(c_sup, "")).strip() if c_sup else ""),
            "location": (str(row.get(c_loc, "")).strip() if c_loc else ""),
        })
    return out

@inventory_bp.get("/receiving/import")
@login_required
def receiving_import_form():
    """Форма загрузки PDF (или CSV, если захочешь)."""
    return render_template("receiving_import_upload.html")

@inventory_bp.route("/receiving/import", methods=["GET", "POST"])
@login_required
def receiving_import_upload():
    # legacy path disabled, use /import-parts instead
    flash("Use Import Parts page instead.", "warning")
    return redirect(url_for("inventory.import_parts_upload"))

@inventory_bp.get("/receiving/<int:batch_id>/attachment")
@login_required
def download_receiving_attachment(batch_id):
    import os
    from flask import send_file, abort
    from models import GoodsReceipt

    gr = GoodsReceipt.query.get_or_404(batch_id)
    p = (gr.attachment_path or "").strip()
    if not p or not os.path.exists(p):
        abort(404)
    return send_file(
        p,
        mimetype="application/pdf",
        as_attachment=False,
        download_name=os.path.basename(p)
    )

@inventory_bp.get("/add-batch", endpoint="add_part_batch_form")
@login_required
def add_part_batch_form():
    """
    Scan / Add Multiple Parts UI.
    Показываем страницу с полем сканера и таблицей batchItems (JS).
    """
    # доступы можно ограничить если хочешь:
    # if (current_user.role or '').lower() not in ('admin','superadmin'):
    #     flash("Access denied", "danger")
    #     return redirect(url_for("inventory.dashboard"))

    return render_template("add_part.html")

@inventory_bp.post("/add-batch", endpoint="add_part_batch")
@login_required
def add_part_batch():
    role_low = (getattr(current_user, "role", "") or "").lower()
    if role_low not in ("admin", "superadmin"):
        flash("Access denied", "danger")
        return redirect(url_for("inventory.dashboard"))

    # --- надёжное чтение payload ---
    payload_raw = (request.form.get("batch_payload") or "").strip()
    if not payload_raw:
        current_app.logger.warning("Empty batch_payload in POST form")
        flash("Invalid batch data: empty payload", "danger")
        return redirect(url_for("inventory.dashboard"))

    try:
        items = json.loads(payload_raw)
    except Exception:
        current_app.logger.exception("Bad batch_payload JSON: %r", payload_raw[:500])
        flash("Invalid batch data: JSON parse error", "danger")
        return redirect(url_for("inventory.dashboard"))

    if not isinstance(items, list) or not items:
        flash("Invalid batch data: expected non-empty list", "danger")
        return redirect(url_for("inventory.dashboard"))

    # ---------- 1. создаём batch (draft) ----------
    supplier_hint = "ADD-BATCH"
    invoice_hint  = datetime.utcnow().strftime("BULK-%Y%m%d-%H%M%S")

    batch = ReceivingBatch(
        supplier_name = supplier_hint,
        invoice_number = invoice_hint,
        invoice_date = datetime.utcnow().date(),
        currency = "USD",
        notes = "Created via /add-batch",
        status = "draft",
        created_at = datetime.utcnow(),
        created_by = getattr(current_user, "id", None),
    )
    db.session.add(batch)
    db.session.flush()  # batch.id

    # ---------- 2. строки ----------
    line_no = 1
    for row in items:
        pn  = (row.get("part_number") or "").strip().upper()
        nm  = (row.get("name") or "").strip()
        loc = (row.get("loc") or "").strip().upper()
        try:
            qty = int(row.get("qty") or 0)
        except Exception:
            qty = 0
        try:
            unit_cost_val = float(row.get("cost")) if row.get("cost") not in (None, "") else 0.0
        except Exception:
            unit_cost_val = 0.0

        if not pn or qty <= 0:
            current_app.logger.warning("Skip invalid row in add-batch: %r", row)
            continue

        line = ReceivingItem(
            goods_receipt_id = batch.id,   # важно: новое имя FK
            line_no          = line_no,
            part_number      = pn,
            part_name        = nm or None,
            quantity         = qty,
            unit_cost        = unit_cost_val,
            location         = loc or None,
        )
        db.session.add(line)
        line_no += 1

    # ---------- 3. commit draft ----------
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception("Failed to commit draft receiving batch")
        flash(f"Failed to save new batch: {e}", "danger")
        return redirect(url_for("inventory.dashboard"))

    # ---------- 4. post (обновление склада) ----------
    from services.receiving import post_receiving_batch
    try:
        post_receiving_batch(batch.id, getattr(current_user, "id", None))
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception("Failed to post receiving batch")
        flash(f"Batch saved as draft but failed to post stock: {e}", "warning")
        return redirect(url_for("inventory.receiving_detail", batch_id=batch.id))

    flash(f"Stock received and posted. Batch #{batch.id}", "success")
    return redirect(url_for("inventory.receiving_detail", batch_id=batch.id))


















